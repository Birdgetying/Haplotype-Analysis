#!/usr/bin/env python3
"""Tests for lightweight star-gene data source metadata."""

from pathlib import Path
from contextlib import redirect_stdout
from io import StringIO
import csv
import gzip
import json
import shutil
import subprocess
import tempfile
import unittest
import pandas as pd


class StarGeneDataTests(unittest.TestCase):
    def test_display_position_pairs_keep_one_rendered_column_per_coordinate(self):
        from haplotype_phenotype_analysis import _deduplicate_display_position_pairs

        positions, sequence_indices = _deduplicate_display_position_pairs(
            [10, 10, 20, 20, 30],
            [2, 7, 8, 11, 12],
        )

        self.assertEqual(positions, [10, 20, 30])
        self.assertEqual(sequence_indices, [2, 8, 12])

    def test_initial_display_range_uses_full_range_for_small_variant_sets(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range([10, 20, 30], 1, 100, 20, 80),
            (None, None),
        )

    def test_initial_display_range_caps_midpoint_window_at_25_variants(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        positions = list(range(100, 4100, 100))
        start, end = _select_initial_display_range(
            positions,
            1,
            5000,
            gene_start=1500,
            gene_end=2500,
            max_variants=25,
            max_span_bp=5000,
        )

        selected = [pos for pos in positions if start <= pos <= end]
        self.assertEqual(len(selected), 25)
        self.assertLessEqual(start, 2000)
        self.assertGreaterEqual(end, 2000)

    def test_initial_display_range_centers_by_sorted_index_across_large_gaps(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        positions = list(range(1, 21)) + list(range(1000, 1020))
        start, end = _select_initial_display_range(
            positions,
            1,
            2000,
            gene_start=999,
            gene_end=1001,
            max_variants=25,
            max_span_bp=2000,
        )

        self.assertEqual((start, end), (9, 1012))

    def test_initial_display_range_falls_back_to_compact_cluster(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        positions = list(range(100, 4100, 100))
        start, end = _select_initial_display_range(
            positions,
            1,
            5000,
            gene_start=1500,
            gene_end=2500,
            max_variants=25,
            max_span_bp=2000,
        )

        selected = [pos for pos in positions if start <= pos <= end]
        self.assertEqual((start, end), (1000, 3000))
        self.assertEqual(len(selected), 21)
        self.assertLessEqual(end - start, 2000)

    def test_initial_display_range_uses_single_sparse_variant_near_gene(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range(
                [100, 5000, 9000],
                1,
                10000,
                gene_start=4000,
                gene_end=6000,
                max_variants=25,
                max_span_bp=2000,
            ),
            (5000, 5000),
        )

    def test_initial_display_range_uses_gene_range_without_variants(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range(
                [],
                1,
                10000,
                gene_start=4000,
                gene_end=6000,
                max_variants=25,
                max_span_bp=2000,
            ),
            (4000, 6000),
        )

    def test_initial_display_range_keeps_vrn_b1_like_cluster_compact(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        positions = (
            list(range(5447, 5459))
            + [7596, 7600, 7603, 7605]
            + list(range(12379, 12429))
        )
        start, end = _select_initial_display_range(
            positions,
            1,
            17867,
            gene_start=4673,
            gene_end=17867,
            max_variants=25,
            max_span_bp=2000,
        )

        selected = [pos for pos in positions if start <= pos <= end]
        self.assertEqual((start, end), (12379, 12403))
        self.assertEqual(len(selected), 25)
        self.assertLessEqual(end - start, 2000)

    def test_compact_range_prefers_midpoint_nearest_anchor_for_equal_counts(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range(
                [10, 60, 90],
                0,
                100,
                gene_start=55,
                gene_end=55,
                max_variants=25,
                max_span_bp=1,
            ),
            (60, 60),
        )

    def test_compact_range_prefers_shorter_span_after_count_and_anchor_tie(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range(
                [10, 14, 24, 26],
                0,
                40,
                gene_start=18,
                gene_end=19,
                max_variants=25,
                max_span_bp=4,
            ),
            (24, 26),
        )

    def test_compact_range_prefers_lower_start_after_first_three_score_ties(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range(
                [10, 12, 18, 20],
                0,
                30,
                gene_start=15,
                gene_end=15,
                max_variants=25,
                max_span_bp=2,
            ),
            (10, 12),
        )

    def test_compact_range_counts_duplicate_positions_as_rendered_columns(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range(
                [10, 10, 20, 20, 20],
                0,
                30,
                gene_start=15,
                gene_end=15,
                max_variants=5,
                max_span_bp=1,
            ),
            (20, 20),
        )

    def test_compact_range_preserves_none_for_region_boundaries(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        for anchor, expected in ((10, (None, 10)), (100, (100, None))):
            with self.subTest(anchor=anchor):
                self.assertEqual(
                    _select_initial_display_range(
                        [10, 100],
                        10,
                        100,
                        gene_start=anchor,
                        gene_end=anchor,
                        max_variants=25,
                        max_span_bp=1,
                    ),
                    expected,
                )

    def test_initial_display_range_invalid_span_falls_back_to_2000(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        for invalid_span in (float("inf"), float("-inf"), float("nan")):
            with self.subTest(max_span_bp=invalid_span):
                try:
                    actual = _select_initial_display_range(
                        [],
                        1,
                        10000,
                        gene_start=4000,
                        gene_end=6000,
                        max_variants=25,
                        max_span_bp=invalid_span,
                    )
                except OverflowError as exc:
                    self.fail(f"invalid max_span_bp raised OverflowError: {exc}")
                self.assertEqual(actual, (4000, 6000))

    def test_initial_display_range_invalid_variant_limit_falls_back_to_25(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        actual = _select_initial_display_range(
            list(range(1, 41)),
            1,
            40,
            gene_start=20,
            gene_end=20,
            max_variants=float("inf"),
        )

        start, end = actual
        self.assertEqual(len([p for p in range(1, 41) if start <= p <= end]), 25)

    def test_initial_display_range_rejects_infinite_region_coordinates(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range([10, 20], float("inf"), 30),
            (None, None),
        )

    def test_initial_display_range_ignores_infinite_variant_coordinates(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range(
                [10, float("inf"), 20],
                0,
                30,
                gene_start=15,
                gene_end=15,
            ),
            (None, None),
        )

    def test_initial_display_range_uses_region_anchor_for_infinite_gene_coordinates(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        self.assertEqual(
            _select_initial_display_range(
                [10, 20],
                0,
                30,
                gene_start=float("inf"),
                gene_end=15,
                max_span_bp=1,
            ),
            (10, 10),
        )

    def test_initial_display_range_uses_region_midpoint_without_gene_coordinates(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        positions = list(range(1, 41))
        start, end = _select_initial_display_range(
            positions,
            1,
            40,
            gene_start=None,
            gene_end=None,
            max_variants=25,
        )

        self.assertEqual(len([p for p in positions if start <= p <= end]), 25)
        self.assertLessEqual(start, 20)
        self.assertGreaterEqual(end, 20)

    def test_initial_display_range_clamps_window_at_variant_boundaries(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        positions = list(range(1, 41))
        start, end = _select_initial_display_range(
            positions,
            1,
            40,
            gene_start=1,
            gene_end=1,
            max_variants=25,
        )

        self.assertEqual((start, end), (None, 25))

    def test_initial_display_range_ignores_invalid_and_duplicate_positions(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        positions = list(range(1, 41)) + [None, "bad", 20, 999]
        start, end = _select_initial_display_range(
            positions,
            1,
            40,
            gene_start="bad",
            gene_end=30,
            max_variants=25,
        )

        lower = 1 if start is None else start
        upper = 40 if end is None else end
        selected = [
            int(pos)
            for pos in positions
            if str(pos).isdigit() and 1 <= int(pos) <= 40 and lower <= int(pos) <= upper
        ]
        self.assertLessEqual(len(selected), 25)
        self.assertLessEqual(lower, 20)
        self.assertGreaterEqual(upper, 20)

    def test_initial_display_range_caps_rendered_columns_with_duplicate_positions(self):
        from haplotype_phenotype_analysis import _select_initial_display_range

        positions = list(range(1, 27)) + [13] * 10
        start, end = _select_initial_display_range(
            positions,
            1,
            26,
            gene_start=1,
            gene_end=26,
            max_variants=25,
        )
        lower = 1 if start is None else start
        upper = 26 if end is None else end
        rendered_columns = [pos for pos in positions if lower <= pos <= upper]

        self.assertLessEqual(len(rendered_columns), 25)
        self.assertLessEqual(lower, 13)
        self.assertGreaterEqual(upper, 13)

    def test_frontiers2022_vrn_b1_builds_structural_marker_tables(self):
        from prepare_wheat2024_vrn_b1_frontiers2022 import build_frontiers2022_marker_tables

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "frontiers_vrn1.xlsx"
            with pd.ExcelWriter(workbook) as writer:
                pd.DataFrame([
                    ["title", None, None, None],
                    ["Genotype ID", "Cultivar name", "Heading date", "Plant height [cm]"],
                    ["TG001", "WinterA", 154.0, 90.0],
                    ["TG002", "Highbury", 150.0, 82.0],
                    ["TG003", "Joss", 151.0, 95.0],
                ]).to_excel(writer, sheet_name="S1", header=False, index=False)
                pd.DataFrame([
                    ["title", None, None, None, None, None, None, None],
                    [None, None, "VRN-A1", "VRN-B1", None, None, "VRN-D1", None],
                    ["NO", "Genotype name", "Insertion(231bp)", "Duplication(838bp)",
                     "Deletion(6851bp)", "Deletion(37bp)", "Insertion(163bp)", "Deletion(17bp)"],
                    [1, "Highbury", "Vrn-A1a", "-", "Vrn-B1a", "-", "-", "-"],
                    [2, "Joss", "-", "Vrn-B1f", "-", "-", "-", "-"],
                ]).to_excel(writer, sheet_name="S12", header=False, index=False)
                pd.DataFrame([
                    ["title", None, None, None, None, None],
                    ["Cultivar name", "VRN-A1 sequence type", "Genotype of VRN-A1",
                     "Haplotype of VRN-A1", "Haplotype of VRN-B1", "Haplotype of VRN-D1"],
                    ["WinterA", "Weebil", "GT4", "Hap3", "Hap1", "Hap1"],
                    ["Highbury", "Weebil", "GT4", "Hap3", "Hap5", "Hap1"],
                    ["Joss", "Weebil", "GT4", "Hap3", "Hap2", "Hap3"],
                ]).to_excel(writer, sheet_name="S13", header=False, index=False)

            marker_df, phenotype_df = build_frontiers2022_marker_tables(workbook)
            single_marker_df, _ = build_frontiers2022_marker_tables(
                workbook,
                marker_columns=["VRN-B1_deletion_6851"],
            )

        self.assertEqual(
            ["SampleID", "VRN-B1_insertion_838_duplication", "VRN-B1_deletion_6851", "VRN-B1_deletion_37"],
            list(marker_df.columns),
        )
        highbury = marker_df.set_index("SampleID").loc["Highbury"]
        self.assertEqual("Vrn-B1a", highbury["VRN-B1_deletion_6851"])
        self.assertEqual("-", highbury["VRN-B1_insertion_838_duplication"])
        self.assertEqual(["SampleID", "Heading_date"], list(phenotype_df.columns))
        self.assertEqual(3, len(phenotype_df))
        self.assertEqual(["SampleID", "VRN-B1_deletion_6851"], list(single_marker_df.columns))

    def test_frontiers2022_vrn_b1_single_marker_target_is_opt_in(self):
        from prepare_wheat2024_vrn_b1_frontiers2022 import build_arg_parser

        parser = build_arg_parser()

        default_args = parser.parse_args([])
        opt_in_args = parser.parse_args(["--include-single-marker-target"])

        self.assertFalse(default_args.include_single_marker_target)
        self.assertTrue(opt_in_args.include_single_marker_target)

    def test_alignment_fasta_is_converted_to_base_and_indel_markers(self):
        from star_gene_data import build_marker_matrix_from_aligned_fasta

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            fasta_path = tmp_path / "aligned.fasta"
            fasta_path.write_text(
                ">S1\n"
                "ACGTAA\n"
                ">S2\n"
                "ATG--A\n"
                ">S3\n"
                "ATGTCA\n"
                ">S4\n"
                "ACGTAA\n",
                encoding="utf-8",
            )

            marker_df, marker_metadata = build_marker_matrix_from_aligned_fasta(
                fasta_path,
                marker_prefix="GeneA",
            )

        self.assertEqual(
            ["SampleID", "GeneA_snp_2", "GeneA_indel_4_5"],
            list(marker_df.columns),
        )
        self.assertEqual(["S1", "S2", "S3", "S4"], marker_df["SampleID"].tolist())
        self.assertEqual(["C", "T", "T", "C"], marker_df["GeneA_snp_2"].tolist())
        self.assertEqual(["TA", "DEL_2", "TC", "TA"], marker_df["GeneA_indel_4_5"].tolist())
        self.assertEqual(
            [
                {
                    "marker_id": "GeneA_snp_2",
                    "kind": "snp",
                    "alignment_start": 2,
                    "alignment_end": 2,
                    "length": 1,
                    "missing_rate": 0.0,
                },
                {
                    "marker_id": "GeneA_indel_4_5",
                    "kind": "indel",
                    "alignment_start": 4,
                    "alignment_end": 5,
                    "length": 2,
                    "missing_rate": 0.0,
                },
            ],
            marker_metadata,
        )

    def test_vrn_b1_full_sequence_maps_fasta_headers_to_cultivars(self):
        from prepare_wheat2024_vrn_b1_full_sequence import (
            build_sample_maps_from_esm2,
            make_vrn_b1_fasta_sample_id_fn,
        )

        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "esm2.xlsx"
            with pd.ExcelWriter(workbook) as writer:
                pd.DataFrame([
                    ["title", None, None, None],
                    ["note", None, None, None],
                    ["note", None, None, None],
                    [None, None, None, None],
                    ["Code", "Accession number", "Name", "Habit"],
                    [4, "ACC4", "Batis", "W"],
                    [71, "ACC71", "Apache", "W"],
                    ["S2", "ACCS2", "Anza", "S"],
                ]).to_excel(writer, sheet_name="Table S1", header=False, index=False)
                pd.DataFrame([
                    ["Table S5", None],
                    [None, None],
                    [34, "TDC"],
                    ["GROUP 1B", None],
                    [4, "Batis"],
                    [71, "Apache"],
                    ["GROUP 20B", None],
                    ["S2", "Anza"],
                ]).to_excel(writer, sheet_name="Table S5", header=False, index=False)

            code_to_sample, phenotype_df = build_sample_maps_from_esm2(workbook)
            sample_id_fn = make_vrn_b1_fasta_sample_id_fn(code_to_sample)

        self.assertEqual("Batis", sample_id_fn("4_Batis_2_NODE_1_length_16886_cov_77"))
        self.assertEqual("Apache", sample_id_fn("71_NODE_1_length_16886_cov_150"))
        self.assertEqual("Anza", sample_id_fn("S2B_merged_extraction 2 reads"))
        self.assertEqual(["SampleID", "GrowthHabitSpringScore"], list(phenotype_df.columns))
        self.assertEqual(1.0, phenotype_df.set_index("SampleID").loc["Anza", "GrowthHabitSpringScore"])

    def test_vrn_b1_full_sequence_merges_promoter_and_gene_markers(self):
        from prepare_wheat2024_vrn_b1_full_sequence import build_full_sequence_marker_matrix

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            promoter_fasta = tmp_path / "prom.fasta"
            gene_fasta = tmp_path / "gene.fasta"
            promoter_fasta.write_text(
                ">S1\n"
                "ACGT\n"
                ">S2\n"
                "ATGT\n"
                ">S3\n"
                "AC-T\n",
                encoding="utf-8",
            )
            gene_fasta.write_text(
                ">S1\n"
                "GGAA\n"
                ">S2\n"
                "GGTA\n"
                ">S3\n"
                "GGAA\n",
                encoding="utf-8",
            )

            marker_df, marker_metadata, layout = build_full_sequence_marker_matrix(
                promoter_fasta,
                gene_fasta,
                sample_id_fn=lambda header: header,
                max_missing_rate=0.5,
                min_minor_count=1,
            )

        self.assertEqual(["S1", "S2", "S3"], marker_df["SampleID"].tolist())
        self.assertIn("VRNB1prom_snp_2", marker_df.columns)
        self.assertIn("VRNB1gene_snp_7", marker_df.columns)
        self.assertEqual(1, layout["promoter_start"])
        self.assertEqual(4, layout["promoter_end"])
        self.assertEqual(5, layout["gene_start"])
        self.assertEqual(8, layout["gene_end"])
        metadata_by_id = {row["marker_id"]: row for row in marker_metadata}
        self.assertEqual("promoter", metadata_by_id["VRNB1prom_snp_2"]["segment"])
        self.assertEqual(2, metadata_by_id["VRNB1prom_snp_2"]["alignment_start"])
        self.assertEqual("gene", metadata_by_id["VRNB1gene_snp_7"]["segment"])
        self.assertEqual(7, metadata_by_id["VRNB1gene_snp_7"]["alignment_start"])

    def test_vrn_b1_full_sequence_adds_exact_vrnb1f_837_marker_from_diagnostic_interval(self):
        from prepare_wheat2024_vrn_b1_full_sequence import build_full_sequence_marker_matrix

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            promoter_fasta = tmp_path / "prom.fasta"
            gene_fasta = tmp_path / "gene.fasta"
            promoter_fasta.write_text(
                ">TDC\n"
                "AAAA\n"
                ">Anza\n"
                "AAAA\n"
                ">Barta\n"
                "AAAA\n",
                encoding="utf-8",
            )
            gene_fasta.write_text(
                ">TDC\n"
                "ACCATCTCCTTGCTTGCGAA-----CCGGTTGTCGTGTTCGTATCGTC\n"
                ">Anza\n"
                "ACCATCTCCTTGCTTGCGAAACGTTCCGGTTGTCGTGTTCGTATCGTC\n"
                ">Barta\n"
                "ACCATCTCCTTGCTTGCGAAACGTTCCGGTTGTCGTGTTCGTATCGTC\n",
                encoding="utf-8",
            )

            marker_df, marker_metadata, layout = build_full_sequence_marker_matrix(
                promoter_fasta,
                gene_fasta,
                sample_id_fn=lambda header: header,
                max_missing_rate=0.5,
                min_minor_count=1,
            )

        exact_markers = [
            row for row in marker_metadata
            if row.get("literature_variant") == "Vrn-B1f_837bp_insertion"
        ]
        self.assertEqual(1, len(exact_markers))
        exact = exact_markers[0]
        marker_id = exact["marker_id"]
        self.assertIn("insertion_837", marker_id)
        self.assertEqual("diagnostic_marker", exact["annotation"])
        self.assertEqual(True, exact["validation_marker"])
        self.assertEqual(layout["promoter_length"] + 21, exact["alignment_start"])
        self.assertEqual(layout["promoter_length"] + 25, exact["alignment_end"])
        self.assertEqual("DEL_5", marker_df.set_index("SampleID").loc["TDC", marker_id])
        self.assertEqual("ACGTT", marker_df.set_index("SampleID").loc["Anza", marker_id])

    def test_promoter_report_can_use_preloaded_full_promoter_coordinates(self):
        from haplotype_phenotype_analysis import PromoterAnnotator

        report = PromoterAnnotator().generate_promoter_report(
            gene_id="VRN-B1-fullSequence-IJMS2021",
            chrom="VRN-B1_IJMS2021_alignment",
            gene_start=4673,
            gene_end=17867,
            strand="+",
            variants_positions=[1, 3952, 4966],
            promoter_start=1,
            promoter_end=4672,
        )

        self.assertEqual(1, report["promoter_start"])
        self.assertEqual(4672, report["promoter_end"])
        self.assertEqual(2, report["variants_in_promoter"])
        self.assertEqual([1, 3952], report["variant_positions"])

    def test_vrn_b1_full_sequence_can_merge_continuous_heading_phenotypes(self):
        from prepare_wheat2024_vrn_b1_full_sequence import merge_continuous_phenotypes

        ijms_samples = pd.DataFrame({
            "SampleID": ["Apache", "Atlas 66", "Pannonia NS", "Unmatched"],
            "GrowthHabitSpringScore": [0.0, 0.0, 1.0, 1.0],
        })

        with tempfile.TemporaryDirectory() as tmp:
            phenotype_path = Path(tmp) / "kiss.tsv"
            phenotype_path.write_text(
                "SampleID\tDEV49_mean\tDEV59_mean\n"
                "APACHE\t210.5\t219.0\n"
                "ATLAS-66\t211.0\t220.5\n"
                "PANNONIA-NS\t205.0\t214.0\n"
                "OTHER\t200.0\t210.0\n",
                encoding="utf-8",
            )

            merged = merge_continuous_phenotypes(
                ijms_samples,
                phenotype_table=phenotype_path,
                sample_column="SampleID",
                phenotype_columns=["DEV49_mean", "DEV59_mean"],
            )

        self.assertEqual(["SampleID", "DEV49_mean", "DEV59_mean"], list(merged.columns))
        self.assertEqual(["Apache", "Atlas 66", "Pannonia NS"], merged["SampleID"].tolist())
        by_sample = merged.set_index("SampleID")
        self.assertEqual(211.0, by_sample.loc["Atlas 66", "DEV49_mean"])
        self.assertEqual(214.0, by_sample.loc["Pannonia NS", "DEV59_mean"])

    def test_vrn_remote_snp_prepare_can_build_heading_date_target(self):
        from openpyxl import Workbook
        from prepare_wheat2024_vrn_remote_snps import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            vcf_dir = tmp_path / "vcfs"
            vcf_dir.mkdir()
            vcf_path = vcf_dir / "VRN-B1.wheatomics_snp.vcf.gz"
            with gzip.open(vcf_path, "wt", encoding="utf-8") as f:
                f.write("\n".join([
                    "##fileformat=VCFv4.2",
                    "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tWATDE0001\tWATDE0002\tWATDE0003\tWATDE0004",
                    "chr5B\t573801000\tchr5B_573801000\tA\tG\t.\t.\t.\tGT\t0/0\t1/1\t0/0\t1/1",
                    "chr5B\t573802000\tchr5B_573802000\tC\tT\t.\t.\t.\tGT\t0/0\t0/0\t1/1\t1/1",
                    "",
                ]))

            phenotype_xlsx = tmp_path / "phenotypes.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "WGIN_Watkins_JIC_CFLN06"
            ws.append(["StoreCode", "Hd_dto_days-CFLN06", "PH_M_cm-CFLN06", "GrwHabit_E_sw-CFLN06"])
            ws.append(["WATDE0001", 84.0, 95.5, "ssss"])
            ws.append(["WATDE0002", 88.0, 90.0, "ssss"])
            ws.append(["WATDE0003", 92.0, 100.0, "wwww"])
            ws.append(["WATDE0004", 96.0, 102.0, "wwww"])
            wb.save(phenotype_xlsx)

            output_root = tmp_path / "db"
            intermediate_root = tmp_path / "intermediate"
            stdout = StringIO()
            with redirect_stdout(stdout):
                rc = prepare_main([
                    "--phenotype-xlsx", str(phenotype_xlsx),
                    "--vcf-dir", str(vcf_dir),
                    "--output-root", str(output_root),
                    "--intermediate-root", str(intermediate_root),
                    "--target", "VRN-B1-remoteSNP",
                    "--phenotype-mode", "heading_date",
                    "--min-haplotype-count", "1",
                ])

            self.assertEqual(rc, 0)
            db_dir = output_root / "VRN-B1-remoteSNP-HeadingDate"
            self.assertTrue((db_dir / "gene_info.json").exists())
            gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
            self.assertEqual(gene_info["source"], "wheatomics_remote_vrn_snp_vcf_heading_date")
            self.assertEqual(gene_info["expected_direction"], "decreases_trait")
            phenotype_data = (db_dir / "phenotype_data.csv").read_text(encoding="utf-8")
            self.assertIn("HeadingDate_CFLN06", phenotype_data)
            self.assertNotIn("GrowthHabitSpringScore_CFLN06", phenotype_data)
            self.assertIn("WATDE0001", phenotype_data)

    def test_rice_figshare_is_large_and_skipped_by_default(self):
        from star_gene_data import build_download_commands, iter_data_files

        rice_files = list(iter_data_files(paper="rice2024"))
        rice_figshare = [f for f in rice_files if f.key == "rice2024_figshare_genotype_matrix"][0]

        self.assertTrue(rice_figshare.is_large)
        self.assertEqual(rice_figshare.default_action, "manual_large_download")
        self.assertIn("10.6084/m9.figshare.19166475", rice_figshare.source)
        self.assertEqual(build_download_commands(paper="rice2024", include_large=False), [])

    def test_maize_small_files_have_direct_download_commands(self):
        from star_gene_data import build_download_commands, iter_data_files

        maize_files = list(iter_data_files(paper="maize2019"))
        keys = {f.key for f in maize_files}

        self.assertIn("maize2019_sv_382254", keys)
        self.assertIn("maize2019_agronomic_blup", keys)

        commands = build_download_commands(paper="maize2019")
        command_text = "\n".join(commands)
        self.assertIn("SV.382254.zip", command_text)
        self.assertIn("blup_traits_final.csv", command_text)
        self.assertIn("external_data/maize_natgenet_2019/maizego", command_text)

    def test_wheat_sources_are_instruction_only_until_object_names_are_known(self):
        from star_gene_data import iter_data_files

        wheat_files = list(iter_data_files(paper="wheat2024"))
        self.assertTrue(wheat_files)
        portal_files = [f for f in wheat_files if f.key in {"wheat2024_wwwg2b_portal", "wheat2024_earlham_watseq"}]
        self.assertEqual(len(portal_files), 2)
        self.assertTrue(all(f.default_action == "instruction_only" for f in portal_files))
        self.assertTrue(any("wwwg2b.com" in f.source for f in wheat_files))

    def test_wheat_q7b_small_files_have_wwwg2b_download_commands(self):
        from star_gene_data import build_download_commands, iter_data_files

        wheat_files = list(iter_data_files(paper="wheat2024"))
        keys = {f.key for f in wheat_files}

        self.assertIn("wheat2024_q7b_ph_figure3g", keys)
        self.assertIn("wheat2024_watkins_jic_phenotypes", keys)

        commands = build_download_commands(paper="wheat2024")
        command_text = "\n".join(commands)
        self.assertIn("get_download_url_form_onedrive", command_text)
        self.assertIn("Watseq_Figure_3g_NIL_Q7B-PH_field_data.xlsx", command_text)
        self.assertIn("external_data/wheat_nature_2024/wwwg2b/q7b_ph", command_text)
        self.assertIn("11032_2014_34_MOESM1_ESM.doc", command_text)
        self.assertIn("static-content.springer.com", command_text)
        self.assertIn("curl.exe -L -C -", command_text)

    def test_summarize_downloads_contains_status_and_target_paths(self):
        from star_gene_data import summarize_downloads

        summary = summarize_downloads(paper="maize2019")
        expected_path = str(Path("external_data/maize_natgenet_2019/maizego/SV.382254.zip")).replace("\\", "/")
        self.assertIn("maize2019_sv_382254", summary)
        self.assertIn("small_direct_download", summary)
        self.assertIn(expected_path, summary)

    def test_unknown_paper_filter_raises_clear_error(self):
        from star_gene_data import iter_data_files

        with self.assertRaisesRegex(ValueError, "No data files matched"):
            list(iter_data_files(paper="not_a_paper"))

    def test_star_gene_validator_prefers_exact_target_id_over_alias_matches(self):
        from star_gene_validation import StarGeneValidator

        manifest = {
            "papers": [
                {
                    "paper_id": "wheat2024",
                    "short_name": "wheat",
                    "targets": [
                        {
                            "target_id": "Rht-D1b",
                            "gene_or_locus": "Rht-D1",
                            "aliases": ["TraesCS4D02G040400"],
                        },
                        {
                            "target_id": "Rht-Zanke2014",
                            "gene_or_locus": "Rht1",
                            "aliases": ["Rht-B1", "Rht-D1", "Rht-D1b"],
                        },
                    ],
                }
            ]
        }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            validator = StarGeneValidator(
                manifest=manifest,
                database_root=tmp_path / "db",
                results_root=tmp_path / "results",
            )

            selected = [
                target["target_id"]
                for _, target in validator.iter_targets(papers=["wheat2024"], targets=["Rht-D1b"])
            ]

        self.assertEqual(selected, ["Rht-D1b"])

    def test_cli_print_downloads_exits_before_validation(self):
        from star_gene_validation import main

        stdout = StringIO()
        with redirect_stdout(stdout):
            rc = main(["--print-downloads", "--paper", "maize2019"])
        self.assertEqual(rc, 0)
        self.assertIn("maize2019_sv_382254", stdout.getvalue())

    def test_cli_accepts_robust_discovery_score_mode(self):
        from star_gene_validation import build_arg_parser

        args = build_arg_parser().parse_args([
            "--check-only",
            "--paper",
            "wheat2024",
            "--score-mode",
            "robust_discovery",
        ])

        self.assertEqual(args.score_mode, "robust_discovery")

    def test_direction_aware_top_haplotype_uses_expected_decrease(self):
        from star_gene_validation import StarGeneValidator

        runner = StarGeneValidator(
            manifest={"papers": []},
            database_root=Path("star_gene_database"),
            results_root=Path("star_gene_results"),
            check_only=True,
        )
        score_results = {
            "per_haplotype": {
                "TallHap": {"total": 2.0, "mean_phenotype": 100.0},
                "ShortHap": {"total": 1.0, "mean_phenotype": 80.0},
                "TinyShortHap": {"total": 3.0, "mean_phenotype": 70.0},
            },
            "per_sample": [
                {"sample_id": "S1", "haplotype": "TallHap", "score": 2.0, "phenotype": 100.0},
                {"sample_id": "S2", "haplotype": "TallHap", "score": 2.0, "phenotype": 102.0},
                {"sample_id": "S3", "haplotype": "ShortHap", "score": 1.0, "phenotype": 80.0},
                {"sample_id": "S4", "haplotype": "ShortHap", "score": 1.0, "phenotype": 78.0},
                {"sample_id": "S5", "haplotype": "TinyShortHap", "score": 3.0, "phenotype": 70.0},
            ],
        }

        self.assertEqual(runner.pick_directional_top_haplotype(score_results, "increases_trait")[0], "TallHap")
        self.assertEqual(runner.pick_directional_top_haplotype(score_results, "decreases_trait")[0], "ShortHap")

    def test_direction_aware_top_haplotype_filters_low_support_extremes(self):
        from star_gene_validation import StarGeneValidator

        runner = StarGeneValidator(
            manifest={"papers": []},
            database_root=Path("star_gene_database"),
            results_root=Path("star_gene_results"),
            check_only=True,
        )
        per_haplotype = {
            "StableHigh": {"total": 1.8, "mean_phenotype": 50.0},
            "TinyExtreme": {"total": 3.0, "mean_phenotype": 100.0},
        }
        per_sample = []
        for i in range(8):
            per_sample.append({
                "sample_id": f"Stable{i}",
                "haplotype": "StableHigh",
                "score": 1.8,
                "phenotype": 49.0 + i % 3,
            })
        for i in range(2):
            per_sample.append({
                "sample_id": f"Tiny{i}",
                "haplotype": "TinyExtreme",
                "score": 3.0,
                "phenotype": 100.0 + i,
            })
        for i in range(30):
            hap = f"Singleton{i}"
            per_haplotype[hap] = {"total": 0.1, "mean_phenotype": 20.0 + i * 0.01}
            per_sample.append({
                "sample_id": f"SingletonSample{i}",
                "haplotype": hap,
                "score": 0.1,
                "phenotype": 20.0 + i * 0.01,
            })

        score_results = {
            "per_haplotype": per_haplotype,
            "per_sample": per_sample,
        }

        self.assertEqual(
            runner.pick_directional_top_haplotype(score_results, "increases_trait")[0],
            "StableHigh",
        )

    def test_direction_aware_top_core_group_uses_expected_direction(self):
        from star_gene_validation import StarGeneValidator

        runner = StarGeneValidator(
            manifest={"papers": []},
            database_root=Path("star_gene_database"),
            results_root=Path("star_gene_results"),
            check_only=True,
        )
        score_results = {
            "core_haplotype_groups": {
                "groups": {
                    "A|A": {
                        "core_sequence": "A|A",
                        "rank_score": 1.0,
                        "mean_score": 1.2,
                        "mean_phenotype": 30.0,
                        "sample_count": 50,
                    },
                    "G|G": {
                        "core_sequence": "G|G",
                        "rank_score": 0.8,
                        "mean_score": 0.9,
                        "mean_phenotype": 45.0,
                        "sample_count": 20,
                    },
                    "T|T": {
                        "core_sequence": "T|T",
                        "rank_score": 2.0,
                        "mean_score": 2.5,
                        "mean_phenotype": 60.0,
                        "sample_count": 1,
                    },
                }
            }
        }

        self.assertEqual(
            runner.pick_directional_top_core_group(score_results, "increases_trait")[0],
            "G|G",
        )
        self.assertEqual(
            runner.pick_directional_top_core_group(score_results, "decreases_trait")[0],
            "A|A",
        )

    def test_direction_aware_top_functional_group_uses_expected_direction(self):
        from star_gene_validation import StarGeneValidator

        runner = StarGeneValidator(
            manifest={"papers": []},
            database_root=Path("star_gene_database"),
            results_root=Path("star_gene_results"),
            check_only=True,
        )
        score_results = {
            "functional_haplotype_groups": {
                "groups": {
                    "A|G": {
                        "functional_sequence": "A|G",
                        "rank_score": 1.0,
                        "mean_score": 1.2,
                        "mean_phenotype": 50.0,
                        "sample_count": 45,
                    },
                    "C|G": {
                        "functional_sequence": "C|G",
                        "rank_score": 1.4,
                        "mean_score": 1.6,
                        "mean_phenotype": 35.0,
                        "sample_count": 30,
                    },
                }
            }
        }

        self.assertEqual(
            runner.pick_directional_top_functional_group(score_results, "increases_trait")[0],
            "A|G",
        )
        self.assertEqual(
            runner.pick_directional_top_functional_group(score_results, "decreases_trait")[0],
            "C|G",
        )

    def test_direction_aware_top_functional_group_shrinks_small_group_outliers(self):
        from star_gene_validation import StarGeneValidator

        runner = StarGeneValidator(
            manifest={"papers": []},
            database_root=Path("star_gene_database"),
            results_root=Path("star_gene_results"),
            check_only=True,
        )
        score_results = {
            "functional_haplotype_groups": {
                "groups": {
                    "StableHigh": {
                        "functional_sequence": "StableHigh",
                        "rank_score": 0.82,
                        "mean_score": 0.87,
                        "mean_phenotype": 41.0,
                        "sample_count": 340,
                    },
                    "SmallSpike": {
                        "functional_sequence": "SmallSpike",
                        "rank_score": 0.22,
                        "mean_score": 0.45,
                        "mean_phenotype": 41.4,
                        "sample_count": 19,
                    },
                    "Baseline": {
                        "functional_sequence": "Baseline",
                        "rank_score": 0.98,
                        "mean_score": 1.05,
                        "mean_phenotype": 39.0,
                        "sample_count": 329,
                    },
                }
            }
        }

        self.assertEqual(
            runner.pick_directional_top_functional_group(score_results, "increases_trait")[0],
            "StableHigh",
        )

    def test_robust_discovery_builds_functional_groups_from_split_background_haplotypes(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        rows = []
        for i in range(25):
            rows.append({
                "SampleID": f"F1_{i}",
                "Hap_Name": "FuncHighBg1",
                "Haplotype_Seq": "A|G|T|A",
                "Trait": 50.0 + (i % 3),
            })
        for i in range(20):
            rows.append({
                "SampleID": f"F2_{i}",
                "Hap_Name": "FuncHighBg2",
                "Haplotype_Seq": "A|G|C|G",
                "Trait": 49.0 + (i % 3),
            })
        for i in range(30):
            rows.append({
                "SampleID": f"L1_{i}",
                "Hap_Name": "FuncLowBg1",
                "Haplotype_Seq": "C|G|T|A",
                "Trait": 35.0 + (i % 2),
            })
        hap_sample_df = pd.DataFrame(rows)
        variant_info = {
            100: {"ref": "C", "alt": "A", "annotation": "promoter", "maf": 0.45, "missing_rate": 0.0},
            200: {"ref": "G", "alt": "A", "annotation": "promoter", "maf": 0.02, "missing_rate": 0.0},
            300: {"ref": "T", "alt": "C", "annotation": "intron", "maf": 0.45, "missing_rate": 0.0},
            400: {"ref": "A", "alt": "G", "annotation": "intron", "maf": 0.45, "missing_rate": 0.0},
        }

        scorer = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200, 300, 400],
            variant_info=variant_info,
            phenotype_col="Trait",
            score_mode="robust_discovery",
        )
        result = scorer.score_all()
        functional = result.get("functional_haplotype_groups") or {}

        self.assertIn(100, functional.get("functional_positions", []))
        self.assertNotIn(300, functional.get("functional_positions", []))
        self.assertIn("A|G", functional.get("groups", {}))
        self.assertEqual(functional["groups"]["A|G"]["sample_count"], 45)
        self.assertEqual(functional["top_group"]["functional_sequence"], "A|G")

    def test_robust_discovery_treats_marker_panel_annotations_as_functional(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "S1", "Hap_Name": "Hap1", "Haplotype_Seq": "0", "Trait": 20.0},
            {"SampleID": "S2", "Hap_Name": "Hap1", "Haplotype_Seq": "0", "Trait": 21.0},
            {"SampleID": "S3", "Hap_Name": "Hap2", "Haplotype_Seq": "1", "Trait": 35.0},
            {"SampleID": "S4", "Hap_Name": "Hap2", "Haplotype_Seq": "1", "Trait": 36.0},
        ])
        variant_info = {
            10: {
                "ref": "0",
                "alt": "1",
                "annotation": "diagnostic_marker",
                "maf": 0.5,
                "missing_rate": 0.0,
            },
        }

        scorer = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[10],
            variant_info=variant_info,
            phenotype_col="Trait",
            score_mode="robust_discovery",
        )

        self.assertEqual(scorer.pos_annotation[10], "diagnostic_marker")
        self.assertGreater(scorer.pos_func_weight[10], scorer.FUNCTIONAL_WEIGHTS["other"])

    def test_robust_discovery_anchor_candidates_rank_high_weight_non_common_alleles(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        rows = []
        for i in range(20):
            rows.append({
                "SampleID": f"BG{i}",
                "Hap_Name": "BackgroundHighTotal",
                "Haplotype_Seq": "C|A|G",
                "Trait": 100.0 + i,
            })
        for i in range(3):
            rows.append({
                "SampleID": f"KEY{i}",
                "Hap_Name": "CarriesFunctionalIndel",
                "Haplotype_Seq": "C|INS_837|G",
                "Trait": 1.0 + i,
            })
        for i in range(6):
            rows.append({
                "SampleID": f"NOISE{i}",
                "Hap_Name": "BackgroundNoise",
                "Haplotype_Seq": "T|A|A",
                "Trait": 50.0 + i,
            })
        hap_sample_df = pd.DataFrame(rows)
        variant_info = {
            100: {"ref": "C", "alt": "T", "annotation": "intron", "maf": 0.2, "missing_rate": 0.0},
            200: {"ref": "A", "alt": "INS_837", "annotation": "INS", "maf": 3 / 29, "missing_rate": 0.0},
            300: {"ref": "G", "alt": "A", "annotation": "intron", "maf": 0.2, "missing_rate": 0.0},
        }

        result = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200, 300],
            variant_info=variant_info,
            phenotype_col="Trait",
            score_mode="robust_discovery",
        ).score_all()

        anchors = result["anchor_haplotype_candidates"]
        self.assertTrue(anchors["phenotype_free"])
        self.assertEqual("CarriesFunctionalIndel", anchors["top_candidate"]["haplotype"])
        self.assertEqual([200], anchors["top_candidate"]["anchor_positions"])
        self.assertGreater(
            anchors["top_candidate"]["anchor_score"],
            anchors["candidates"]["BackgroundNoise"]["anchor_score"],
        )
        self.assertIn("anchor_haplotype_candidates", result["site_weighting_policy"]["allowed_outputs"])

    def test_anchor_candidates_use_variant_info_sequence_order_for_late_marker(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "B1", "Hap_Name": "Background", "Haplotype_Seq": "A|G|DEL_837", "Trait": 0.0},
            {"SampleID": "B2", "Hap_Name": "Background", "Haplotype_Seq": "A|G|DEL_837", "Trait": 0.0},
            {"SampleID": "K1", "Hap_Name": "InsertionCarrier", "Haplotype_Seq": "A|G|INS_837", "Trait": 1.0},
        ])
        variant_info = {
            100: {"ref": "A", "alt": "T", "annotation": "intron", "maf": 0.2, "missing_rate": 0.0},
            300: {"ref": "G", "alt": "C", "annotation": "intron", "maf": 0.2, "missing_rate": 0.0},
            200: {"ref": "DEL_837", "alt": "INS_837", "annotation": "diagnostic_marker", "maf": 1 / 3, "missing_rate": 0.0},
        }

        result = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200, 300],
            variant_info=variant_info,
            phenotype_col="Trait",
            score_mode="robust_discovery",
        ).score_all()

        anchors = result["anchor_haplotype_candidates"]
        self.assertEqual("InsertionCarrier", anchors["top_candidate"]["haplotype"])
        self.assertIn(200, anchors["top_candidate"]["anchor_positions"])
        self.assertEqual("INS_837", anchors["top_candidate"]["anchor_alleles"]["200"])

    def test_anchor_candidates_prioritize_single_high_weight_site_over_background_burden(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        rows = []
        for i in range(6):
            rows.append({
                "SampleID": f"R{i}",
                "Hap_Name": "Reference",
                "Haplotype_Seq": "A|A|A|DEL_837",
                "Trait": 0.0,
            })
        for i in range(2):
            rows.append({
                "SampleID": f"B{i}",
                "Hap_Name": "BackgroundBurden",
                "Haplotype_Seq": "DEL_20|DEL_20|DEL_20|DEL_837",
                "Trait": 0.0,
            })
        for i in range(2):
            rows.append({
                "SampleID": f"K{i}",
                "Hap_Name": "HighImpactCarrier",
                "Haplotype_Seq": "A|A|A|INS_837",
                "Trait": 1.0,
            })
        hap_sample_df = pd.DataFrame(rows)
        variant_info = {
            100: {"ref": "A", "alt": "DEL_20", "annotation": "base_indel", "maf": 0.5, "missing_rate": 0.0},
            200: {"ref": "A", "alt": "DEL_20", "annotation": "base_indel", "maf": 0.5, "missing_rate": 0.0},
            300: {"ref": "A", "alt": "DEL_20", "annotation": "base_indel", "maf": 0.5, "missing_rate": 0.0},
            400: {"ref": "DEL_837", "alt": "INS_837", "annotation": "diagnostic_marker", "maf": 0.5, "missing_rate": 0.0},
        }

        result = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200, 300, 400],
            variant_info=variant_info,
            phenotype_col="Trait",
            score_mode="robust_discovery",
        ).score_all()

        anchors = result["anchor_haplotype_candidates"]
        self.assertEqual("HighImpactCarrier", anchors["top_candidate"]["haplotype"])
        self.assertEqual([400], anchors["top_candidate"]["anchor_positions"])
        self.assertGreater(
            anchors["top_candidate"]["anchor_max_site_score"],
            anchors["candidates"]["BackgroundBurden"]["anchor_max_site_score"],
        )

    def test_robust_discovery_score_is_unchanged_when_phenotype_is_inverted(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        rows = []
        for i in range(12):
            rows.append({
                "SampleID": f"F_{i}",
                "Hap_Name": "FunctionalPromoter",
                "Haplotype_Seq": "A|G|T",
                "Trait": 10.0 + i,
            })
        for i in range(12):
            rows.append({
                "SampleID": f"N_{i}",
                "Hap_Name": "BackgroundNoise",
                "Haplotype_Seq": "C|A|C",
                "Trait": 100.0 + i,
            })
        hap_sample_df = pd.DataFrame(rows)
        inverted_df = hap_sample_df.copy()
        inverted_df["Trait"] = -inverted_df["Trait"]
        variant_info = {
            100: {"ref": "C", "alt": "A", "annotation": "promoter_core", "maf": 0.5, "missing_rate": 0.0},
            200: {"ref": "G", "alt": "A", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
            300: {"ref": "T", "alt": "C", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
        }

        def score(df):
            return HaplotypeScorer(
                hap_sample_df=df,
                variant_positions=[100, 200, 300],
                variant_info=variant_info,
                phenotype_col="Trait",
                score_mode="robust_discovery",
                expected_direction="increases_trait",
            ).score_all()

        original = score(hap_sample_df)
        inverted = score(inverted_df)

        self.assertEqual(original["score_axis"], "total")
        self.assertEqual(inverted["score_axis"], "total")
        self.assertEqual(
            original["functional_haplotype_groups"]["functional_positions"],
            inverted["functional_haplotype_groups"]["functional_positions"],
        )
        for hap in ("FunctionalPromoter", "BackgroundNoise"):
            self.assertAlmostEqual(
                original["per_haplotype"][hap]["total"],
                inverted["per_haplotype"][hap]["total"],
                places=9,
            )

    def test_robust_discovery_uses_only_explicit_external_site_evidence(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "S1", "Hap_Name": "HapRef", "Haplotype_Seq": "G|C", "Trait": 1.0},
            {"SampleID": "S2", "Hap_Name": "HapRef", "Haplotype_Seq": "G|C", "Trait": 2.0},
            {"SampleID": "S3", "Hap_Name": "HapAlt", "Haplotype_Seq": "A|T", "Trait": 99.0},
            {"SampleID": "S4", "Hap_Name": "HapAlt", "Haplotype_Seq": "A|T", "Trait": 100.0},
        ])
        variant_info = {
            100: {"ref": "G", "alt": "A", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
            200: {"ref": "C", "alt": "T", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
        }

        local_pvalue_scorer = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200],
            variant_info=variant_info,
            phenotype_col="Trait",
            gwas_data=[{"pos": 200, "pvalue": 1e-12}],
            score_mode="robust_discovery",
        )
        external_pvalue_scorer = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200],
            variant_info=variant_info,
            phenotype_col="Trait",
            gwas_data=[{"pos": 200, "pvalue": 1e-12, "external": True}],
            score_mode="robust_discovery",
        )

        local_scores = {
            item["position"]: item["score"]
            for item in local_pvalue_scorer._score_positions_for_grouping()
        }
        external_scores = {
            item["position"]: item["score"]
            for item in external_pvalue_scorer._score_positions_for_grouping()
        }

        self.assertEqual(local_pvalue_scorer.pos_gwas_logp[200], 0.0)
        self.assertGreater(external_pvalue_scorer.pos_gwas_logp[200], 0.0)
        self.assertAlmostEqual(local_scores[100], local_scores[200], places=9)
        self.assertGreater(external_scores[200], external_scores[100])

    def test_robust_discovery_site_weighted_score_is_phenotype_free(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "R1", "Hap_Name": "RefHap", "Haplotype_Seq": "G|C", "Trait": 1.0},
            {"SampleID": "R2", "Hap_Name": "RefHap", "Haplotype_Seq": "G|C", "Trait": 2.0},
            {"SampleID": "R3", "Hap_Name": "RefHap", "Haplotype_Seq": "G|C", "Trait": 3.0},
            {"SampleID": "A1", "Hap_Name": "AltHap", "Haplotype_Seq": "A|T", "Trait": 90.0},
            {"SampleID": "A2", "Hap_Name": "AltHap", "Haplotype_Seq": "A|T", "Trait": 91.0},
            {"SampleID": "A3", "Hap_Name": "AltHap", "Haplotype_Seq": "A|T", "Trait": 92.0},
        ])
        shuffled_df = hap_sample_df.copy()
        shuffled_df["Trait"] = [50.0, 10.0, 80.0, 20.0, 70.0, 30.0]
        variant_info = {
            100: {"ref": "G", "alt": "A", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
            200: {"ref": "C", "alt": "T", "annotation": "promoter_core", "maf": 0.5, "missing_rate": 0.0},
        }

        def score(df):
            return HaplotypeScorer(
                hap_sample_df=df,
                variant_positions=[100, 200],
                variant_info=variant_info,
                phenotype_col="Trait",
                gwas_data=[{"pos": 100, "pvalue": 1e-9, "external": True}],
                score_mode="robust_discovery",
            ).score_all()

        original = score(hap_sample_df)
        shuffled = score(shuffled_df)

        self.assertIn("site_weighted", original["component_weights"])
        self.assertIn("site_weighted", original["per_haplotype"]["AltHap"])
        self.assertIn("site_weights", original)
        self.assertTrue(original["site_weighting_policy"]["phenotype_free"])
        self.assertFalse(original["site_weighting_policy"]["current_phenotype_used"])

        by_pos = {row["position"]: row for row in original["site_weights"]}
        self.assertGreater(by_pos[100]["external_weight"], 0)
        self.assertEqual(by_pos[100]["current_phenotype_used"], False)
        self.assertEqual(by_pos[200]["current_phenotype_used"], False)

        self.assertEqual(original["site_weights"], shuffled["site_weights"])
        for hap in ("RefHap", "AltHap"):
            self.assertAlmostEqual(
                original["per_haplotype"][hap]["site_weighted"],
                shuffled["per_haplotype"][hap]["site_weighted"],
                places=9,
            )
            self.assertAlmostEqual(
                original["per_haplotype"][hap]["total"],
                shuffled["per_haplotype"][hap]["total"],
                places=9,
            )

    def test_robust_discovery_uses_external_attention_prior_without_phenotype_leakage(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        rows = []
        for i in range(8):
            rows.append({
                "SampleID": f"R{i}",
                "Hap_Name": "Reference",
                "Haplotype_Seq": "G|C",
                "Trait": float(i),
            })
            rows.append({
                "SampleID": f"P{i}",
                "Hap_Name": "AttentionCarrier",
                "Haplotype_Seq": "A|C",
                "Trait": 100.0 + i,
            })
            rows.append({
                "SampleID": f"B{i}",
                "Hap_Name": "BackgroundAlt",
                "Haplotype_Seq": "G|T",
                "Trait": 50.0 + i,
            })
        hap_sample_df = pd.DataFrame(rows)
        inverted_df = hap_sample_df.copy()
        inverted_df["Trait"] = -inverted_df["Trait"]
        variant_info = {
            100: {
                "ref": "G",
                "alt": "A",
                "annotation": "intron",
                "maf": 1 / 3,
                "missing_rate": 0.0,
                "source_type": "external_attention_prior",
                "attention_prior_score": 0.98,
                "attention_percentile": 0.99,
                "attention_cluster_id": "atlas_like_cluster_1",
                "source": "ATLAS-like phenotype-free attention prior",
            },
            200: {"ref": "C", "alt": "T", "annotation": "intron", "maf": 1 / 3, "missing_rate": 0.0},
        }

        def score(df):
            return HaplotypeScorer(
                hap_sample_df=df,
                variant_positions=[100, 200],
                variant_info=variant_info,
                phenotype_col="Trait",
                score_mode="robust_discovery",
            ).score_all()

        original = score(hap_sample_df)
        inverted = score(inverted_df)
        by_pos = {row["position"]: row for row in original["site_weights"]}

        self.assertIn("external_attention_prior", original["site_weighting_policy"]["allowed_inputs"])
        self.assertGreater(by_pos[100]["attention_prior_weight"], 0.9)
        self.assertEqual(by_pos[100]["attention_cluster_id"], "atlas_like_cluster_1")
        self.assertGreater(by_pos[100]["total_site_weight"], by_pos[200]["total_site_weight"])
        self.assertEqual(
            "AttentionCarrier",
            original["anchor_haplotype_candidates"]["top_candidate"]["haplotype"],
        )
        self.assertEqual(original["site_weights"], inverted["site_weights"])
        self.assertEqual(
            original["anchor_haplotype_candidates"]["top_candidate"]["haplotype"],
            inverted["anchor_haplotype_candidates"]["top_candidate"]["haplotype"],
        )
        for hap in ("Reference", "AttentionCarrier", "BackgroundAlt"):
            self.assertAlmostEqual(
                original["per_haplotype"][hap]["total"],
                inverted["per_haplotype"][hap]["total"],
                places=9,
            )

    def test_robust_discovery_ignores_attention_prior_without_explicit_external_marker(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "R1", "Hap_Name": "Reference", "Haplotype_Seq": "G", "Trait": 1.0},
            {"SampleID": "R2", "Hap_Name": "Reference", "Haplotype_Seq": "G", "Trait": 2.0},
            {"SampleID": "A1", "Hap_Name": "AttentionLike", "Haplotype_Seq": "A", "Trait": 9.0},
            {"SampleID": "A2", "Hap_Name": "AttentionLike", "Haplotype_Seq": "A", "Trait": 10.0},
        ])

        result = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100],
            variant_info={
                100: {
                    "ref": "G",
                    "alt": "A",
                    "annotation": "intron",
                    "maf": 0.5,
                    "missing_rate": 0.0,
                    "source_type": "attention_prior",
                    "attention_prior_score": 0.99,
                }
            },
            phenotype_col="Trait",
            score_mode="robust_discovery",
        ).score_all()

        self.assertEqual(result["site_weights"][0]["attention_prior_weight"], 0.0)

    def test_variant_info_csv_row_preserves_external_attention_prior_columns(self):
        from haplotype_phenotype_analysis import _variant_info_csv_row_to_record
        import pandas as pd

        record = _variant_info_csv_row_to_record(pd.Series({
            "position": 100,
            "ref": "G",
            "alt": "A",
            "len_diff": 0,
            "is_sv": "False",
            "maf": 0.25,
            "missing_rate": 0.01,
            "annotation": "intron",
            "source_type": "external_attention_prior",
            "attention_prior_score": 0.93,
            "attention_cluster_id": "cluster_B",
            "source": "ATLAS-like prior export",
        }))

        self.assertEqual(record["ref"], "G")
        self.assertFalse(record["is_sv"])
        self.assertEqual(record["source_type"], "external_attention_prior")
        self.assertAlmostEqual(record["attention_prior_score"], 0.93)
        self.assertEqual(record["attention_cluster_id"], "cluster_B")
        self.assertEqual(record["source"], "ATLAS-like prior export")

    def test_numeric_csv_external_attention_flag_drives_scoring(self):
        from haplotype_phenotype_analysis import HaplotypeScorer, _variant_info_csv_row_to_record
        import pandas as pd

        record = _variant_info_csv_row_to_record(pd.Series({
            "position": 100,
            "ref": "G",
            "alt": "A",
            "annotation": "intron",
            "maf": 0.5,
            "missing_rate": 0.0,
            "external_attention_prior": 1,
            "attention_prior_score": 0.91,
        }))
        hap_sample_df = pd.DataFrame([
            {"SampleID": "R1", "Hap_Name": "Reference", "Haplotype_Seq": "G", "Trait": 1.0},
            {"SampleID": "R2", "Hap_Name": "Reference", "Haplotype_Seq": "G", "Trait": 2.0},
            {"SampleID": "A1", "Hap_Name": "AttentionCarrier", "Haplotype_Seq": "A", "Trait": 9.0},
            {"SampleID": "A2", "Hap_Name": "AttentionCarrier", "Haplotype_Seq": "A", "Trait": 10.0},
        ])

        result = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100],
            variant_info={100: record},
            phenotype_col="Trait",
            score_mode="robust_discovery",
        ).score_all()

        self.assertGreater(result["site_weights"][0]["attention_prior_weight"], 0.9)

    def test_site_weight_keeps_rare_high_impact_functional_sites(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        rows = []
        for i in range(100):
            rows.append({
                "SampleID": f"R{i}",
                "Hap_Name": "RefHap",
                "Haplotype_Seq": "G",
                "Trait": float(i),
            })
        rows.append({
            "SampleID": "A1",
            "Hap_Name": "RareStop",
            "Haplotype_Seq": "T",
            "Trait": 1.0,
        })
        hap_sample_df = pd.DataFrame(rows)
        scorer = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[18781242],
            variant_info={
                18781242: {
                    "ref": "G",
                    "alt": "T",
                    "annotation": "stop_gain",
                    "maf": 0.003,
                    "missing_rate": 0.0,
                }
            },
            phenotype_col="Trait",
            score_mode="robust_discovery",
        )

        result = scorer.score_all()

        self.assertEqual(len(result["site_weights"]), 1)
        self.assertEqual(result["site_weights"][0]["annotation"], "stop_gain")
        self.assertGreater(result["per_haplotype"]["RareStop"]["site_weighted"], 0)

    def test_site_weight_keeps_snp_eff_stop_gained_annotation(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        rows = []
        for i in range(100):
            rows.append({
                "SampleID": f"R{i}",
                "Hap_Name": "RefHap",
                "Haplotype_Seq": "G",
                "Trait": float(i),
            })
        rows.append({
            "SampleID": "A1",
            "Hap_Name": "RareStop",
            "Haplotype_Seq": "T",
            "Trait": 1.0,
        })
        hap_sample_df = pd.DataFrame(rows)
        scorer = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[18781242],
            variant_info={
                18781242: {
                    "ref": "G",
                    "alt": "T",
                    "annotation": "stop_gained",
                    "maf": 0.003,
                    "missing_rate": 0.0,
                }
            },
            phenotype_col="Trait",
            score_mode="robust_discovery",
        )

        result = scorer.score_all()

        self.assertEqual(len(result["site_weights"]), 1)
        self.assertEqual(result["site_weights"][0]["annotation"], "stop_gain")
        self.assertGreater(result["per_haplotype"]["RareStop"]["site_weighted"], 0)

    def test_site_weight_ignores_unmarked_variant_info_pvalues(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "S1", "Hap_Name": "HapRef", "Haplotype_Seq": "G|C", "Trait": 1.0},
            {"SampleID": "S2", "Hap_Name": "HapRef", "Haplotype_Seq": "G|C", "Trait": 2.0},
            {"SampleID": "S3", "Hap_Name": "HapAlt", "Haplotype_Seq": "A|T", "Trait": 99.0},
            {"SampleID": "S4", "Hap_Name": "HapAlt", "Haplotype_Seq": "A|T", "Trait": 100.0},
        ])
        base_variant_info = {
            100: {"ref": "G", "alt": "A", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
            200: {"ref": "C", "alt": "T", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
        }
        local_variant_info = {
            100: dict(base_variant_info[100]),
            200: {
                **base_variant_info[200],
                "minus_log10_p": 12.0,
                "gwas_pvalue": 1e-12,
                "site_score": 1.0,
            },
        }
        external_variant_info = {
            100: dict(base_variant_info[100]),
            200: {
                **base_variant_info[200],
                "minus_log10_p": 12.0,
                "gwas_pvalue": 1e-12,
                "site_score": 1.0,
                "external": True,
            },
        }

        def site_weights(variant_info):
            result = HaplotypeScorer(
                hap_sample_df=hap_sample_df,
                variant_positions=[100, 200],
                variant_info=variant_info,
                phenotype_col="Trait",
                score_mode="robust_discovery",
            ).score_all()
            return {row["position"]: row for row in result["site_weights"]}

        base_weights = site_weights(base_variant_info)
        local_weights = site_weights(local_variant_info)
        external_weights = site_weights(external_variant_info)

        self.assertAlmostEqual(
            local_weights[200]["external_weight"],
            base_weights[200]["external_weight"],
            places=9,
        )
        self.assertAlmostEqual(
            local_weights[200]["total_site_weight"],
            base_weights[200]["total_site_weight"],
            places=9,
        )
        self.assertGreater(external_weights[200]["external_weight"], base_weights[200]["external_weight"])
        self.assertGreater(external_weights[200]["total_site_weight"], base_weights[200]["total_site_weight"])

    def test_robust_discovery_adds_boundary_gene_body_signal_to_functional_groups(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        rows = []
        for i in range(20):
            rows.append({
                "SampleID": f"H1_{i}",
                "Hap_Name": "HighBg1",
                "Haplotype_Seq": "A|G|C|T",
                "Trait": 50.0 + (i % 3),
            })
        for i in range(20):
            rows.append({
                "SampleID": f"H2_{i}",
                "Hap_Name": "HighBg2",
                "Haplotype_Seq": "A|G|C|A",
                "Trait": 51.0 + (i % 3),
            })
        for i in range(30):
            rows.append({
                "SampleID": f"L1_{i}",
                "Hap_Name": "LowBg1",
                "Haplotype_Seq": "C|G|T|T",
                "Trait": 35.0 + (i % 2),
            })
        hap_sample_df = pd.DataFrame(rows)
        variant_info = {
            100: {"ref": "C", "alt": "A", "annotation": "promoter", "maf": 0.45, "missing_rate": 0.0},
            200: {"ref": "G", "alt": "A", "annotation": "promoter", "maf": 0.05, "missing_rate": 0.0},
            260: {"ref": "T", "alt": "C", "annotation": "intron", "maf": 0.45, "missing_rate": 0.0},
            1000: {"ref": "T", "alt": "A", "annotation": "intron", "maf": 0.30, "missing_rate": 0.0},
        }

        scorer = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200, 260, 1000],
            variant_info=variant_info,
            phenotype_col="Trait",
            score_mode="robust_discovery",
            gene_start=250,
            gene_end=5000,
            strand="+",
        )
        result = scorer.score_all()
        functional = result.get("functional_haplotype_groups") or {}

        self.assertIn(100, functional.get("functional_positions", []))
        self.assertIn(260, functional.get("functional_positions", []))
        self.assertNotIn(1000, functional.get("functional_positions", []))
        self.assertIn("A|G|C", functional.get("groups", {}))
        self.assertEqual(functional["groups"]["A|G|C"]["sample_count"], 40)

    def test_functional_position_selection_keeps_large_indel_representative_after_boundary_cluster(self):
        from haplotype_phenotype_analysis import HaplotypeScorer
        import pandas as pd

        positions = [100] + list(range(200, 211)) + [2000]

        def seq(promoter, boundary, structural):
            return "|".join([promoter] + [boundary] * 11 + [structural])

        rows = []
        for i in range(18):
            rows.append({
                "SampleID": f"Ref_{i}",
                "Hap_Name": "ReferenceBackground",
                "Haplotype_Seq": seq("C", "G", "REFBLOCK"),
                "Trait": float(i % 5),
            })
        for i in range(16):
            rows.append({
                "SampleID": f"Boundary_{i}",
                "Hap_Name": "BoundaryClusterOnly",
                "Haplotype_Seq": seq("A", "A", "REFBLOCK"),
                "Trait": float(i % 5),
            })
        for i in range(14):
            rows.append({
                "SampleID": f"Structural_{i}",
                "Hap_Name": "LargeIndelCarrier",
                "Haplotype_Seq": seq("A", "A", "INS_" + "T" * 80),
                "Trait": float(i % 5),
            })
        hap_sample_df = pd.DataFrame(rows)
        variant_info = {
            100: {"ref": "C", "alt": "A", "annotation": "promoter", "maf": 0.45, "missing_rate": 0.0},
            2000: {
                "ref": "N",
                "alt": "N",
                "annotation": "base_indel",
                "marker_id": "VRNB1gene_indel_2000_3400",
                "maf": 0.30,
                "missing_rate": 0.0,
            },
        }
        for pos in range(200, 211):
            variant_info[pos] = {
                "ref": "G",
                "alt": "A",
                "annotation": "intron",
                "maf": 0.45,
                "missing_rate": 0.0,
            }

        scorer = HaplotypeScorer(
            hap_sample_df=hap_sample_df,
            variant_positions=positions,
            variant_info=variant_info,
            phenotype_col="Trait",
            score_mode="robust_discovery",
            gene_start=200,
            gene_end=3000,
            strand="+",
        )
        functional = scorer.score_all().get("functional_haplotype_groups") or {}

        self.assertIn(100, functional.get("functional_positions", []))
        self.assertIn(2000, functional.get("functional_positions", []))
        boundary_count = sum(1 for pos in functional.get("functional_positions", []) if 200 <= pos <= 210)
        self.assertLess(boundary_count, 11)

    def test_score_tooltips_use_viewport_coordinates(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("function drawHaplotypeScorePlot(scoreData)")
        integrated_end = source.index("function updateScoreLegend", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        standalone_start = source.index("def generate_haplotype_score_html")
        standalone_end = source.index("def generate_effect_boxplot_html", standalone_start)
        standalone_block = source[standalone_start:standalone_end]

        for block in (integrated_block, standalone_block):
            self.assertIn("clientX", block)
            self.assertIn("clientY", block)
            self.assertNotIn("pageX", block)
            self.assertNotIn("pageY", block)

    def test_haplotype_score_html_shows_only_current_robust_mode(self):
        from haplotype_phenotype_analysis import ReportGenerator

        def score_bundle(mode, score):
            return {
                "score_mode": mode,
                "per_sample": [
                    {"sample_id": "S1", "haplotype": "Hap1", "score": score, "phenotype": 10.0},
                    {"sample_id": "S2", "haplotype": "Hap2", "score": score + 1.0, "phenotype": 12.0},
                ],
                "per_haplotype": {
                    "Hap1": {"total": score, "sample_reliability": 0.9},
                    "Hap2": {"total": score + 1.0, "sample_reliability": 0.8},
                },
                "component_weights": {"variant_effect": 1.0},
                "r_squared": 0.5,
                "regression_pvalue": 0.01,
                "slope": 1.0,
                "intercept": 9.0,
            }

        with tempfile.TemporaryDirectory() as tmp:
            robust_dir = Path(tmp) / "GeneA__robust_discovery"
            default_dir = Path(tmp) / "GeneA"
            robust_dir.mkdir()
            default_dir.mkdir()
            (default_dir / "haplotype_scores.json").write_text(
                json.dumps({"Trait": score_bundle("default", 1.0)}),
                encoding="utf-8",
            )

            generator = ReportGenerator(output_dir=str(robust_dir))
            generator._cached_all_haplotype_scores = {"Trait": score_bundle("robust_discovery", 2.0)}
            generator._cached_haplotype_score = generator._cached_all_haplotype_scores["Trait"]

            html_path = generator.generate_haplotype_score_html(gene_id="GeneA", phenotype_col="Trait")
            html = Path(html_path).read_text(encoding="utf-8")

            self.assertIn("var allScoreModeData", html)
            self.assertIn('"robust_discovery"', html)
            self.assertIn('"current_mode": "robust_discovery"', html)
            self.assertIn('"score": 2.0', html)
            self.assertNotIn('"score": 1.0', html)
            self.assertNotIn('"default"', html)
            self.assertNotIn("switchScoreMode('default')", html)
            self.assertNotIn("switchScoreMode('robust_discovery')", html)
            self.assertNotIn("mode-toggle-btn", html)

    def test_haplotype_score_html_escapes_dynamic_score_mode_label(self):
        from haplotype_phenotype_analysis import ReportGenerator

        malicious_mode = "robust_discovery<img src=x onerror=alert(1)>"
        score_data = {
            "score_mode": malicious_mode,
            "per_sample": [
                {"sample_id": "S1", "haplotype": "Hap1", "score": 1.0, "phenotype": 10.0},
                {"sample_id": "S2", "haplotype": "Hap2", "score": 2.0, "phenotype": 12.0},
            ],
            "per_haplotype": {
                "Hap1": {"total": 1.0, "sample_reliability": 0.9},
                "Hap2": {"total": 2.0, "sample_reliability": 0.8},
            },
            "component_weights": {"variant_effect": 1.0},
            "r_squared": 0.5,
            "regression_pvalue": 0.01,
        }

        with tempfile.TemporaryDirectory() as tmp:
            generator = ReportGenerator(output_dir=str(Path(tmp) / "GeneA__robust_discovery"))
            generator._cached_all_haplotype_scores = {"Trait": score_data}
            generator._cached_haplotype_score = score_data

            html_path = generator.generate_haplotype_score_html(gene_id="GeneA", phenotype_col="Trait")
            html = Path(html_path).read_text(encoding="utf-8")

            self.assertNotIn("<img", html)
            self.assertIn("\\u003cimg", html)
            self.assertNotIn("status.innerHTML", html)

    def test_score_plot_display_haplotypes_filter_tiny_groups_and_cap_at_five(self):
        from haplotype_phenotype_analysis import ReportGenerator

        hap_counts = {
            "Hap1": 6,
            "Hap2": 1,
            "Hap3": 5,
            "Hap4": 3,
            "Hap5": 4,
            "Hap6": 3,
            "Hap7": 3,
        }
        hap_scores = {
            "Hap1": 1.1,
            "Hap2": 9.9,
            "Hap3": 1.7,
            "Hap4": 1.3,
            "Hap5": 1.6,
            "Hap6": 1.5,
            "Hap7": 1.4,
        }
        per_sample = []
        for hap, n in hap_counts.items():
            for i in range(n):
                per_sample.append({
                    "sample_id": f"{hap}_S{i}",
                    "haplotype": hap,
                    "score": hap_scores[hap],
                    "phenotype": float(i),
                })

        generator = ReportGenerator(output_dir=tempfile.mkdtemp())
        score_data = {
            "score_mode": "robust_discovery",
            "per_sample": per_sample,
            "per_haplotype": {
                hap: {"total": score, "sample_count": hap_counts[hap]}
                for hap, score in hap_scores.items()
            },
        }

        mode_data = generator._collect_score_mode_data({"Trait": score_data})
        trait_data = mode_data["modes"]["robust_discovery"]["Trait"]

        self.assertEqual(
            ["Hap3", "Hap5", "Hap6", "Hap7", "Hap4"],
            trait_data["score_plot_haplotypes"],
        )
        self.assertNotIn("Hap2", trait_data["score_plot_haplotypes"])
        self.assertEqual(3, trait_data["score_plot_policy"]["min_samples"])
        self.assertEqual(5, trait_data["score_plot_policy"]["max_haplotypes"])
        self.assertEqual(sum(hap_counts.values()), len(trait_data["per_sample"]))

    def test_integrated_html_shows_robust_only_without_score_mode_toggle(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("var allScoreModeData", integrated_block)
        self.assertIn("currentScoreMode = allScoreModeData.current_mode || 'robust_discovery'", integrated_block)
        self.assertIn("Robust discovery", integrated_block)
        self.assertIn("html = html.replace('{score_mode_status_html}', score_mode_status_html)", integrated_block)
        self.assertIn("updateScoreModeStatus(scoreData)", integrated_block)
        self.assertNotIn("switchScoreMode('default')", integrated_block)
        self.assertNotIn("switchScoreMode('robust_discovery')", integrated_block)
        self.assertNotIn("mode-toggle-btn", integrated_block)
        self.assertNotIn("original / robust", integrated_block)
        self.assertNotIn("status.innerHTML", integrated_block)

    def test_integrated_html_replaces_score_mode_status_placeholder(self):
        from haplotype_phenotype_analysis import ReportGenerator
        import pandas as pd

        score_data = {
            "score_mode": "robust_discovery",
            "per_sample": [
                {"sample_id": "S1", "haplotype": "Hap1", "score": 1.0, "phenotype": 10.0},
                {"sample_id": "S2", "haplotype": "Hap2", "score": 2.0, "phenotype": 12.0},
            ],
            "per_haplotype": {
                "Hap1": {"total": 1.0, "sample_reliability": 0.9},
                "Hap2": {"total": 2.0, "sample_reliability": 0.8},
            },
            "component_weights": {"variant_effect": 1.0},
            "r_squared": 0.5,
            "regression_pvalue": 0.01,
        }

        with tempfile.TemporaryDirectory() as tmp:
            df = pd.DataFrame({
                "SampleID": ["S1", "S2"],
                "Hap_Name": ["Hap1", "Hap2"],
                "Haplotype_Seq": ["A", "G"],
                "Trait": [10.0, 12.0],
            })
            generator = ReportGenerator(output_dir=tmp)
            generator._cached_all_haplotype_scores = {"Trait": score_data}
            generator._cached_haplotype_score = score_data
            generator.generate_integrated_html(
                hap_sample_df=df,
                effect_results={"grand_mean": 11.0, "haplotype_effects": []},
                variant_positions=[100],
                region_start=50,
                region_end=150,
                phenotype_col="Trait",
                gene_start=80,
                gene_end=150,
                promoter_start=50,
                promoter_end=79,
                chrom="chr1",
                gene_id="GeneX",
                variant_info={100: {"ref": "A", "alt": "G", "annotation": "base_snp"}},
                variant_pvalues={100: 0.05},
                snp_effects={100: "base_snp"},
            )
            html = (Path(tmp) / "GeneX.html").read_text(encoding="utf-8")

        self.assertNotIn("{score_mode_status_html}", html)
        self.assertIn('class="score-mode-current">Robust discovery</span>', html)
        self.assertNotIn("status.innerHTML", html)
        self.assertNotIn("mode-toggle-btn", html)
        self.assertIn('id="rangeStartSlider" aria-label="Display range start handle" min="50" max="150" value="50"', html)
        self.assertIn('id="rangeEndSlider" aria-label="Display range end handle" min="50" max="150" value="150"', html)
        self.assertIn('<span>50</span>', html)
        self.assertIn('<span>150</span>', html)

    def test_integrated_html_initializes_a_neutral_25_variant_window(self):
        from haplotype_phenotype_analysis import ReportGenerator

        positions = list(range(1, 41))
        variant_info = {
            pos: {"ref": "A", "alt": "G", "annotation": "base_snp"}
            for pos in positions
        }
        with tempfile.TemporaryDirectory() as tmp:
            df = pd.DataFrame({
                "SampleID": ["S1", "S2", "S3", "S4"],
                "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap2"],
                "Haplotype_Seq": [
                    "|".join(["A"] * 40),
                    "|".join(["A"] * 40),
                    "|".join(["G"] * 40),
                    "|".join(["G"] * 40),
                ],
                "Trait": [10.0, 11.0, 20.0, 21.0],
            })
            generator = ReportGenerator(output_dir=tmp)
            generator.generate_integrated_html(
                hap_sample_df=df,
                effect_results={"grand_mean": 15.5, "haplotype_effects": []},
                variant_positions=positions,
                region_start=1,
                region_end=40,
                phenotype_col="Trait",
                gene_start=1,
                gene_end=40,
                promoter_start=1,
                promoter_end=0,
                chrom="chr1",
                gene_id="GeneX",
                variant_info=variant_info,
                variant_pvalues={pos: 0.5 for pos in positions},
                snp_effects={pos: "base_snp" for pos in positions},
            )
            html = (Path(tmp) / "GeneX.html").read_text(encoding="utf-8")

        self.assertIn('var initialDisplayRange = {"start": 8, "end": 32};', html)
        self.assertIn(
            "var currentFilter = { maf: 0.05, missingRate: 0.2, "
            "rangeStart: initialDisplayRange.start, rangeEnd: initialDisplayRange.end };",
            html,
        )
        self.assertIn(
            "var pendingRangeFilter = { start: initialDisplayRange.start, "
            "end: initialDisplayRange.end };",
            html,
        )

    def test_integrated_html_keeps_validation_markers_visible_by_default(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("function isPriorityValidationSite", integrated_block)
        self.assertIn("diagnostic_marker", integrated_block)
        self.assertIn("functional_marker", integrated_block)
        self.assertIn(
            "return rangePass && (priorityPass || (mafPass && missPass && annPass && typePass && synPass));",
            integrated_block,
        )

    def test_integrated_html_has_local_candidate_evidence_panel(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("def _summarize_post_gwas_evidence", source)
        self.assertIn("post-gwas-evidence", integrated_block)
        self.assertIn("Local Candidate Evidence", integrated_block)
        self.assertIn("Discovery-safe evidence", integrated_block)
        self.assertIn("Candidate rule", integrated_block)
        self.assertIn("Top local marker", integrated_block)
        self.assertIn("postGwasEvidence", integrated_block)
        self.assertIn("html = html.replace('{post_gwas_evidence_json}', post_gwas_evidence_json)", integrated_block)

    def test_integrated_post_gwas_panel_is_not_between_gwas_and_gene_structure(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        evidence_idx = integrated_block.index('<section class="post-gwas-evidence"')
        top_section_idx = integrated_block.index('<div class="top-section">')
        gene_svg_idx = integrated_block.index('gene-structure-svg')

        self.assertLess(
            evidence_idx,
            top_section_idx,
            "Post-GWAS evidence should sit outside the GWAS-to-gene guide-line corridor.",
        )
        self.assertLess(top_section_idx, gene_svg_idx)

    def test_integrated_report_guide_lines_and_ld_align_to_sequence_columns(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("gwas_bottom_extension = gwas_panel_h - gwas_mt - gwas_i_h", integrated_block)
        self.assertIn("top_section_to_gene_gap = 10", integrated_block)
        self.assertIn("up_line_extension = var_top_y + top_section_to_gene_gap", integrated_block)
        self.assertIn("score_panel_w = max(360, gene_area_start - score_gap)", integrated_block)
        self.assertIn("flex: 0 0 {score_panel_w}px", integrated_block)
        self.assertIn("canvasW_screen = Math.max(canvasW_screen, cellW * ncAbs);", integrated_block)
        self.assertNotIn("canvasW_screen = Math.max(canvasW_screen, 100);", integrated_block)
        self.assertIn("html = html.replace('{gwas_svg_h}', str(gwas_svg_h))", integrated_block)
        self.assertIn("html = html.replace('{gwas_mt}', str(gwas_mt))", integrated_block)
        self.assertIn("html = html.replace('{gwas_mr}', str(gwas_mr))", integrated_block)
        self.assertIn("html = html.replace('{gwas_mb}', str(gwas_mb))", integrated_block)

    def test_integrated_report_ld_sidebar_uses_its_own_visual_scale(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]
        ld_start = integrated_block.index("function drawLDTriangle()")
        ld_end = integrated_block.index("function scheduleConnectorRedraw()", ld_start)
        ld_block = integrated_block[ld_start:ld_end]

        self.assertIn("var contentZoomFactor = 1;", ld_block)
        self.assertIn("var ldZoomFactor = 1;", ld_block)
        self.assertIn("var isSidebarLD = !!wrapper.closest('.report-sidebar');", ld_block)
        self.assertIn("var padLeft = isSidebarLD ? 0", ld_block)
        self.assertIn("canvasW_screen / ldZoomFactor", ld_block)
        self.assertIn("canvasH / ldZoomFactor", ld_block)
        self.assertNotIn("canvasW_screen / zoomFactor", ld_block)

    def test_integrated_report_uses_right_sidebar_workbench_without_losing_controls(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("report-shell", integrated_block)
        self.assertIn("report-sidebar", integrated_block)
        self.assertIn("openReportSidebar", integrated_block)
        self.assertIn("chooseReportComponent('score')", integrated_block)
        self.assertIn("chooseReportComponent('network')", integrated_block)
        self.assertIn("chooseReportComponent('filters')", integrated_block)
        self.assertIn("chooseReportComponent('ld')", integrated_block)
        self.assertIn("chooseReportComponent('audit')", integrated_block)

        # Existing controls must remain addressable by their original JS ids.
        for required_id in [
            'id="mafSlider"',
            'id="missingSlider"',
            'id="zoomSlider"',
            'id="zoomContent"',
            'id="score-scatter-viz"',
            'id="ld-triangle-canvas"',
            'id="network-viz"',
            'id="networkModeBtn"',
        ]:
            self.assertIn(required_id, integrated_block)

        self.assertNotIn('id="mode-btn-default"', integrated_block)
        self.assertNotIn('id="mode-btn-robust_discovery"', integrated_block)
        self.assertIn("function captureMainDataCenter()", integrated_block)
        self.assertIn("function restoreMainDataCenter(centerRatio)", integrated_block)
        self.assertIn("zc.style.zoom = String(cz / 100);", integrated_block)
        self.assertIn("zc.style.transform = 'none';", integrated_block)
        self.assertNotIn("zc.style.transform = 'scale('", integrated_block)
        self.assertIn("function getIntrinsicReportWidth()", integrated_block)
        self.assertIn("scheduleConnectorRedraw();", integrated_block)
        self.assertIn("scheduleLDTriangleRedraw();", integrated_block)

    def test_integrated_report_has_position_range_filter_controls(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn('id="rangeStartInput"', integrated_block)
        self.assertIn('id="rangeEndInput"', integrated_block)
        self.assertIn('id="rangeStartSlider"', integrated_block)
        self.assertIn('id="rangeEndSlider"', integrated_block)
        self.assertIn('id="rangeSliderFill"', integrated_block)
        self.assertIn('id="rangeSliderLabel"', integrated_block)
        self.assertIn('id="rangeApplyBtn"', integrated_block)
        self.assertIn('id="rangePendingStatus"', integrated_block)
        self.assertIn("updateRangeFilterFromInputs('start')", integrated_block)
        self.assertIn("updateRangeFilterFromInputs('end')", integrated_block)
        self.assertIn("updateRangeFilterFromSliders('start')", integrated_block)
        self.assertIn("updateRangeFilterFromSliders('end')", integrated_block)
        self.assertIn('onclick="commitRangeFilter()"', integrated_block)
        self.assertIn("function readRangeFilter(changedHandle)", integrated_block)
        self.assertIn("function syncRangeControls(range, preserveInputs)", integrated_block)
        self.assertIn("function updateRangeFilterFromSliders(changedHandle)", integrated_block)
        self.assertIn("function activateRangeHandle(handle)", integrated_block)
        self.assertIn("var initialDisplayRange = {initial_display_range_json};", integrated_block)
        self.assertIn(
            "var pendingRangeFilter = { start: initialDisplayRange.start, end: initialDisplayRange.end };",
            integrated_block,
        )
        self.assertIn("function updateRangePendingState()", integrated_block)
        self.assertIn("function setPendingRangeFilter(range, preserveInputs)", integrated_block)
        self.assertIn("function commitRangeFilter()", integrated_block)
        self.assertNotIn("function scheduleRangeFilterApply()", integrated_block)
        self.assertIn("function beginRangeSliderDrag(event)", integrated_block)
        self.assertIn("function moveRangeSliderDrag(event)", integrated_block)
        self.assertIn("function endRangeSliderDrag(event)", integrated_block)
        self.assertIn("function chooseRangeHandle(pointerValue, start, end, lastHandle)", integrated_block)
        self.assertIn("function resolveDraggedRange(handle, value, start, end)", integrated_block)
        self.assertIn("function scheduleLDTriangleRedraw()", integrated_block)
        self.assertIn("clearTimeout(ldRedrawTimer)", integrated_block)
        for function_name in (
            "updateRangeFilterFromInputs",
            "updateRangeFilterFromSliders",
            "clearRangeFilter",
        ):
            function_start = integrated_block.index(f"function {function_name}(")
            function_end = integrated_block.find("\nfunction ", function_start + 1)
            function_block = integrated_block[function_start:function_end]
            self.assertNotIn("applyFilters();", function_block)
        commit_start = integrated_block.index("function commitRangeFilter()")
        commit_end = integrated_block.find("\nfunction ", commit_start + 1)
        commit_block = integrated_block[commit_start:commit_end]
        self.assertEqual(commit_block.count("applyFilters();"), 1)
        self.assertIn("canonicalizeRangeFilter(pendingRangeFilter)", commit_block)
        self.assertNotIn("canonicalizeRangeFilter(readRangeFilter())", commit_block)
        message_start = integrated_block.index("window.addEventListener('message'")
        message_end = integrated_block.index("function scheduleConnectorRedraw", message_start)
        message_block = integrated_block[message_start:message_end]
        self.assertIn(
            "f.rangeStart !== undefined ? f.rangeStart : currentFilter.rangeStart",
            message_block,
        )
        self.assertIn(
            "f.rangeEnd !== undefined ? f.rangeEnd : currentFilter.rangeEnd",
            message_block,
        )
        self.assertIn("aria-valuetext", integrated_block)
        self.assertNotIn('aria-live="polite"', integrated_block)
        self.assertIn("rangeStartInput.value = ''", integrated_block)
        self.assertIn("rangeEndInput.value = ''", integrated_block)
        self.assertIn("rangePass = inDisplayRange(d.pos)", integrated_block)
        self.assertIn("rangePass && (priorityPass || (mafPass && missPass && annPass && typePass && synPass))", integrated_block)
        self.assertIn("rangeStart: initialDisplayRange.start", integrated_block)
        self.assertIn("rangeEnd: initialDisplayRange.end", integrated_block)
        self.assertIn("setPendingRangeFilter(initialDisplayRange);", integrated_block)
        export_start = integrated_block.index("function prepareReportExportVisuals")
        export_end = integrated_block.index("function exportSVG", export_start)
        export_block = integrated_block[export_start:export_end]
        self.assertIn("applyFilters();", export_block)
        self.assertNotIn("drawGWASPlot(gwasData);", export_block)

    def test_applied_display_range_drives_gene_structure_and_gwas_domains(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("function getAppliedCoordinateDomain()", integrated_block)
        self.assertIn("function updateGeneStructureDomain()", integrated_block)
        self.assertIn('id="gene-axis-layer"', integrated_block)
        self.assertIn('id="gene-model-layer"', integrated_block)
        self.assertIn("var geneModelData = __GENE_MODEL_DATA__;", integrated_block)
        self.assertIn("var coordinateDomain = getAppliedCoordinateDomain();", integrated_block)
        self.assertIn("updateGeneStructureDomain();", integrated_block)
        self.assertIn(".domain(getD3CoordinateDomain(coordinateDomain))", integrated_block)
        self.assertNotIn(".domain([regionStart, regionEnd])", integrated_block)

        commit_start = integrated_block.index("function commitRangeFilter()")
        commit_end = integrated_block.find("\nfunction ", commit_start + 1)
        commit_block = integrated_block[commit_start:commit_end]
        self.assertEqual(commit_block.count("applyFilters();"), 1)

        pending_functions = (
            "updateRangeFilterFromInputs",
            "updateRangeFilterFromSliders",
            "clearRangeFilter",
        )
        for function_name in pending_functions:
            function_start = integrated_block.index(f"function {function_name}(")
            function_end = integrated_block.find("\nfunction ", function_start + 1)
            function_block = integrated_block[function_start:function_end]
            self.assertNotIn("updateGeneStructureDomain();", function_block)
            self.assertNotIn("drawGWASPlot(", function_block)

    def test_display_range_slider_executes_overlap_and_timer_logic(self):
        node = shutil.which("node")
        if not node:
            self.skipTest("Node.js is not available for generated JavaScript logic checks")

        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")
        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        def extract_js_function(name):
            start = integrated_block.index(f"function {name}(")
            brace = integrated_block.index("{", start)
            depth = 0
            quote = None
            escaped = False
            for idx in range(brace, len(integrated_block)):
                char = integrated_block[idx]
                if quote:
                    if escaped:
                        escaped = False
                    elif char == "\\":
                        escaped = True
                    elif char == quote:
                        quote = None
                    continue
                if char in ("'", '"', "`"):
                    quote = char
                elif char == "{":
                    depth += 1
                elif char == "}":
                    depth -= 1
                    if depth == 0:
                        return integrated_block[start:idx + 1]
            self.fail(f"Could not extract JavaScript function {name}")

        functions = "\n".join(extract_js_function(name) for name in [
            "getRangeBounds",
            "parseRangeNumber",
            "clampRangeNumber",
            "normalizeRangeValues",
            "readRangeFilter",
            "formatRangeNumber",
            "rangeValueFromClientX",
            "chooseRangeHandle",
            "resolveDraggedRange",
            "canonicalizeRangeFilter",
            "rangeFiltersEqual",
            "syncRangeControls",
            "updateRangePendingState",
            "setPendingRangeFilter",
            "updateRangeFilterFromInputs",
            "updateRangeFilterFromSliders",
            "clearRangeFilter",
            "commitRangeFilter",
            "getAppliedCoordinateDomain",
            "getD3CoordinateDomain",
            "formatCoordinateTick",
            "coordinateToGeneX",
            "clipCoordinateInterval",
            "getGeneStructureGeometry",
            "resetFilters",
            "scheduleLDTriangleRedraw",
        ])
        script = functions + r"""
function assertEqual(actual, expected, label) {
    if (JSON.stringify(actual) !== JSON.stringify(expected)) {
        throw new Error(label + ': expected ' + JSON.stringify(expected) + ', got ' + JSON.stringify(actual));
    }
}
assertEqual(rangeValueFromClientX(0, 0, 100, 1, 101), 1, 'left bound');
assertEqual(rangeValueFromClientX(50, 0, 100, 1, 101), 51, 'midpoint');
assertEqual(rangeValueFromClientX(120, 0, 100, 1, 101), 101, 'right clamp');
assertEqual(chooseRangeHandle(49, 50, 50, 'start'), 'start', 'overlap drag left');
assertEqual(chooseRangeHandle(51, 50, 50, 'start'), 'end', 'overlap drag right');
assertEqual(chooseRangeHandle(50, 50, 50, 'start'), 'end', 'overlap exact toggles handle');
assertEqual(resolveDraggedRange('start', 60, 50, 55), {start: 55, end: 55}, 'start cannot cross end');
assertEqual(resolveDraggedRange('end', 40, 50, 55), {start: 50, end: 50}, 'end cannot cross start');
var elements = {
    mafSlider: { value: '0.05' },
    missingSlider: { value: '0.2' },
    mafValue: { textContent: '0.05' },
    missingValue: { textContent: '0.2' },
    rangeStartInput: { value: '', min: '1', max: '101' },
    rangeEndInput: { value: '', min: '1', max: '101' },
    rangeStartSlider: { value: '1', min: '1', max: '101', style: {}, setAttribute: function() {} },
    rangeEndSlider: { value: '101', min: '1', max: '101', style: {}, setAttribute: function() {} },
    rangeSliderFill: { style: {} },
    rangeSliderLabel: { textContent: '' },
    rangeApplyBtn: { disabled: true },
    rangePendingStatus: { hidden: true }
};
var document = {
    getElementById: function(id) { return elements[id] || null; },
    querySelectorAll: function() { return []; }
};
var initialDisplayRange = {start: 10, end: 90};
var currentFilter = { maf: 0.05, missingRate: 0.2, rangeStart: 10, rangeEnd: 90 };
var pendingRangeFilter = { start: 10, end: 90 };
var regionStart = 1;
var regionEnd = 101;
var gwasLeftMargin = 450;
var geneAreaWidth = 200;
var manualBlacklist = new Set();
var manualFilterHistory = [];
var manualFilterMode = false;
function updateUndoButton() {}
function toggleManualFilter() {}
function scheduleAppliedRangeCentering() {}
var applyCount = 0;
function applyFilters() { applyCount += 1; }

assertEqual(getAppliedCoordinateDomain(), {start: 10, end: 90}, 'applied coordinate domain');
assertEqual(getD3CoordinateDomain({start: 50, end: 50}), [49.5, 50.5], 'single-position D3 domain');
assertEqual(coordinateToGeneX(50, {start: 10, end: 90}), 550, 'coordinate maps into gene plot');
assertEqual(clipCoordinateInterval(1, 20, {start: 10, end: 90}), {start: 10, end: 20}, 'feature interval is clipped');
assertEqual(formatCoordinateTick(50, {start: 10, end: 90}), '50 bp', 'small coordinates use bp labels');
var plusModel = {
    gene_start: 20, gene_end: 90,
    promoter_start: 1, promoter_end: 19,
    strand: '+', exons: [[20, 60]], cds: [[25, 55]]
};
var partialGeometry = getGeneStructureGeometry({start: 10, end: 30}, plusModel);
assertEqual(partialGeometry.gene, {start: 20, end: 30}, 'gene body is clipped');
assertEqual(partialGeometry.promoter, {start: 10, end: 19}, 'promoter is clipped');
assertEqual(partialGeometry.exons, [{interval: {start: 20, end: 30}, cds: [{start: 25, end: 30}]}], 'exon and CDS are clipped');
assertEqual(partialGeometry.threePrimePosition, null, 'off-screen arrow is omitted');
var minusModel = {
    gene_start: 20, gene_end: 90,
    promoter_start: 91, promoter_end: 101,
    strand: '-', exons: [[60, 20]], cds: [[55, 25]]
};
var minusGeometry = getGeneStructureGeometry({start: 10, end: 30}, minusModel);
assertEqual(minusGeometry.ticks.map(function(t) { return Number(t.relativeKb.toFixed(3)); }), [0.091, 0.081, 0.071], 'minus-strand ticks run in reverse');
assertEqual(minusGeometry.exons, [{interval: {start: 20, end: 30}, cds: [{start: 25, end: 30}]}], 'reversed exon coordinates are normalized');
assertEqual(minusGeometry.threePrimePosition, 20, 'minus-strand arrow uses gene start');
var singleGeometry = getGeneStructureGeometry({start: 50, end: 50}, plusModel);
assertEqual(singleGeometry.ticks.map(function(t) { return t.ratio; }), [0.5], 'single-position range has one centered tick');
assertEqual(singleGeometry.gene, {start: 50, end: 50}, 'single-position gene interval remains visible');

elements.rangeStartInput.value = '1';
updateRangeFilterFromInputs('start');
assertEqual(elements.rangeStartInput.value, '1', 'number typing preserves a boundary prefix');
elements.rangeStartInput.value = '20';
updateRangeFilterFromInputs('start');
assertEqual(applyCount, 0, 'number edit stays pending');
assertEqual(currentFilter.rangeStart, 10, 'number edit leaves applied start unchanged');
assertEqual(elements.rangeApplyBtn.disabled, false, 'pending edit enables apply');
commitRangeFilter();
assertEqual(applyCount, 1, 'changed range applies once');
assertEqual(currentFilter.rangeStart, 20, 'commit updates applied start');
elements.rangeEndInput.value = '80';
updateRangeFilterFromInputs('end');
clearRangeFilter();
assertEqual(applyCount, 1, 'clear stays pending');
assertEqual(currentFilter.rangeStart, 20, 'clear leaves applied range unchanged');
commitRangeFilter();
assertEqual(applyCount, 2, 'cleared range applies once');
assertEqual(currentFilter.rangeStart, null, 'clear commit restores full range');
commitRangeFilter();
assertEqual(applyCount, 2, 'unchanged commit is skipped');
elements.rangeStartSlider.value = '30';
elements.rangeEndSlider.value = '101';
updateRangeFilterFromSliders('start');
assertEqual(applyCount, 2, 'slider edit stays pending');
commitRangeFilter();
assertEqual(applyCount, 3, 'slider range applies on explicit commit');
elements.rangeEndInput.value = '80';
updateRangeFilterFromInputs('end');
elements.rangeStartInput.value = '90';
updateRangeFilterFromInputs('start');
assertEqual(pendingRangeFilter, {start: 80, end: 80}, 'crossed input preview');
commitRangeFilter();
assertEqual(
    {start: currentFilter.rangeStart, end: currentFilter.rangeEnd},
    {start: 80, end: 80},
    'commit uses pending preview'
);
resetFilters();
assertEqual(applyCount, 5, 'reset applies once');
assertEqual(
    {start: currentFilter.rangeStart, end: currentFilter.rangeEnd},
    initialDisplayRange,
    'reset restores generated initial window'
);
assertEqual(pendingRangeFilter, initialDisplayRange, 'reset synchronizes pending range');
assertEqual(elements.rangeApplyBtn.disabled, true, 'reset leaves no pending change');
var clearedTimers = [];
var nextTimer = 0;
var ldRedrawTimer = null;
function setTimeout(callback, delay) { nextTimer += 1; return nextTimer; }
function clearTimeout(timer) { clearedTimers.push(timer); }
scheduleLDTriangleRedraw();
scheduleLDTriangleRedraw();
assertEqual(clearedTimers, [1], 'LD redraw timer is coalesced');
assertEqual(ldRedrawTimer, 2, 'latest LD redraw timer is retained');
"""
        completed = subprocess.run(
            [node, "-e", script],
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertEqual(completed.returncode, 0, completed.stderr)

    def test_multi_panel_reset_clears_integrated_display_range(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        multi_start = source.index("def generate_multi_panel_html")
        multi_end = source.index("def generate_ld_triangle_html", multi_start)
        multi_block = source[multi_start:multi_end]

        self.assertIn("applyFilters({{ clearRange: true }});", multi_block)
        self.assertIn("function applyFilters(options)", multi_block)
        self.assertIn("integratedFilters.rangeStart = null", multi_block)
        self.assertIn("integratedFilters.rangeEnd = null", multi_block)

    def test_integrated_report_keeps_gwas_aligned_above_gene_structure_in_main_view(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        main_idx = integrated_block.index('<div class="main-data-section">')
        gwas_idx = integrated_block.index('id="gene-gwas-panel-workbench"')
        gene_svg_idx = integrated_block.index('gene-structure-svg')

        self.assertLess(main_idx, gwas_idx)
        self.assertLess(gwas_idx, gene_svg_idx)
        self.assertIn("gwas_left_margin = gene_area_start", integrated_block)
        self.assertNotIn('data-component="gwas"', integrated_block)
        self.assertNotIn('id="component-gwas"', integrated_block)
        self.assertNotIn('id="gwas-side-slot"', integrated_block)
        self.assertNotIn("moveElementToSlot('gene-gwas-panel-workbench', 'gwas-side-slot')", integrated_block)
        self.assertNotIn("moveFirstMatchToSlot('.gene-gwas-panel', 'gwas-side-slot')", integrated_block)

    def test_integrated_report_constrains_horizontal_overflow_to_content_pane(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("html, body {{ width: 100%; max-width: 100%;", integrated_block)
        self.assertIn("body {{ background: #f5f7fa; padding: 0; overflow: hidden; color: #243044;", integrated_block)
        self.assertIn(".content-wrapper {{ min-width: 0; overflow: auto; height: calc(100vh - 44px); max-height: none; background: #f5f7fa;", integrated_block)
        self.assertIn(".main-data-section {{ max-width: 100%; overflow: auto; border: 1px solid #d8e0ea; border-radius: 8px; background: #ffffff;", integrated_block)
        self.assertIn(".report-sidebar {{ position: fixed; top: 0; right: 0; bottom: 0; width: min(380px, 92vw); z-index: 9000;", integrated_block)
        self.assertIn("background: #ffffff;", integrated_block)
        self.assertNotIn("#071017", integrated_block)
        self.assertNotIn("#0b141e", integrated_block)
        self.assertNotIn("#101b27", integrated_block)
        self.assertIn(".report-shell {{ width: 100%; max-width: 100vw;", integrated_block)
        self.assertIn(".container {{ width: 100%; max-width: 100vw;", integrated_block)
        self.assertIn(".content-wrapper {{ min-width: 0; overflow: auto;", integrated_block)
        self.assertIn(".main-data-section {{ max-width: 100%; overflow: auto;", integrated_block)
        self.assertIn(".header > div {{ min-width: 0;", integrated_block)
        self.assertIn(".base-legend {{ display: flex; flex-wrap: wrap;", integrated_block)
        self.assertIn(".footer > * {{ min-width: 0;", integrated_block)
        self.assertIn(".report-sidebar .pheno-selector", integrated_block)
        self.assertIn("max-width: 100%; box-sizing: border-box;", integrated_block)

    def test_integrated_report_keeps_gene_structure_above_sequence_and_connected(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        gene_idx = integrated_block.index("gene-structure-svg")
        sequence_idx = integrated_block.index('<table class="data-table"')
        score_idx = integrated_block.index('id="haplotype-score-panel"')

        self.assertLess(gene_idx, sequence_idx)
        self.assertLess(sequence_idx, score_idx)
        self.assertIn("js-connector", integrated_block)
        self.assertIn("updateConnectorLines", integrated_block)
        self.assertIn("seq-col-th", integrated_block)
        self.assertIn("seq-site-cell", integrated_block)

    def test_integrated_report_export_includes_sidebar_component_svgs(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        export_start = integrated_block.index("function collectReportExportSVGElements")
        export_end = integrated_block.index("// ==================== 页面初始化", export_start)
        export_block = integrated_block[export_start:export_end]

        self.assertIn("document.querySelector('.report-shell')", export_block)
        self.assertIn("root.querySelectorAll('svg')", export_block)
        self.assertIn("var attrW = parseFloat(svg.getAttribute('width')) || 0;", export_block)
        self.assertIn("var svgElements = collectReportExportSVGElements();", export_block)
        self.assertIn("function withIntrinsicReportZoom(callback)", export_block)
        self.assertIn("return withIntrinsicReportZoom(function()", export_block)
        self.assertNotIn("content.querySelectorAll('svg')", export_block)

    def test_integrated_report_export_uses_fallback_svg_dimensions_for_hidden_panels(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("function getSvgExportSize", integrated_block)
        export_start = integrated_block.index("function getSvgExportSize")
        export_end = integrated_block.index("// ==================== 页面初始化", export_start)
        export_block = integrated_block[export_start:export_end]

        self.assertIn("function prepareReportExportVisuals", export_block)
        self.assertIn("reportSidebar", export_block)
        self.assertIn("export-measure", export_block)
        self.assertIn("var viewBox = svg.viewBox && svg.viewBox.baseVal;", export_block)
        self.assertIn("var size = getSvgExportSize(svgElements[i]);", export_block)
        self.assertIn("totalWidth = Math.max(totalWidth, size.width);", export_block)
        self.assertIn("currentY += size.height + 20;", export_block)

    def test_integrated_report_print_layout_uncrops_scroll_container_and_prints_sidebar_panels(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("@media print", integrated_block)
        self.assertIn(".content-wrapper {{ height: auto; max-height: none; overflow: visible;", integrated_block)
        self.assertIn(".report-sidebar {{ order: 3; position: static;", integrated_block)
        self.assertIn(".report-sidebar .report-component-panel {{ display: block !important;", integrated_block)
        self.assertIn("zoom: 1 !important", integrated_block)
        self.assertIn("window.addEventListener('beforeprint', prepareReportForPrint);", integrated_block)
        self.assertIn("window.addEventListener('afterprint', restoreReportAfterPrint);", integrated_block)

    def test_integrated_report_auto_center_scrolls_content_wrapper(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("function scrollToAppliedRange()", integrated_block)
        self.assertIn("var section = document.querySelector('.main-data-section');", integrated_block)
        self.assertIn("function scheduleAppliedRangeCentering()", integrated_block)
        self.assertIn("scheduleAppliedRangeCentering();", integrated_block)
        self.assertNotIn("window.scrollTo({ left: scrollX", integrated_block)

    def test_summarize_post_gwas_evidence_uses_local_variant_context(self):
        from haplotype_phenotype_analysis import _summarize_post_gwas_evidence

        evidence = _summarize_post_gwas_evidence(
            variant_positions=[90, 110, 206],
            variant_pvalues={90: 1e-3, 110: 1e-6, 206: 0.2},
            variant_info={
                90: {"ref": "A", "alt": "G", "maf": 0.20, "missing_rate": 0.01, "annotation": "promoter"},
                110: {"ref": "C", "alt": "T", "maf": 0.35, "missing_rate": 0.02, "annotation": "missense"},
                206: {"ref": "A", "alt": "<DEL>", "is_sv": True, "maf": 0.08, "missing_rate": 0.05},
            },
            snp_effects={90: "promoter", 110: "missense", 206: "SV"},
            gene_start=100,
            gene_end=200,
            promoter_start=50,
            promoter_end=99,
            flank_bp=5,
            gwas_threshold=1e-5,
        )

        self.assertEqual(evidence["top_marker"]["pos"], 110)
        self.assertEqual(evidence["top_marker"]["candidate_status"], "strict_local_gwas_window")
        self.assertEqual(evidence["candidate_rule"]["status"], "strict_local_gwas_window")
        self.assertEqual(evidence["candidate_rule"]["significant_in_window"], 1)
        self.assertEqual(evidence["variant_type_counts"]["SNP"], 2)
        self.assertEqual(evidence["variant_type_counts"]["SV"], 1)
        self.assertEqual(evidence["position_counts"]["promoter"], 1)
        self.assertEqual(evidence["position_counts"]["body"], 1)
        self.assertEqual(evidence["position_counts"]["outside"], 1)

    def test_script_json_dumps_escapes_script_breakout(self):
        from haplotype_phenotype_analysis import _script_json_dumps

        dumped = _script_json_dumps({
            "annotation": "</script><script>alert(1)</script>",
            "ref": "A&B",
            "alt": "<DEL>",
        })

        self.assertNotIn("</script>", dumped.lower())
        self.assertIn("\\u003c/script\\u003e", dumped)
        self.assertIn("\\u0026", dumped)

    def test_integrated_post_gwas_panel_labels_bound_phenotype(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("Local Candidate Evidence", integrated_block)
        self.assertIn("Evidence phenotype", integrated_block)
        self.assertIn("Switching phenotype updates score, effect, and candidate panels", integrated_block)

    def test_integrated_post_gwas_panel_has_auditable_evidence_layout(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("_render_post_gwas_evidence_table", source)
        self.assertIn("_build_variant_haplotype_bridge", source)
        self.assertIn("_render_evidence_confidence_flags", source)
        self.assertIn("evidence-detail-layout", integrated_block)
        self.assertIn("evidence-detail-disclosure", integrated_block)
        self.assertIn("<details class=\"evidence-detail-disclosure\">", integrated_block)
        self.assertIn("Top Variant Evidence", integrated_block)
        self.assertIn("Variant-Haplotype Bridge", integrated_block)
        self.assertIn("Confidence Flags", integrated_block)
        self.assertIn("{post_gwas_table_html}", integrated_block)
        self.assertIn("{variant_haplotype_bridge_html}", integrated_block)
        self.assertIn("{confidence_flags_html}", integrated_block)

    def test_integrated_report_has_discovery_candidate_panels(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("_build_discovery_candidate_rows", source)
        self.assertIn("_render_discovery_candidate_list", source)
        self.assertIn("_render_score_component_breakdown", source)
        self.assertIn("_render_reliability_population_panel", source)
        self.assertIn("Local Candidate Evidence", integrated_block)
        self.assertIn("discovery-candidate-section", integrated_block)
        self.assertIn("Discovery Candidate List", source)
        self.assertIn("Score Component Breakdown", source)
        self.assertIn("Candidate Key Sites", source)
        self.assertIn("Reliability &amp; Population", source)
        self.assertIn("component-matrix", source)
        self.assertNotIn("component-card", source)
        self.assertIn("{discovery_candidate_list_html}", integrated_block)
        self.assertIn("{top_haplotype_key_sites_html}", integrated_block)
        self.assertIn("{score_component_breakdown_html}", integrated_block)
        self.assertIn("{reliability_population_html}", integrated_block)
        self.assertIn("top-hap-key-site", integrated_block)
        self.assertIn("highlightKeySiteColumns", integrated_block)

    def test_integrated_candidate_panels_are_robust_mode_and_phenotype_aware(self):
        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")

        integrated_start = source.index("def generate_integrated_html")
        integrated_end = source.index("def generate_haplotype_network_html", integrated_start)
        integrated_block = source[integrated_start:integrated_end]

        self.assertIn("_build_discovery_candidate_panel_data", source)
        self.assertIn("var allDiscoveryCandidatePanelData", integrated_block)
        self.assertIn("function updateDiscoveryCandidatePanels", integrated_block)
        self.assertIn("allDiscoveryCandidatePanelData.modes", integrated_block)
        self.assertIn("discovery-candidate-panel-meta", integrated_block)

        phenotype_switch = integrated_block[
            integrated_block.index("function switchPhenotype"):
            integrated_block.index("function getScoreData")
        ]
        self.assertIn("updateDiscoveryCandidatePanels()", phenotype_switch)
        self.assertIn("getDiscoveryCandidatePanel(currentScoreMode, currentPhenotype)", integrated_block)
        self.assertNotIn("function switchScoreMode", integrated_block)
        self.assertNotIn("Switch updates score plot only.", integrated_block)

    def test_render_discovery_candidate_panels_are_literature_free(self):
        from haplotype_phenotype_analysis import (
            _build_discovery_candidate_rows,
            _build_discovery_candidate_panel_data,
            _build_top_haplotype_key_site_rows,
            _render_discovery_candidate_list,
            _render_score_component_breakdown,
            _render_reliability_population_panel,
            _render_top_haplotype_key_sites,
            _select_discovery_candidate_panel,
        )
        import pandas as pd

        hap_sample_df = pd.DataFrame({
            "SampleID": ["S1", "S2", "S3", "S4", "S5"],
            "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap3", "Hap3"],
            "Haplotype_Seq": ["A|G|C", "A|G|C", "C|G|C", "C|G|T", "C|G|T"],
            "Trait": [10.0, 12.0, 30.0, 20.0, 22.0],
            "Population": ["P1", "P2", "P1", "P2", "P2"],
        })
        score_results = {
            "score_mode": "robust_discovery",
            "per_haplotype": {
                "Hap1": {"total": 3.2, "variant_effect": 0.8, "burden": 0.2, "effect_size": 0.6, "sample_reliability": 0.9, "ambiguity_factor": 1.0},
                "Hap2": {"total": 2.5, "variant_effect": 1.0, "burden": 0.7, "effect_size": 0.9, "sample_reliability": 0.4, "ambiguity_factor": 0.5},
                "Hap3": {"total": 1.7, "variant_effect": 0.3, "burden": 0.9, "effect_size": 0.4, "sample_reliability": 0.8, "ambiguity_factor": 1.0},
            },
        }

        rows = _build_discovery_candidate_rows(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            phenotype_col="Trait",
            score_results=score_results,
            effect_data={"Hap2": {"effect": 5.0}},
        )

        self.assertEqual(rows[0]["haplotype"], "Hap1")
        self.assertEqual(rows[0]["rank"], 1)
        self.assertEqual(rows[0]["sample_count"], 2)
        self.assertEqual(rows[0]["population_summary"], "P1=1; P2=1")
        self.assertEqual(rows[0]["reliability_flag"], "warn")

        key_rows = _build_top_haplotype_key_site_rows(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            phenotype_col="Trait",
            score_results=score_results,
            display_positions=[100, 200, 300],
            display_orig_indices=[0, 1, 2],
            variant_info={
                100: {"annotation": "promoter", "maf": 0.40, "missing_rate": 0.00},
                200: {"annotation": "intron", "maf": 0.45, "missing_rate": 0.00},
                300: {"annotation": "missense", "maf": 0.25, "missing_rate": 0.10},
            },
            variant_pvalues={100: 1e-6, 300: 2e-4},
            top_n=5,
        )
        self.assertEqual([row["position"] for row in key_rows], [100, 300])
        self.assertEqual(key_rows[0]["top_haplotype"], "Hap1")
        self.assertEqual(key_rows[0]["top_allele"], "A")
        self.assertEqual(key_rows[0]["other_alleles"], "C=3")
        self.assertEqual(key_rows[0]["top_count"], 2)
        self.assertEqual(key_rows[0]["other_count"], 3)
        self.assertAlmostEqual(key_rows[0]["phenotype_contrast"], -13.0)
        self.assertEqual(key_rows[0]["annotation"], "promoter")
        self.assertEqual(key_rows[0]["variant_type"], "SNP")
        self.assertEqual(key_rows[0]["reliability_flag"], "warn")

        html = (
            _render_discovery_candidate_list(rows)
            + _render_top_haplotype_key_sites(key_rows)
            + _render_score_component_breakdown(rows)
            + _render_reliability_population_panel(rows)
        )
        self.assertIn("Discovery Candidate List", html)
        self.assertIn("Candidate Key Sites", html)
        self.assertIn("data-pos=\"100\"", html)
        self.assertIn("Hap1", html)
        self.assertIn("A vs C=3", html)
        self.assertIn("Score Component Breakdown", html)
        self.assertIn("Reliability &amp; Population", html)
        self.assertIn("component-matrix", html)
        self.assertNotIn("component-card", html)
        self.assertIn("Hap1", html)
        self.assertIn("sample_reliability", html)
        self.assertIn("P1=1", html)
        self.assertNotIn("literature", html.lower())
        self.assertNotIn("published", html.lower())

        panel_data = _build_discovery_candidate_panel_data(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            display_positions=[100, 200, 300],
            display_orig_indices=[0, 1, 2],
            variant_info={100: {"annotation": "promoter"}, 300: {"annotation": "missense"}},
            score_mode_data={
                "current_mode": "robust_discovery",
                "modes": {
                    "default": {
                        "Trait": {
                            "score_mode": "default",
                            "per_haplotype": {
                                "Hap1": {"total": 3.0, "sample_reliability": 0.9},
                                "Hap2": {"total": 1.0, "sample_reliability": 0.9},
                            },
                        },
                    },
                    "robust_discovery": {
                        "Trait": {
                            "score_mode": "robust_discovery",
                            "per_haplotype": {
                                "Hap1": {"total": 1.0, "sample_reliability": 0.9},
                                "Hap2": {"total": 4.0, "sample_reliability": 0.8},
                            },
                        },
                    },
                },
            },
            effect_by_phenotype={"Trait": {"Hap1": {"effect": 1.0}, "Hap2": {"effect": 2.0}}},
            current_phenotype="Trait",
        )
        default_panel = _select_discovery_candidate_panel(panel_data, "default", "Trait")
        robust_panel = _select_discovery_candidate_panel(panel_data, "robust_discovery", "Trait")
        self.assertIn("#1</td><td>Hap1", default_panel["candidate_list_html"])
        self.assertIn("#1</td><td>Hap2", robust_panel["candidate_list_html"])
        self.assertIn("Candidate Key Sites", default_panel["top_haplotype_key_sites_html"])
        self.assertIn("Candidate Key Sites", robust_panel["top_haplotype_key_sites_html"])
        self.assertIn("Score mode: <strong>Original</strong>", default_panel["meta_html"])
        self.assertIn("Score mode: <strong>Robust discovery</strong>", robust_panel["meta_html"])

    def test_discovery_candidate_rows_prefer_anchor_candidates_when_available(self):
        from haplotype_phenotype_analysis import _build_discovery_candidate_rows
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "S1", "Hap_Name": "HighTotalBackground", "Haplotype_Seq": "C|A", "Trait": 1.0},
            {"SampleID": "S2", "Hap_Name": "HighTotalBackground", "Haplotype_Seq": "C|A", "Trait": 2.0},
            {"SampleID": "S3", "Hap_Name": "FunctionalAnchor", "Haplotype_Seq": "C|INS_837", "Trait": 3.0},
            {"SampleID": "S4", "Hap_Name": "FunctionalAnchor", "Haplotype_Seq": "C|INS_837", "Trait": 4.0},
        ])
        score_results = {
            "per_haplotype": {
                "HighTotalBackground": {"total": 5.0, "site_weighted": 0.1},
                "FunctionalAnchor": {"total": 0.5, "site_weighted": 0.2},
            },
            "anchor_haplotype_candidates": {
                "phenotype_free": True,
                "candidates": {
                    "HighTotalBackground": {
                        "haplotype": "HighTotalBackground",
                        "anchor_score": 0.0,
                        "anchor_positions": [],
                    },
                    "FunctionalAnchor": {
                        "haplotype": "FunctionalAnchor",
                        "anchor_score": 0.9,
                        "anchor_max_site_score": 0.9,
                        "anchor_positions": [200],
                    },
                },
            },
        }

        rows = _build_discovery_candidate_rows(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            phenotype_col="Trait",
            score_results=score_results,
        )

        self.assertEqual("FunctionalAnchor", rows[0]["haplotype"])
        self.assertEqual("anchor", rows[0]["rank_basis"])
        self.assertEqual([200], rows[0]["anchor_positions"])
        self.assertAlmostEqual(0.9, rows[0]["total"])
        self.assertAlmostEqual(0.5, rows[0]["raw_total"])

    def test_compute_haplotype_scores_uses_full_positions_not_display_subset(self):
        from haplotype_phenotype_analysis import ReportGenerator
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "S1", "Hap_Name": "Hap1", "Haplotype_Seq": "A|G|C|T", "Trait": 50.0},
            {"SampleID": "S2", "Hap_Name": "Hap1", "Haplotype_Seq": "A|G|C|T", "Trait": 51.0},
            {"SampleID": "S3", "Hap_Name": "Hap2", "Haplotype_Seq": "A|G|C|A", "Trait": 52.0},
            {"SampleID": "S4", "Hap_Name": "Hap2", "Haplotype_Seq": "A|G|C|A", "Trait": 53.0},
            {"SampleID": "S5", "Hap_Name": "Hap3", "Haplotype_Seq": "C|G|T|T", "Trait": 35.0},
            {"SampleID": "S6", "Hap_Name": "Hap3", "Haplotype_Seq": "C|G|T|T", "Trait": 36.0},
        ])
        variant_info = {
            100: {"ref": "C", "alt": "A", "annotation": "promoter", "maf": 0.45, "missing_rate": 0.0},
            200: {"ref": "G", "alt": "A", "annotation": "promoter", "maf": 0.05, "missing_rate": 0.0},
            260: {"ref": "T", "alt": "C", "annotation": "intron", "maf": 0.45, "missing_rate": 0.0},
            1000: {"ref": "T", "alt": "A", "annotation": "intron", "maf": 0.30, "missing_rate": 0.0},
        }
        reporter = ReportGenerator(output_dir=tempfile.mkdtemp())

        result = reporter.compute_haplotype_scores(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200, 260, 1000],
            region_start=100,
            region_end=1000,
            phenotype_col="Trait",
            gene_start=250,
            gene_end=5000,
            variant_info=variant_info,
            score_mode="robust_discovery",
        )

        score_data = result["all_score_data"]["Trait"]
        functional = score_data.get("functional_haplotype_groups") or {}
        self.assertIn(260, functional.get("functional_positions", []))
        self.assertIn("A|G|C", functional.get("groups", {}))

    def test_display_position_mapping_preserves_variant_info_sequence_order(self):
        from haplotype_phenotype_analysis import ReportGenerator
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "S1", "Hap_Name": "Hap1", "Haplotype_Seq": "A|C|DEL_837", "Trait": 0.0},
            {"SampleID": "S2", "Hap_Name": "Hap1", "Haplotype_Seq": "A|C|DEL_837", "Trait": 0.0},
            {"SampleID": "S3", "Hap_Name": "Hap2", "Haplotype_Seq": "A|C|INS_837", "Trait": 1.0},
            {"SampleID": "S4", "Hap_Name": "Hap2", "Haplotype_Seq": "A|C|INS_837", "Trait": 1.0},
        ])
        variant_info = {
            100: {"ref": "A", "alt": "G", "annotation": "base_snp", "maf": 0.5, "missing_rate": 0.0},
            300: {"ref": "C", "alt": "T", "annotation": "base_snp", "maf": 0.5, "missing_rate": 0.0},
            200: {
                "ref": "DEL_837",
                "alt": "INS_837",
                "annotation": "diagnostic_marker",
                "maf": 0.5,
                "missing_rate": 0.0,
            },
        }
        reporter = ReportGenerator(output_dir=tempfile.mkdtemp())

        result = reporter.compute_haplotype_scores(
            hap_sample_df=hap_sample_df,
            variant_positions=[200],
            region_start=100,
            region_end=300,
            phenotype_col="Trait",
            gene_start=100,
            gene_end=300,
            variant_info=variant_info,
            score_mode="robust_discovery",
        )

        self.assertEqual([200], result["display_positions"])

    def test_haplotype_allele_display_formats_structural_tokens(self):
        from haplotype_phenotype_analysis import _format_haplotype_allele_for_display

        self.assertEqual(("-837bp", "-"), _format_haplotype_allele_for_display("DEL_837"))
        self.assertEqual(("+837bp", "+"), _format_haplotype_allele_for_display("INS_837"))
        self.assertEqual(("+837bp", "+"), _format_haplotype_allele_for_display("A" * 837))
        self.assertEqual(("A", "A"), _format_haplotype_allele_for_display("A"))

    def test_compute_haplotype_scores_ld_uses_tokenized_indel_alleles(self):
        from haplotype_phenotype_analysis import ReportGenerator
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "S1", "Hap_Name": "Hap1", "Haplotype_Seq": "LONGREF|A|T", "Trait": 10.0},
            {"SampleID": "S2", "Hap_Name": "Hap1", "Haplotype_Seq": "LONGREF|A|T", "Trait": 11.0},
            {"SampleID": "S3", "Hap_Name": "Hap2", "Haplotype_Seq": "DEL_7|G|C", "Trait": 20.0},
            {"SampleID": "S4", "Hap_Name": "Hap2", "Haplotype_Seq": "DEL_7|G|C", "Trait": 21.0},
        ])
        variant_info = {
            100: {"ref": "LONGREF", "alt": "DEL_7", "annotation": "promoter", "maf": 0.5, "missing_rate": 0.0},
            200: {"ref": "A", "alt": "G", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
            300: {"ref": "T", "alt": "C", "annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
        }
        reporter = ReportGenerator(output_dir=tempfile.mkdtemp())

        result = reporter.compute_haplotype_scores(
            hap_sample_df=hap_sample_df,
            variant_positions=[100, 200, 300],
            region_start=100,
            region_end=300,
            phenotype_col="Trait",
            gene_start=150,
            gene_end=300,
            variant_info=variant_info,
            score_mode="robust_discovery",
        )

        self.assertEqual(result["display_positions"], [100, 200, 300])
        ld_matrix = result["ld_r2_matrix"]
        self.assertEqual(len(ld_matrix), 3)
        self.assertGreater(ld_matrix[1][2], 0.9)

    def test_key_site_contrast_is_site_specific(self):
        from haplotype_phenotype_analysis import _build_top_haplotype_key_site_rows
        import pandas as pd

        hap_sample_df = pd.DataFrame({
            "SampleID": [f"S{i}" for i in range(1, 9)],
            "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap2", "Hap3", "Hap3", "Hap4", "Hap4"],
            "Haplotype_Seq": [
                "A|G|C", "A|G|C",
                "T|G|C", "T|G|C",
                "A|T|C", "A|T|C",
                "A|G|T", "A|G|T",
            ],
            "Trait": [10.0, 12.0, 30.0, 32.0, 80.0, 82.0, 15.0, 17.0],
        })
        score_results = {
            "per_haplotype": {
                "Hap1": {"total": 5.0},
                "Hap2": {"total": 2.0},
                "Hap3": {"total": 1.0},
                "Hap4": {"total": 0.5},
            }
        }

        rows = _build_top_haplotype_key_site_rows(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            phenotype_col="Trait",
            score_results=score_results,
            display_positions=[100, 200, 300],
            display_orig_indices=[0, 1, 2],
            variant_info={
                100: {"annotation": "intron", "maf": 0.2, "missing_rate": 0.0},
                200: {"annotation": "intron", "maf": 0.2, "missing_rate": 0.0},
                300: {"annotation": "intron", "maf": 0.2, "missing_rate": 0.0},
            },
            top_n=3,
        )

        by_pos = {row["position"]: row for row in rows}
        self.assertAlmostEqual(by_pos[100]["other_mean"], 31.0)
        self.assertAlmostEqual(by_pos[100]["phenotype_contrast"], -20.0)
        self.assertAlmostEqual(by_pos[200]["other_mean"], 81.0)
        self.assertAlmostEqual(by_pos[200]["phenotype_contrast"], -70.0)
        self.assertAlmostEqual(by_pos[300]["other_mean"], 16.0)
        self.assertAlmostEqual(by_pos[300]["phenotype_contrast"], -5.0)
        self.assertEqual([row["position"] for row in rows], [200, 100, 300])

    def test_robust_key_site_priority_uses_attention_not_local_phenotype_signal(self):
        from haplotype_phenotype_analysis import _build_top_haplotype_key_site_rows
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "T1", "Hap_Name": "Top", "Haplotype_Seq": "A|G", "Trait": 10.0},
            {"SampleID": "T2", "Hap_Name": "Top", "Haplotype_Seq": "A|G", "Trait": 10.0},
            {"SampleID": "P1", "Hap_Name": "LocalPOnly", "Haplotype_Seq": "T|G", "Trait": 0.0},
            {"SampleID": "P2", "Hap_Name": "LocalPOnly", "Haplotype_Seq": "T|G", "Trait": 0.0},
            {"SampleID": "A1", "Hap_Name": "AttentionSite", "Haplotype_Seq": "A|C", "Trait": 9.0},
            {"SampleID": "A2", "Hap_Name": "AttentionSite", "Haplotype_Seq": "A|C", "Trait": 9.0},
        ])
        score_results = {
            "score_mode": "robust_discovery",
            "per_haplotype": {
                "Top": {"total": 5.0},
                "LocalPOnly": {"total": 1.0},
                "AttentionSite": {"total": 1.0},
            },
            "site_weights": [
                {"position": 100, "attention_prior_weight": 0.0, "attention_cluster_id": ""},
                {"position": 200, "attention_prior_weight": 0.9, "attention_cluster_id": "cluster_1"},
            ],
        }

        rows = _build_top_haplotype_key_site_rows(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            phenotype_col="Trait",
            score_results=score_results,
            display_positions=[100, 200],
            display_orig_indices=[0, 1],
            variant_info={
                100: {"annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
                200: {"annotation": "intron", "maf": 0.5, "missing_rate": 0.0},
            },
            variant_pvalues={100: 1e-12, 200: 0.5},
            top_n=2,
        )

        self.assertEqual([row["position"] for row in rows], [200, 100])
        self.assertEqual(rows[0]["attention_cluster_id"], "cluster_1")
        self.assertLess(abs(rows[0]["phenotype_contrast"]), abs(rows[1]["phenotype_contrast"]))

    def test_key_site_rows_use_anchor_top_haplotype_when_available(self):
        from haplotype_phenotype_analysis import _build_top_haplotype_key_site_rows
        import pandas as pd

        hap_sample_df = pd.DataFrame([
            {"SampleID": "B1", "Hap_Name": "HighTotalBackground", "Haplotype_Seq": "C|A", "Trait": 1.0},
            {"SampleID": "B2", "Hap_Name": "HighTotalBackground", "Haplotype_Seq": "C|A", "Trait": 2.0},
            {"SampleID": "K1", "Hap_Name": "FunctionalAnchor", "Haplotype_Seq": "C|INS_837", "Trait": 3.0},
            {"SampleID": "K2", "Hap_Name": "FunctionalAnchor", "Haplotype_Seq": "C|INS_837", "Trait": 4.0},
        ])
        score_results = {
            "per_haplotype": {
                "HighTotalBackground": {"total": 5.0},
                "FunctionalAnchor": {"total": 0.5},
            },
            "anchor_haplotype_candidates": {
                "top_candidate": {
                    "haplotype": "FunctionalAnchor",
                    "anchor_score": 0.9,
                    "anchor_positions": [200],
                }
            },
        }

        rows = _build_top_haplotype_key_site_rows(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            phenotype_col="Trait",
            score_results=score_results,
            display_positions=[100, 200],
            display_orig_indices=[0, 1],
            variant_info={200: {"annotation": "INS", "maf": 0.1, "missing_rate": 0.0}},
            top_n=3,
        )

        self.assertEqual("FunctionalAnchor", rows[0]["top_haplotype"])
        self.assertEqual(200, rows[0]["position"])
        self.assertEqual("INS_837", rows[0]["top_allele"])

    def test_discovery_panel_uses_phenotype_specific_variant_pvalues(self):
        from haplotype_phenotype_analysis import _build_discovery_candidate_panel_data
        import pandas as pd

        hap_sample_df = pd.DataFrame({
            "SampleID": [f"S{i}" for i in range(1, 7)],
            "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap2", "Hap3", "Hap3"],
            "Haplotype_Seq": ["A|G", "A|G", "T|G", "T|G", "A|T", "A|T"],
            "TraitA": [10.0, 11.0, 30.0, 31.0, 12.0, 13.0],
            "TraitB": [50.0, 51.0, 52.0, 53.0, 90.0, 91.0],
        })
        score_mode_data = {
            "current_mode": "default",
            "modes": {
                "default": {
                    "TraitA": {
                        "per_haplotype": {
                            "Hap1": {"total": 5.0},
                            "Hap2": {"total": 1.0},
                            "Hap3": {"total": 0.5},
                        }
                    },
                    "TraitB": {
                        "per_haplotype": {
                            "Hap1": {"total": 5.0},
                            "Hap2": {"total": 1.0},
                            "Hap3": {"total": 0.5},
                        }
                    },
                }
            },
        }

        panel_data = _build_discovery_candidate_panel_data(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            score_mode_data=score_mode_data,
            current_phenotype="TraitA",
            display_positions=[100, 200],
            display_orig_indices=[0, 1],
            variant_info={
                100: {"annotation": "promoter", "maf": 0.2, "missing_rate": 0.0},
                200: {"annotation": "promoter", "maf": 0.2, "missing_rate": 0.0},
            },
            variant_pvalues={100: 0.5, 200: 0.5},
            variant_pvalues_by_phenotype={
                "TraitA": {100: 1e-8, 200: 0.5},
                "TraitB": {100: 0.5, 200: 1e-8},
            },
        )

        trait_a_html = panel_data["modes"]["default"]["TraitA"]["top_haplotype_key_sites_html"]
        trait_b_html = panel_data["modes"]["default"]["TraitB"]["top_haplotype_key_sites_html"]
        self.assertIn('data-pos="100"', trait_a_html)
        self.assertIn("1.00e-08", trait_a_html)
        self.assertIn('data-pos="200"', trait_b_html)
        self.assertIn("1.00e-08", trait_b_html)

    def test_key_site_renderer_sanitizes_flag_class_and_string_positions(self):
        from haplotype_phenotype_analysis import (
            _build_discovery_candidate_panel_data,
            _render_top_haplotype_key_sites,
        )
        import pandas as pd

        html = _render_top_haplotype_key_sites([{
            "rank": 1,
            "position": "SV_marker_1",
            "top_haplotype": "Hap1",
            "top_allele": "DEL",
            "other_alleles": "REF=4",
            "top_count": 2,
            "other_count": 4,
            "phenotype_contrast": 3.0,
            "pvalue": 0.01,
            "attention_prior_weight": 0.91,
            "attention_cluster_id": "cluster_A",
            "variant_type": "SV",
            "annotation": "sv",
            "reliability_flag": 'pass" onclick="alert(1)',
            "flag_note": "stable",
        }])

        self.assertIn('data-pos="SV_marker_1"', html)
        self.assertIn("<th>Attention</th>", html)
        self.assertIn("0.91", html)
        self.assertIn("cluster_A", html)
        self.assertIn('class="key-site-flag muted"', html)
        self.assertNotIn("onclick", html)

        hap_sample_df = pd.DataFrame({
            "SampleID": ["S1", "S2", "S3", "S4"],
            "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap2"],
            "Haplotype_Seq": ["A", "A", "T", "T"],
            "Trait": [10.0, 11.0, 20.0, 21.0],
        })
        panel_data = _build_discovery_candidate_panel_data(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            score_mode_data={
                "current_mode": "default",
                "modes": {
                    "default": {
                        "Trait": {
                            "per_haplotype": {
                                "Hap1": {"total": 5.0},
                                "Hap2": {"total": 1.0},
                            }
                        }
                    }
                },
            },
            current_phenotype="Trait",
            display_positions=["SV_marker_1"],
            display_orig_indices=[0],
        )

        self.assertEqual(
            panel_data["modes"]["default"]["Trait"]["top_haplotype_key_site_positions"],
            ["SV_marker_1"],
        )

    def test_discovery_candidate_flags_render_as_wrappable_chips(self):
        from haplotype_phenotype_analysis import _render_discovery_candidate_list

        html = _render_discovery_candidate_list([{
            "rank": 1,
            "haplotype": "Hap1",
            "total": 1.2,
            "sample_count": 1,
            "phenotype_mean": 10.5,
            "effect": 2.0,
            "reliability_flag": "warn",
            "flag_note": "low n, low reliability, ambiguity penalty",
            "sequence": "A|G|T",
        }])

        self.assertIn('class="candidate-flag-set"', html)
        self.assertIn(">low n<", html)
        self.assertIn(">low reliability<", html)
        self.assertIn(">ambiguity penalty<", html)
        self.assertNotIn(">low n, low reliability, ambiguity penalty<", html)

        source = Path("haplotype_phenotype_analysis.py").read_text(encoding="utf-8")
        self.assertIn(".candidate-table th:nth-child(7), .candidate-table td:nth-child(7)", source)
        self.assertIn(".candidate-flag-set", source)
        self.assertIn("white-space: normal", source)

    def test_key_site_priority_penalizes_low_reliability_sites(self):
        from haplotype_phenotype_analysis import _build_top_haplotype_key_site_rows
        import pandas as pd

        top_rows = [
            {
                "SampleID": f"T{i}",
                "Hap_Name": "Hap1",
                "Haplotype_Seq": "A|G",
                "Trait": 10.0 + (i % 2),
            }
            for i in range(20)
        ]
        other_rows = [
            {
                "SampleID": f"O{i}",
                "Hap_Name": "Hap2",
                "Haplotype_Seq": "T|C",
                "Trait": 30.0 + (i % 2),
            }
            for i in range(20)
        ]
        hap_sample_df = pd.DataFrame(top_rows + other_rows)

        rows = _build_top_haplotype_key_site_rows(
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            phenotype_col="Trait",
            score_results={
                "per_haplotype": {
                    "Hap1": {"total": 5.0},
                    "Hap2": {"total": 1.0},
                }
            },
            display_positions=[100, 200],
            display_orig_indices=[0, 1],
            variant_info={
                100: {"annotation": "promoter", "maf": 0.01, "missing_rate": 0.45},
                200: {"annotation": "promoter", "maf": 0.25, "missing_rate": 0.0},
            },
            variant_pvalues={100: 1e-20, 200: 1e-6},
            top_n=2,
        )

        by_pos = {row["position"]: row for row in rows}
        self.assertEqual(by_pos[100]["reliability_flag"], "warn")
        self.assertEqual(by_pos[200]["reliability_flag"], "pass")
        self.assertLess(by_pos[100]["priority_score"], by_pos[200]["priority_score"])
        self.assertEqual(rows[0]["position"], 200)

    def test_evidence_table_includes_ld_and_annotation_columns(self):
        from haplotype_phenotype_analysis import _render_post_gwas_evidence_table

        html = _render_post_gwas_evidence_table({
            "records": [{
                "pos": 110,
                "variant_type": "SNP",
                "position_context": "body",
                "pvalue": 1e-6,
                "minus_log10_p": 6.0,
                "r2_to_lead": 1.0,
                "annotation": "missense",
            }]
        }, chrom="chr1")

        self.assertIn("LD r2", html)
        self.assertIn("Annotation", html)
        self.assertIn("chr1:110", html)
        self.assertIn("missense", html)

    def test_evidence_table_preserves_small_pvalues(self):
        from haplotype_phenotype_analysis import _render_post_gwas_evidence_table

        html = _render_post_gwas_evidence_table({
            "records": [{
                "pos": 110,
                "variant_type": "SNP",
                "position_context": "body",
                "pvalue": 0.004,
                "minus_log10_p": 2.39794,
                "r2_to_lead": 0.75,
                "annotation": "intron",
            }]
        }, chrom="chr1")

        self.assertIn("4.00e-03", html)
        self.assertNotIn(">0<", html)

    def test_integrated_post_gwas_summary_escapes_dynamic_text(self):
        from haplotype_phenotype_analysis import ReportGenerator
        import pandas as pd

        with tempfile.TemporaryDirectory() as tmp:
            df = pd.DataFrame({
                "SampleID": ["S1", "S2", "S3", "S4"],
                "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap2"],
                "Haplotype_Seq": ["A", "A", "G", "G"],
                "Trait<script>": [10, 11, 20, 21],
            })
            generator = ReportGenerator(output_dir=tmp)
            generator.generate_integrated_html(
                hap_sample_df=df,
                effect_results={"grand_mean": 15.5, "haplotype_effects": []},
                variant_positions=[100],
                region_start=50,
                region_end=150,
                phenotype_col="Trait<script>",
                gene_start=80,
                gene_end=150,
                promoter_start=50,
                promoter_end=79,
                chrom="chr<script>",
                gene_id="GeneX",
                variant_info={100: {"ref": "A", "alt": "G", "annotation": "<b>bad</b>"}},
                variant_pvalues={100: 1e-6},
                snp_effects={100: "<img src=x onerror=alert(1)>"},
            )
            html = (Path(tmp) / "GeneX.html").read_text(encoding="utf-8")

        self.assertNotIn("Trait<script>", html)
        self.assertNotIn("chr<script>", html)
        self.assertNotIn("<img src=x onerror=alert(1)>", html)
        self.assertIn("Trait&lt;script&gt;", html)
        self.assertIn("chr&lt;script&gt;", html)
        self.assertIn("&lt;img src=x onerror=alert(1)&gt;", html)

    def test_variant_haplotype_bridge_groups_marker_alleles_by_haplotype(self):
        from haplotype_phenotype_analysis import _build_variant_haplotype_bridge
        import pandas as pd

        hap_sample_df = pd.DataFrame({
            "SampleID": ["S1", "S2", "S3", "S4"],
            "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap3"],
            "Haplotype_Seq": ["A|C", "A|C", "G|C", "G|T"],
            "Trait": [10.0, 12.0, 20.0, 30.0],
        })
        bridge = _build_variant_haplotype_bridge(
            records=[{"pos": 100, "pvalue": 1e-6}],
            display_positions=[100, 200],
            display_orig_indices=[0, 1],
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            top_haps=["Hap1", "Hap2", "Hap3"],
            phenotype_col="Trait",
            effect_data={"Hap2": {"effect": 4.0}, "Hap3": {"effect": 8.0}},
            score_results={"per_haplotype": {"Hap2": {"overall_score": 0.7}}},
        )

        self.assertIn("pos", bridge[0])
        self.assertEqual(bridge[0]["allele"], "G")
        self.assertEqual(bridge[0]["haplotypes"], ["Hap2", "Hap3"])
        self.assertEqual(bridge[0]["sample_count"], 2)
        self.assertEqual(bridge[0]["best_score_haplotype"], "Hap2")

    def test_variant_haplotype_bridge_uses_total_score_field(self):
        from haplotype_phenotype_analysis import _build_variant_haplotype_bridge
        import pandas as pd

        hap_sample_df = pd.DataFrame({
            "SampleID": ["S1", "S2", "S3"],
            "Hap_Name": ["Hap1", "Hap2", "Hap3"],
            "Haplotype_Seq": ["A", "G", "G"],
            "Trait": [10.0, 20.0, 30.0],
        })

        bridge = _build_variant_haplotype_bridge(
            records=[{"pos": 100, "ref": "A", "alt": "G", "pvalue": 1e-6}],
            display_positions=[100],
            display_orig_indices=[0],
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            top_haps=["Hap1", "Hap2", "Hap3"],
            phenotype_col="Trait",
            score_results={"per_haplotype": {"Hap2": {"total": 0.7}, "Hap3": {"total": 0.5}}},
        )

        self.assertEqual(bridge[0]["best_score_haplotype"], "Hap2")
        self.assertEqual(bridge[0]["best_score"], 0.7)

    def test_variant_haplotype_bridge_prefers_annotated_alt_allele(self):
        from haplotype_phenotype_analysis import _build_variant_haplotype_bridge
        import pandas as pd

        hap_sample_df = pd.DataFrame({
            "SampleID": ["S1", "S2", "S3", "S4", "S5", "S6"],
            "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap2", "Hap3", "Hap3"],
            "Haplotype_Seq": ["A", "A", "G", "G", "G", "G"],
            "Trait": [10, 11, 18, 19, 25, 26],
        })

        bridge = _build_variant_haplotype_bridge(
            records=[{"pos": 100, "ref": "A", "alt": "G", "pvalue": 1e-6}],
            display_positions=[100],
            display_orig_indices=[0],
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            top_haps=["Hap1", "Hap2", "Hap3"],
            phenotype_col="Trait",
        )

        self.assertEqual(bridge[0]["allele"], "G")
        self.assertEqual(bridge[0]["reference_allele"], "A")
        self.assertEqual(bridge[0]["haplotypes"], ["Hap2", "Hap3"])

    def test_variant_haplotype_bridge_handles_heterozygous_and_multi_alt_calls(self):
        from haplotype_phenotype_analysis import _build_variant_haplotype_bridge
        import pandas as pd

        hap_sample_df = pd.DataFrame({
            "SampleID": ["S1", "S2", "S3", "S4"],
            "Hap_Name": ["Hap1", "Hap2", "Hap3", "Hap4"],
            "Haplotype_Seq": ["A/A", "A/G", "T/A", "N"],
            "Trait": [10, 20, 30, 40],
        })

        bridge = _build_variant_haplotype_bridge(
            records=[{"pos": 100, "ref": "A", "alt": "G,T", "pvalue": 1e-6}],
            display_positions=[100],
            display_orig_indices=[0],
            hap_sample_df=hap_sample_df,
            hap_col="Hap_Name",
            top_haps=["Hap1", "Hap2", "Hap3", "Hap4"],
            phenotype_col="Trait",
        )

        self.assertEqual(bridge[0]["allele"], "G,T")
        self.assertEqual(bridge[0]["haplotypes"], ["Hap2", "Hap3"])
        self.assertEqual(bridge[0]["sample_count"], 2)

    def test_variant_haplotype_bridge_labels_bound_phenotype_and_mode(self):
        from haplotype_phenotype_analysis import _render_variant_haplotype_bridge

        html = _render_variant_haplotype_bridge(
            [{
                "pos": 100,
                "allele": "G",
                "haplotypes": ["Hap2"],
                "sample_count": 2,
                "phenotype_mean": 20.0,
                "best_score_haplotype": "Hap2",
                "best_score": 0.8,
                "strongest_effect_haplotype": "Hap2",
                "strongest_effect": 5.0,
            }],
            chrom="chr1",
            phenotype_col="TraitA",
            score_mode="robust_discovery",
        )

        self.assertIn("Bound to TraitA", html)
        self.assertIn("robust_discovery", html)
        self.assertIn("does not change when the phenotype selector is switched", html)

    def test_score_mode_collection_prefers_score_json_over_stale_report_mode(self):
        from haplotype_phenotype_analysis import ReportGenerator

        with tempfile.TemporaryDirectory() as tmp:
            generator = ReportGenerator(output_dir=str(Path(tmp) / "GeneA__robust_discovery"))
            generator.score_mode = "default"

            collected = generator._collect_score_mode_data({
                "Trait": {
                    "score_mode": "robust_discovery",
                    "per_sample": [],
                    "per_haplotype": {},
                }
            })

            self.assertEqual(collected["current_mode"], "robust_discovery")
            self.assertIn("robust_discovery", collected["modes"])
            self.assertNotIn("default", collected["modes"])

    def test_robust_discovery_penalizes_tiny_ambiguous_haplotypes(self):
        import pandas as pd
        from haplotype_phenotype_analysis import HaplotypeScorer

        rows = []
        rows.extend(
            {
                "SampleID": f"H1_{i}",
                "Hap_Name": "Hap1",
                "Haplotype_Seq": "A|G|C",
                "TGW_mean": 41.0 + (0.1 if i % 2 else -0.1),
            }
            for i in range(40)
        )
        rows.extend(
            {
                "SampleID": f"H2_{i}",
                "Hap_Name": "Hap2",
                "Haplotype_Seq": "C|G|C",
                "TGW_mean": 39.0 + (0.1 if i % 2 else -0.1),
            }
            for i in range(35)
        )
        rows.extend(
            {
                "SampleID": f"H5_{i}",
                "Hap_Name": "Hap5",
                "Haplotype_Seq": "C/A|G|C",
                "TGW_mean": 40.6 + (0.1 if i % 2 else -0.1),
            }
            for i in range(3)
        )
        hap_sample_df = pd.DataFrame(rows)
        variant_positions = [291759689, 291760677, 291761315]
        variant_info = {
            291759689: {"ref": "C", "alt": "A", "maf": 0.48, "annotation": "promoter"},
            291760677: {"ref": "G", "alt": "A", "maf": 0.01, "annotation": "promoter"},
            291761315: {"ref": "T", "alt": "C", "maf": 0.47, "annotation": "promoter"},
        }
        gwas_data = [
            {"pos": 291759689, "pvalue": 1e-8},
            {"pos": 291760677, "pvalue": 1e-4},
            {"pos": 291761315, "pvalue": 1e-6},
        ]
        effect_results = {
            "haplotype_effects": [
                {"haplotype": "Hap1", "cohens_d": 0.9, "n_samples": 40},
                {"haplotype": "Hap2", "cohens_d": 0.4, "n_samples": 35},
                {"haplotype": "Hap5", "cohens_d": 2.5, "n_samples": 3},
            ]
        }

        robust = HaplotypeScorer(
            hap_sample_df,
            variant_positions,
            variant_info=variant_info,
            snp_effects={pos: "promoter" for pos in variant_positions},
            gwas_data=gwas_data,
            effect_results=effect_results,
            phenotype_col="TGW_mean",
            score_mode="robust_discovery",
        ).score_all()
        default = HaplotypeScorer(
            hap_sample_df,
            variant_positions,
            variant_info=variant_info,
            snp_effects={pos: "promoter" for pos in variant_positions},
            gwas_data=gwas_data,
            effect_results=effect_results,
            phenotype_col="TGW_mean",
        ).score_all()

        self.assertEqual(robust["score_mode"], "robust_discovery")
        self.assertNotIn("sample_reliability", default["per_haplotype"]["Hap5"])
        self.assertLess(robust["per_haplotype"]["Hap5"]["sample_reliability"], 0.2)
        self.assertLess(robust["per_haplotype"]["Hap5"]["ambiguity_factor"], 1.0)
        self.assertGreater(
            robust["per_haplotype"]["Hap1"]["total"],
            robust["per_haplotype"]["Hap5"]["total"],
        )

    def test_expected_decrease_direction_reorders_raw_high_phenotype_score(self):
        import pandas as pd
        from haplotype_phenotype_analysis import HaplotypeScorer

        rows = []
        rows.extend(
            {
                "SampleID": f"D1b_{i}",
                "Hap_Name": "HapD1b",
                "Haplotype_Seq": "B1a|D1b",
                "PlantHeight": 82.0 + (0.1 if i % 2 else -0.1),
            }
            for i in range(80)
        )
        rows.extend(
            {
                "SampleID": f"WT_{i}",
                "Hap_Name": "HapWT",
                "Haplotype_Seq": "B1a|Rht-D1a",
                "PlantHeight": 98.0 + (0.1 if i % 2 else -0.1),
            }
            for i in range(50)
        )
        hap_sample_df = pd.DataFrame(rows)
        effect_results = {
            "haplotype_effects": [
                {"haplotype": "HapD1b", "cohens_d": 0.0, "n_samples": 80},
                {"haplotype": "HapWT", "cohens_d": 2.0, "n_samples": 50},
            ]
        }

        scored = HaplotypeScorer(
            hap_sample_df,
            variant_positions=[],
            effect_results=effect_results,
            phenotype_col="PlantHeight",
            expected_direction="decreases_trait",
        ).score_all()

        self.assertEqual(scored["score_axis"], "total")
        self.assertGreater(
            scored["per_haplotype"]["HapWT"]["total"],
            scored["per_haplotype"]["HapD1b"]["total"],
        )
        self.assertGreater(
            scored["per_haplotype"]["HapD1b"]["directional_total"],
            scored["per_haplotype"]["HapWT"]["directional_total"],
        )

    def test_pca_plot_handles_two_variant_haplotypes(self):
        import pandas as pd
        from haplotype_phenotype_analysis import ReportGenerator

        with tempfile.TemporaryDirectory() as tmp:
            hap_sample_df = pd.DataFrame({
                "SampleID": ["S1", "S2", "S3", "S4"],
                "Hap_Name": ["Hap1", "Hap1", "Hap2", "Hap3"],
                "Haplotype_Seq": ["T|G", "T|G", "T|A", "C|G"],
                "TGW_mean": [39.7, 40.1, 31.4, 39.6],
            })
            generator = ReportGenerator(output_dir=tmp)
            stdout = StringIO()

            with redirect_stdout(stdout):
                html_path = generator.generate_pca_plot_html(
                    hap_sample_df,
                    phenotype_col="TGW_mean",
                )

            self.assertTrue(Path(html_path).exists())
            html = Path(html_path).read_text(encoding="utf-8")
            self.assertIn("Principal Component Analysis", html)
            self.assertIn("PC2", html)
            self.assertIn("PCA分析图已保存", stdout.getvalue())

    def test_build_database_from_wide_marker_matrix(self):
        from star_gene_data import build_database_from_marker_matrix

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            marker_path = tmp_path / "markers.tsv"
            phenotype_path = tmp_path / "phenotypes.tsv"
            out_root = tmp_path / "db"

            marker_path.write_text(
                "SampleID\tqHKW1_100\tqHKW1_200\n"
                "S1\tA\tT\n"
                "S2\tA\tT\n"
                "S3\tG\tT\n"
                "S4\tG\tC\n",
                encoding="utf-8",
            )
            phenotype_path.write_text(
                "SampleID\tHKW\n"
                "S1\t25.1\n"
                "S2\t25.3\n"
                "S3\t30.4\n"
                "S4\t31.0\n",
                encoding="utf-8",
            )

            db_dir = build_database_from_marker_matrix(
                marker_matrix=marker_path,
                phenotype_table=phenotype_path,
                output_root=out_root,
                target_id="qHKW1",
                chrom="chr1",
                start=100,
                end=200,
                phenotype_columns=["HKW"],
                marker_columns=["qHKW1_100", "qHKW1_200"],
                marker_positions={"qHKW1_100": 100, "qHKW1_200": 200},
                expected_direction="increases_trait",
            )

            self.assertTrue((db_dir / "gene_info.json").exists())
            self.assertTrue((db_dir / "haplotype_data.csv").exists())
            self.assertTrue((db_dir / "haplotype_samples.csv").exists())
            self.assertTrue((db_dir / "phenotype_data.csv").exists())
            self.assertTrue((db_dir / "variant_info.csv").exists())

            hap_samples = (db_dir / "haplotype_samples.csv").read_text(encoding="utf-8")
            self.assertIn("S1,A|T,Hap1", hap_samples)
            self.assertIn("S3,G|T,Hap2", hap_samples)
            phenotype_data = (db_dir / "phenotype_data.csv").read_text(encoding="utf-8")
            self.assertIn("HKW", phenotype_data)

    def test_build_database_filters_configured_phenotype_missing_values(self):
        from star_gene_data import build_database_from_marker_matrix

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            marker_path = tmp_path / "markers.tsv"
            phenotype_path = tmp_path / "phenotypes.tsv"
            out_root = tmp_path / "db"

            marker_path.write_text(
                "SampleID\tm1\n"
                "S1\tNN\n"
                "S2\tAA\n"
                "S3\tAA\n",
                encoding="utf-8",
            )
            phenotype_path.write_text(
                "SampleID\tHKW\n"
                "S1\t-999.0\n"
                "S2\t25.1\n"
                "S3\t25.3\n",
                encoding="utf-8",
            )

            db_dir = build_database_from_marker_matrix(
                marker_matrix=marker_path,
                phenotype_table=phenotype_path,
                output_root=out_root,
                target_id="qHKW1",
                chrom="chr1",
                start=100,
                end=100,
                phenotype_columns=["HKW"],
                marker_columns=["m1"],
                marker_positions={"m1": 100},
                phenotype_missing_values=["-999"],
            )

            phenotype_data = (db_dir / "phenotype_data.csv").read_text(encoding="utf-8")
            hap_samples = (db_dir / "haplotype_samples.csv").read_text(encoding="utf-8")

            self.assertNotIn("S1", phenotype_data)
            self.assertNotIn("-999.0", phenotype_data)
            self.assertIn("S2", phenotype_data)
            self.assertNotIn("S1", hap_samples)

    def test_build_database_from_marker_matrix_reads_long_tsv_with_truncated_sniffer_sample(self):
        from star_gene_data import build_database_from_marker_matrix

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            marker_path = tmp_path / "markers.tsv"
            phenotype_path = tmp_path / "phenotypes.tsv"
            out_root = tmp_path / "db"
            marker_cols = [f"chr6A_{237732717 + i}" for i in range(190)]

            with marker_path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f, delimiter="\t", lineterminator="\n")
                writer.writerow(["SampleID", *marker_cols])
                writer.writerow(["S1", *(["A"] * len(marker_cols))])
                writer.writerow(["S2", *(["T"] * len(marker_cols))])
                writer.writerow(["S3", *(["A"] * len(marker_cols))])
                writer.writerow(["S4", *(["T"] * len(marker_cols))])
            with phenotype_path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f, delimiter="\t", lineterminator="\n")
                writer.writerow(["SampleID", "Trait"])
                writer.writerow(["S1", 1.0])
                writer.writerow(["S2", 2.0])
                writer.writerow(["S3", 1.5])
                writer.writerow(["S4", 2.5])

            db_dir = build_database_from_marker_matrix(
                marker_matrix=marker_path,
                phenotype_table=phenotype_path,
                output_root=out_root,
                target_id="LongHeaderGene",
                chrom="1",
                start=1000,
                end=2000,
                phenotype_columns=["Trait"],
                min_haplotype_count=1,
            )

            self.assertTrue((db_dir / "haplotype_samples.csv").exists())
            with (db_dir / "variant_info.csv").open() as f:
                variant_rows = list(csv.DictReader(f))
            self.assertEqual(len(variant_rows), len(marker_cols))

    def test_build_database_infers_sv_length_from_marker_id(self):
        from star_gene_data import build_database_from_marker_matrix

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            marker_path = tmp_path / "markers.tsv"
            phenotype_path = tmp_path / "phenotypes.tsv"
            out_root = tmp_path / "db"
            marker_name = "chr1_27976568_27994848_deletion_18280"

            marker_path.write_text(
                f"SampleID\t{marker_name}\n"
                "S1\tNN\n"
                "S2\tAA\n",
                encoding="utf-8",
            )
            phenotype_path.write_text(
                "SampleID\tHKW\n"
                "S1\t22\n"
                "S2\t25\n",
                encoding="utf-8",
            )

            db_dir = build_database_from_marker_matrix(
                marker_matrix=marker_path,
                phenotype_table=phenotype_path,
                output_root=out_root,
                target_id="qHKW1",
                chrom="1",
                start=27976568,
                end=27994848,
                phenotype_columns=["HKW"],
                marker_columns=[marker_name],
                marker_positions={marker_name: 28892142},
            )

            variant_info = (db_dir / "variant_info.csv").read_text(encoding="utf-8")

            self.assertIn("-18280", variant_info)
            self.assertIn("True", variant_info)
            self.assertIn("SV", variant_info)

    def test_scan_maizego_sv_candidates_filters_by_window_and_length(self):
        from star_gene_data import scan_maizego_sv_candidates

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            matrix_path = tmp_path / "SV.386014.hmp.txt"
            matrix_path.write_text(
                "rs#\talleles\tchrom\tpos\tstrand\tassembly#\tcenter\tprotLSID\tassayLSID\tpanelLSID\tQCcode\tS1\tS2\tS3\tS4\n"
                "chr1_30450000_30458900_insertion_8900\tA/T\t1\t30454500\t+\tNA\tNA\tNA\tNA\tNA\tNA\tAA\tAA\tNN\tTT\n"
                "chr1_30452000_30452100_insertion_100\tA/T\t1\t30452000\t+\tNA\tNA\tNA\tNA\tNA\tNA\tAA\tAA\tAA\tAA\n"
                "chr1_40000000_40008900_insertion_8900\tA/T\t1\t40004500\t+\tNA\tNA\tNA\tNA\tNA\tNA\tAA\tNN\tNN\tTT\n",
                encoding="utf-8",
            )

            candidates = scan_maizego_sv_candidates(
                matrix_paths=[matrix_path],
                chrom="1",
                window_start=30440000,
                window_end=30540000,
                length_min=8500,
                length_max=9500,
            )

            self.assertEqual(len(candidates), 1)
            candidate = candidates[0]
            self.assertEqual(candidate["marker"], "chr1_30450000_30458900_insertion_8900")
            self.assertEqual(candidate["variant_type"], "insertion")
            self.assertEqual(candidate["sv_length"], 8900)
            self.assertEqual(candidate["counts"], "AA:2;NN:1;TT:1")
            self.assertEqual(candidate["source_path"], str(matrix_path))

    def test_scan_maizego_sv_catalogue_candidates_reports_records_without_samples(self):
        from star_gene_data import scan_maizego_sv_catalogue_candidates

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            catalogue_path = tmp_path / "svs.final.ms.txt"
            catalogue_path.write_text(
                "chr1\t30450000\t30458900\tinsertion\t8900\t+\tMo17.chr1.64\t29961002\t29969902\t295670\t181512\t324959\t0.558569\tACGT\n"
                "chr1\t30452000\t30452100\tinsertion\t100\t+\tMo17.chr1.64\t29961002\t29961102\t295670\t181512\t324959\t0.558569\tAC\n"
                "chr1\t40000000\t40008900\tdeletion\t8900\t+\tMo17.chr1.64\t39961002\t39961002\t295670\t181512\t324959\t0.558569\tACGT\n",
                encoding="utf-8",
            )

            candidates = scan_maizego_sv_catalogue_candidates(
                catalogue_paths=[catalogue_path],
                chrom="1",
                window_start=30440000,
                window_end=30540000,
                length_min=8500,
                length_max=9500,
            )

            self.assertEqual(len(candidates), 1)
            candidate = candidates[0]
            self.assertEqual(candidate["chrom"], "1")
            self.assertEqual(candidate["start"], 30450000)
            self.assertEqual(candidate["end"], 30458900)
            self.assertEqual(candidate["variant_type"], "insertion")
            self.assertEqual(candidate["sv_length"], 8900)
            self.assertEqual(candidate["sample_genotype_columns"], 0)
            self.assertEqual(candidate["has_sample_genotypes"], "False")
            self.assertEqual(candidate["source_path"], str(catalogue_path))

    def test_extract_maizego_marker_matrix_transposes_selected_record(self):
        from star_gene_data import extract_maizego_marker_matrix

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            matrix_path = tmp_path / "SV.386014.hmp.txt"
            output_path = tmp_path / "qHKW1_marker.tsv"
            marker = "chr1_30450000_30458900_insertion_8900"
            matrix_path.write_text(
                "rs#\talleles\tchrom\tpos\tstrand\tassembly#\tcenter\tprotLSID\tassayLSID\tpanelLSID\tQCcode\tS1\tS2\tS3\n"
                f"{marker}\tA/T\t1\t30454500\t+\tNA\tNA\tNA\tNA\tNA\tNA\tAA\tNN\tTT\n",
                encoding="utf-8",
            )

            written = extract_maizego_marker_matrix(
                matrix_path=matrix_path,
                marker_id=marker,
                output_path=output_path,
            )

            self.assertEqual(written, output_path)
            matrix_text = output_path.read_text(encoding="utf-8")
            self.assertEqual(
                matrix_text,
                f"SampleID\t{marker}\nS1\tAA\nS2\tNN\nS3\tTT\n",
            )

    def test_complete_database_can_supply_coordinates_and_phenotypes(self):
        from star_gene_validation import StarGeneValidator

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_root = tmp_path / "star_gene_database"
            results_root = tmp_path / "star_gene_results"
            db_dir = db_root / "paper1" / "qHKW1"
            db_dir.mkdir(parents=True)

            (db_dir / "gene_info.json").write_text(
                '{"gene_id":"qHKW1","chrom":"1","start":100,"end":200}',
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\nAA,2,Hap1,AA\nNN,2,Hap2,NN\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\nS1,AA,Hap1\nS2,AA,Hap1\nS3,NN,Hap2\nS4,NN,Hap2\n",
                encoding="utf-8",
            )
            (db_dir / "phenotype_data.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name,HKW\nS1,AA,Hap1,25\nS2,AA,Hap1,26\nS3,NN,Hap2,22\nS4,NN,Hap2,23\n",
                encoding="utf-8",
            )

            paper = {
                "paper_id": "paper1",
                "short_name": "p1",
                "local_expected_paths": [
                    {
                        "key": "target_variant_matrix",
                        "path": "missing/markers.tsv",
                        "format": "unknown",
                        "required": True,
                    },
                    {
                        "key": "phenotype_table",
                        "path": "missing/phenotype.tsv",
                        "format": "phenotype_table",
                        "required": True,
                    },
                ],
            }
            target = {
                "target_id": "qHKW1",
                "gene_or_locus": "qHKW1",
                "requires_coordinate_resolution": True,
                "coordinates": None,
                "traits": [{"name": "hundred-kernel weight", "local_columns": ["HKW"]}],
            }

            validator = StarGeneValidator(
                manifest={"papers": [paper]},
                database_root=db_root,
                results_root=results_root,
                check_only=False,
            )
            check = validator.check_target(paper, target)
            coords = validator.resolve_target_coordinates(target, check["database_path"])

            self.assertEqual(check["status"], "ready_for_analysis")
            self.assertIn("database_complete", check["notes"])
            self.assertIn("database_coordinates:qHKW1=1:100-200", check["notes"])
            self.assertIn("HKW", check["phenotype_columns"])
            self.assertEqual(coords, ("1", 100, 200))

    def test_required_database_source_blocks_substitute_marker_database(self):
        from star_gene_validation import StarGeneValidator

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_root = tmp_path / "star_gene_database"
            results_root = tmp_path / "star_gene_results"
            db_dir = db_root / "paper1" / "qHKW1"
            db_dir.mkdir(parents=True)

            (db_dir / "gene_info.json").write_text(
                '{"gene_id":"qHKW1","chrom":"1","start":100,"end":200,"source":"marker_matrix"}',
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\nAA,2,Hap1,AA\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\nS1,AA,Hap1\n",
                encoding="utf-8",
            )
            (db_dir / "phenotype_data.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name,HKW\nS1,AA,Hap1,25\n",
                encoding="utf-8",
            )

            paper = {"paper_id": "paper1", "short_name": "p1", "local_expected_paths": []}
            target = {
                "target_id": "qHKW1",
                "gene_or_locus": "qHKW1",
                "requires_coordinate_resolution": True,
                "coordinates": None,
                "required_database_source": "maizego_paper_marker",
                "traits": [{"name": "hundred-kernel weight", "local_columns": ["HKW"]}],
            }

            validator = StarGeneValidator(
                manifest={"papers": [paper]},
                database_root=db_root,
                results_root=results_root,
                check_only=True,
            )
            check = validator.check_target(paper, target)

            self.assertEqual(check["status"], "unsupported_input_format_for_analysis")
            self.assertTrue(any(note.startswith("database_source_mismatch:") for note in check["notes"]))

    def test_literature_audit_matches_top_haplotype_for_single_variant(self):
        from star_gene_literature_audit import run_literature_audit

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db"
            result_dir = tmp_path / "result"
            db_dir.mkdir()
            result_dir.mkdir()
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,len_diff,is_sv,maf,missing_rate,annotation,marker_id\n"
                "100,A,G,0,False,0.5,0.0,promoter,m1\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\n"
                "G,2,Hap1,G\n"
                "A,1,Hap2,A\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\n"
                "S1,G,Hap1\n"
                "S2,G,Hap1\n"
                "S3,A,Hap2\n",
                encoding="utf-8",
            )
            (result_dir / "haplotype_scores.json").write_text(
                json.dumps({
                    "Trait": {
                        "per_haplotype": {
                            "Hap1": {"total": 4.2},
                            "Hap2": {"total": 1.1},
                        },
                        "per_sample": [
                            {"sample_id": "S1", "haplotype": "Hap1"},
                            {"sample_id": "S2", "haplotype": "Hap1"},
                            {"sample_id": "S3", "haplotype": "Hap2"},
                        ],
                    }
                }),
                encoding="utf-8",
            )

            records = run_literature_audit(
                paper={"paper_id": "paper1"},
                target={
                    "target_id": "Gene1",
                    "literature_variants": [
                        {
                            "variant_name": "KnownSNP",
                            "chrom": "chr1",
                            "position": 100,
                            "marker_id": "m1",
                            "expected_allele": "G",
                            "variant_class": "promoter_snp",
                            "source_note": "unit test",
                        }
                    ],
                },
                database_path=db_dir,
                result_path=result_dir,
            )

            self.assertEqual(len(records), 1)
            row = records[0]
            self.assertEqual(row["record_type"], "variant")
            self.assertEqual(row["variant_name"], "KnownSNP")
            self.assertTrue(row["covered_in_database"])
            self.assertTrue(row["segregating_in_current_samples"])
            self.assertEqual(row["carrier_count"], 2)
            self.assertEqual(row["top_scored_haplotype"], "Hap1")
            self.assertEqual(row["top_haplotype_sample_count"], 2)
            self.assertEqual(row["top_haplotype_carrier_count"], 2)
            self.assertTrue(row["top_haplotype_contains_expected"])
            self.assertTrue(row["top_haplotype_exact_expected"])
            self.assertEqual(row["validation_status"], "matched_top_haplotype")
            self.assertTrue((result_dir / "literature_variant_audit.csv").exists())
            self.assertTrue((result_dir / "literature_variant_audit.json").exists())

    def test_literature_audit_reports_directional_top_match(self):
        from star_gene_literature_audit import run_literature_audit

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db"
            result_dir = tmp_path / "result"
            db_dir.mkdir()
            result_dir.mkdir()
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,len_diff,is_sv,maf,missing_rate,annotation,marker_id\n"
                "1,D1a,D1b,0,False,0.5,0.0,functional_marker,Rht-D1\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\n"
                "D1a,20,HapTall,D1a\n"
                "D1b,30,HapShort,D1b\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\n"
                "S1,D1a,HapTall\n"
                "S2,D1b,HapShort\n",
                encoding="utf-8",
            )
            (result_dir / "haplotype_scores.json").write_text(
                json.dumps({"Trait": {"per_haplotype": {"HapTall": {"total": 5.0}, "HapShort": {"total": 1.0}}}}),
                encoding="utf-8",
            )
            (result_dir.parent.parent / "validation_summary.csv").parent.mkdir(parents=True, exist_ok=True)
            (result_dir.parent.parent / "validation_summary.csv").write_text(
                "target_id,result_path,directional_top_haplotype,directional_top_haplotype_score,directional_top_haplotype_sample_count\n"
                f"RhtTest,{result_dir},HapShort,1.0,30\n",
                encoding="utf-8",
            )

            records = run_literature_audit(
                paper={"paper_id": "paper1"},
                target={
                    "target_id": "RhtTest",
                    "literature_variants": [
                        {
                            "variant_name": "Rht-D1b",
                            "marker_id": "Rht-D1",
                            "expected_allele": "D1b",
                        }
                    ],
                },
                database_path=db_dir,
                result_path=result_dir,
            )

            row = records[0]
            self.assertEqual(row["validation_status"], "present_but_not_top")
            self.assertEqual(row["directional_top_haplotype"], "HapShort")
            self.assertTrue(row["directional_top_haplotype_exact_expected"])
            self.assertEqual(row["directional_validation_status"], "matched_directional_top_haplotype")

    def test_literature_audit_distinguishes_heterozygous_top_variant_match(self):
        from star_gene_literature_audit import run_literature_audit

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db"
            result_dir = tmp_path / "result"
            db_dir.mkdir()
            result_dir.mkdir()
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,len_diff,is_sv,maf,missing_rate,annotation,marker_id\n"
                "100,C,A,0,False,0.5,0.0,promoter,m1\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\n"
                "A,3,Hap1,A\n"
                "C/A,2,Hap2,C/A\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\n"
                "S1,A,Hap1\n"
                "S2,A,Hap1\n"
                "S3,A,Hap1\n"
                "S4,C/A,Hap2\n"
                "S5,C/A,Hap2\n",
                encoding="utf-8",
            )
            (result_dir / "haplotype_scores.json").write_text(
                json.dumps({
                    "Trait": {
                        "per_haplotype": {
                            "Hap1": {"total": 2.0},
                            "Hap2": {"total": 5.0},
                        }
                    }
                }),
                encoding="utf-8",
            )

            records = run_literature_audit(
                paper={"paper_id": "paper1"},
                target={
                    "target_id": "Gene1",
                    "literature_variants": [
                        {
                            "variant_name": "KnownSNP",
                            "marker_id": "m1",
                            "expected_allele": "A",
                        }
                    ],
                },
                database_path=db_dir,
                result_path=result_dir,
            )

            row = records[0]
            self.assertTrue(row["top_haplotype_contains_expected"])
            self.assertFalse(row["top_haplotype_exact_expected"])
            self.assertEqual(row["exact_matching_haplotypes"], "Hap1:3")
            self.assertEqual(row["validation_status"], "contained_in_top_haplotype_not_exact")

    def test_literature_audit_reports_missing_and_nonsegregating_variants(self):
        from star_gene_literature_audit import run_literature_audit

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db"
            result_dir = tmp_path / "result"
            db_dir.mkdir()
            result_dir.mkdir()
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,len_diff,is_sv,maf,missing_rate,annotation,marker_id\n"
                "100,A,G,0,False,0.0,0.0,promoter,m1\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\nA,3,Hap1,A\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\nS1,A,Hap1\nS2,A,Hap1\nS3,A,Hap1\n",
                encoding="utf-8",
            )

            records = run_literature_audit(
                paper={"paper_id": "paper1"},
                target={
                    "target_id": "Gene1",
                    "literature_variants": [
                        {
                            "variant_name": "ExpectedMissingAllele",
                            "chrom": "chr1",
                            "position": 100,
                            "marker_id": "m1",
                            "expected_allele": "G",
                        },
                        {
                            "variant_name": "AbsentMarker",
                            "chrom": "chr1",
                            "position": 200,
                            "marker_id": "m2",
                            "expected_allele": "T",
                        },
                    ],
                },
                database_path=db_dir,
                result_path=result_dir,
            )

            by_name = {row["variant_name"]: row for row in records}
            self.assertEqual(by_name["ExpectedMissingAllele"]["validation_status"], "present_not_segregating")
            self.assertTrue(by_name["ExpectedMissingAllele"]["covered_in_database"])
            self.assertFalse(by_name["ExpectedMissingAllele"]["segregating_in_current_samples"])
            self.assertEqual(by_name["ExpectedMissingAllele"]["carrier_count"], 0)
            self.assertEqual(by_name["ExpectedMissingAllele"]["allele_counts"], "A:3")
            self.assertEqual(by_name["AbsentMarker"]["validation_status"], "missing_from_database")
            self.assertFalse(by_name["AbsentMarker"]["covered_in_database"])

    def test_literature_audit_does_not_validate_monomorphic_expected_allele(self):
        from star_gene_literature_audit import run_literature_audit

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db"
            result_dir = tmp_path / "result"
            db_dir.mkdir()
            result_dir.mkdir()
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,len_diff,is_sv,maf,missing_rate,annotation,marker_id\n"
                "100,G,A,0,False,0.0,0.0,promoter,m1\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\nG,3,Hap1,G\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\nS1,G,Hap1\nS2,G,Hap1\nS3,G,Hap1\n",
                encoding="utf-8",
            )
            (result_dir / "haplotype_scores.json").write_text(
                json.dumps({"Trait": {"per_haplotype": {"Hap1": {"total": 5.0}}}}),
                encoding="utf-8",
            )

            records = run_literature_audit(
                paper={"paper_id": "paper1"},
                target={
                    "target_id": "Gene1",
                    "literature_variants": [
                        {
                            "variant_name": "MonomorphicExpected",
                            "marker_id": "m1",
                            "expected_allele": "G",
                        }
                    ],
                },
                database_path=db_dir,
                result_path=result_dir,
            )

            row = records[0]
            self.assertEqual(row["carrier_count"], 3)
            self.assertFalse(row["segregating_in_current_samples"])
            self.assertTrue(row["top_haplotype_contains_expected"])
            self.assertEqual(row["validation_status"], "present_not_segregating")

    def test_literature_audit_reports_group_haplotype_status(self):
        from star_gene_literature_audit import run_literature_audit

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db"
            result_dir = tmp_path / "result"
            db_dir.mkdir()
            result_dir.mkdir()
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,len_diff,is_sv,maf,missing_rate,annotation,marker_id\n"
                "100,A,G,0,False,0.5,0.0,promoter,m1\n"
                "200,C,T,0,False,0.5,0.0,promoter,m2\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\n"
                "G|T,2,Hap1,G|T\n"
                "G|C,1,Hap2,G|C\n"
                "A|T,1,Hap3,A|T\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\n"
                "S1,G|T,Hap1\n"
                "S2,G|T,Hap1\n"
                "S3,G|C,Hap2\n"
                "S4,A|T,Hap3\n",
                encoding="utf-8",
            )
            (result_dir / "haplotype_scores.json").write_text(
                json.dumps({
                    "Trait": {
                        "per_haplotype": {
                            "Hap1": {"total": 2.0},
                            "Hap2": {"total": 5.0},
                            "Hap3": {"total": 1.0},
                        }
                    }
                }),
                encoding="utf-8",
            )

            records = run_literature_audit(
                paper={"paper_id": "paper1"},
                target={
                    "target_id": "Gene1",
                    "literature_haplotypes": [
                        {
                            "haplotype_name": "PublishedPair",
                            "expected_markers": {"m1": "G", "m2": "T"},
                            "source_note": "unit test pair",
                        }
                    ],
                },
                database_path=db_dir,
                result_path=result_dir,
            )

            self.assertEqual(len(records), 1)
            row = records[0]
            self.assertEqual(row["record_type"], "haplotype")
            self.assertEqual(row["variant_name"], "PublishedPair")
            self.assertTrue(row["covered_in_database"])
            self.assertTrue(row["segregating_in_current_samples"])
            self.assertEqual(row["carrier_count"], 2)
            self.assertEqual(row["top_scored_haplotype"], "Hap2")
            self.assertFalse(row["top_haplotype_contains_expected"])
            self.assertEqual(row["validation_status"], "present_but_not_top")

    def test_literature_audit_reports_functional_group_haplotype_status(self):
        from star_gene_literature_audit import run_literature_audit

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db"
            result_dir = tmp_path / "result"
            db_dir.mkdir()
            result_dir.mkdir()
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,len_diff,is_sv,maf,missing_rate,annotation,marker_id\n"
                "100,C,A,0,False,0.5,0.0,promoter,m1\n"
                "200,G,A,0,False,0.02,0.0,promoter,m2\n"
                "300,T,C,0,False,0.5,0.0,intron,bg1\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\n"
                "A|G|T,2,Hap1,A|G|T\n"
                "A|G|C,2,Hap2,A|G|C\n"
                "C|G|T,2,Hap3,C|G|T\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\n"
                "S1,A|G|T,Hap1\n"
                "S2,A|G|T,Hap1\n"
                "S3,A|G|C,Hap2\n"
                "S4,A|G|C,Hap2\n"
                "S5,C|G|T,Hap3\n"
                "S6,C|G|T,Hap3\n",
                encoding="utf-8",
            )
            (result_dir / "haplotype_scores.json").write_text(
                json.dumps({
                    "Trait": {
                        "per_haplotype": {
                            "Hap1": {"total": 2.0},
                            "Hap2": {"total": 1.8},
                            "Hap3": {"total": 5.0},
                        },
                        "functional_haplotype_groups": {
                            "functional_positions": [100, 200],
                            "groups": {
                                "A|G": {
                                    "functional_sequence": "A|G",
                                    "sample_count": 4,
                                    "rank_score": 1.7,
                                    "mean_score": 1.9,
                                    "mean_phenotype": 50.0,
                                },
                                "C|G": {
                                    "functional_sequence": "C|G",
                                    "sample_count": 2,
                                    "rank_score": 1.2,
                                    "mean_score": 5.0,
                                    "mean_phenotype": 35.0,
                                },
                            },
                            "top_group": {
                                "functional_sequence": "A|G",
                                "sample_count": 4,
                                "rank_score": 1.7,
                                "mean_score": 1.9,
                                "mean_phenotype": 50.0,
                            },
                        },
                    }
                }),
                encoding="utf-8",
            )

            records = run_literature_audit(
                paper={"paper_id": "paper1"},
                target={
                    "target_id": "Gene1",
                    "literature_haplotypes": [
                        {
                            "haplotype_name": "PublishedFunctional",
                            "expected_markers": {"m1": "A", "m2": "G"},
                            "source_note": "unit test functional group",
                        }
                    ],
                },
                database_path=db_dir,
                result_path=result_dir,
            )

            row = records[0]
            self.assertEqual(row["top_scored_haplotype"], "Hap3")
            self.assertEqual(row["validation_status"], "present_but_not_top")
            self.assertEqual(row["top_functional_group"], "A|G")
            self.assertTrue(row["top_functional_group_exact_expected"])
            self.assertEqual(row["functional_group_validation_status"], "matched_top_functional_group")

    def test_literature_audit_requires_exact_top_haplotype_for_group_match(self):
        from star_gene_literature_audit import run_literature_audit

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db"
            result_dir = tmp_path / "result"
            db_dir.mkdir()
            result_dir.mkdir()
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,len_diff,is_sv,maf,missing_rate,annotation,marker_id\n"
                "100,C,A,0,False,0.5,0.0,promoter,m1\n"
                "200,G,A,0,False,0.5,0.0,promoter,m2\n"
                "300,T,C,0,False,0.5,0.0,promoter,m3\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\n"
                "A|G|C,3,Hap1,A|G|C\n"
                "C/A|G|C,2,Hap2,C/A|G|C\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\n"
                "S1,A|G|C,Hap1\n"
                "S2,A|G|C,Hap1\n"
                "S3,A|G|C,Hap1\n"
                "S4,C/A|G|C,Hap2\n"
                "S5,C/A|G|C,Hap2\n",
                encoding="utf-8",
            )
            (result_dir / "haplotype_scores.json").write_text(
                json.dumps({
                    "Trait": {
                        "per_haplotype": {
                            "Hap1": {"total": 2.0},
                            "Hap2": {"total": 5.0},
                        }
                    }
                }),
                encoding="utf-8",
            )

            records = run_literature_audit(
                paper={"paper_id": "paper1"},
                target={
                    "target_id": "Gene1",
                    "literature_haplotypes": [
                        {
                            "haplotype_name": "PublishedAGC",
                            "expected_markers": {"m1": "A", "m2": "G", "m3": "C"},
                            "source_note": "unit test strict haplotype",
                        }
                    ],
                },
                database_path=db_dir,
                result_path=result_dir,
            )

            row = records[0]
            self.assertTrue(row["top_haplotype_contains_expected"])
            self.assertFalse(row["top_haplotype_exact_expected"])
            self.assertEqual(row["exact_matching_haplotypes"], "Hap1:3")
            self.assertEqual(row["validation_status"], "contained_in_top_haplotype_not_exact")

    def test_marker_database_cli(self):
        from build_star_gene_database import main as build_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            marker_path = tmp_path / "markers.tsv"
            phenotype_path = tmp_path / "phenotypes.tsv"
            out_root = tmp_path / "db"
            marker_path.write_text(
                "SampleID\tm1\nS1\tA\nS2\tG\n",
                encoding="utf-8",
            )
            phenotype_path.write_text(
                "SampleID\tHKW\nS1\t25\nS2\t30\n",
                encoding="utf-8",
            )

            stdout = StringIO()
            with redirect_stdout(stdout):
                rc = build_main([
                    "--marker-matrix", str(marker_path),
                    "--phenotype-table", str(phenotype_path),
                    "--output-root", str(out_root),
                    "--target-id", "qHKW1",
                    "--chrom", "chr1",
                    "--start", "100",
                    "--end", "100",
                    "--phenotype-column", "HKW",
                    "--marker-column", "m1",
                    "--marker-position", "m1=100",
                ])

            self.assertEqual(rc, 0)
            self.assertTrue((out_root / "qHKW1" / "gene_info.json").exists())
            self.assertIn("Built star-gene database", stdout.getvalue())

    def test_maize_qhkw1_prepare_cli_reports_missing_full_sv_package(self):
        from prepare_maize2019_qhkw1_paper_genotype import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            stdout = StringIO()
            with redirect_stdout(stdout):
                rc = prepare_main(["--sv-root", str(tmp_path)])

            self.assertEqual(rc, 2)
            output = stdout.getvalue()
            self.assertIn("SV.386014.zip", output)
            self.assertIn("paper genotype package is missing", output)

    def test_maize_qhkw1_prepare_ignores_non_matrix_text_files(self):
        from prepare_maize2019_qhkw1_paper_genotype import find_candidate_matrices

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            (tmp_path / "paper.txt").write_text("This is article text, not genotype data.\n", encoding="utf-8")
            matrix_path = tmp_path / "MS.step1.org.txt"
            matrix_path.write_text(
                "rs#\talleles\tchrom\tpos\tstrand\tassembly#\tcenter\tprotLSID\tassayLSID\tpanelLSID\tQCcode\tS1\n"
                "chr1_30450000_30458900_insertion_8900\tA/T\t1\t30454500\t+\tNA\tNA\tNA\tNA\tNA\tNA\tAA\n",
                encoding="utf-8",
            )

            self.assertEqual(find_candidate_matrices(tmp_path), [matrix_path])

    def test_maize_qhkw1_prepare_reports_catalogue_candidates_but_keeps_blocked(self):
        from prepare_maize2019_qhkw1_paper_genotype import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            sv_root = tmp_path / "maizego"
            extracted_dir = sv_root / "SV.386014" / "SV.386014"
            extracted_dir.mkdir(parents=True)
            (extracted_dir / "svs.final.ms.txt").write_text(
                "chr1\t30450000\t30458900\tinsertion\t8900\t+\tMo17.chr1.64\t29961002\t29969902\t295670\t181512\t324959\t0.558569\tACGT\n",
                encoding="utf-8",
            )
            matrix_candidate_output = tmp_path / "matrix_candidates.tsv"
            catalogue_output = tmp_path / "catalogue_candidates.tsv"

            stdout = StringIO()
            with redirect_stdout(stdout):
                rc = prepare_main([
                    "--sv-root", str(sv_root),
                    "--candidate-output", str(matrix_candidate_output),
                    "--catalogue-output", str(catalogue_output),
                ])

            self.assertEqual(rc, 3)
            output = stdout.getvalue()
            self.assertIn("Candidate matrix files: 0", output)
            self.assertIn("SV catalogue candidate records: 1", output)
            self.assertIn("not sample-level genotype matrices", output)
            self.assertTrue(matrix_candidate_output.exists())
            catalogue_text = catalogue_output.read_text(encoding="utf-8")
            self.assertIn("sample_genotype_columns", catalogue_text)
            self.assertIn("chr1", catalogue_text)

    def test_maize_qhkw1_prepare_does_not_reextract_existing_catalogues(self):
        from prepare_maize2019_qhkw1_paper_genotype import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            sv_root = tmp_path / "maizego"
            extracted_dir = sv_root / "SV.386014" / "SV.386014"
            extracted_dir.mkdir(parents=True)
            (extracted_dir / "svs.final.ms.txt").write_text(
                "chr1\t30450000\t30458900\tinsertion\t8900\t+\tMo17.chr1.64\t29961002\t29969902\t295670\t181512\t324959\t0.558569\tACGT\n",
                encoding="utf-8",
            )
            invalid_zip = sv_root / "SV.386014.zip"
            invalid_zip.write_text("not a zip file\n", encoding="utf-8")

            stdout = StringIO()
            with redirect_stdout(stdout):
                rc = prepare_main([
                    "--sv-root", str(sv_root),
                    "--sv-zip", str(invalid_zip),
                    "--candidate-output", str(tmp_path / "matrix_candidates.tsv"),
                    "--catalogue-output", str(tmp_path / "catalogue_candidates.tsv"),
                ])

            self.assertEqual(rc, 3)
            self.assertIn("SV catalogue candidate records: 1", stdout.getvalue())

    def test_maize_qhkw1_prepare_builds_paper_marker_database(self):
        from prepare_maize2019_qhkw1_paper_genotype import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            sv_root = tmp_path / "maizego"
            sv_root.mkdir()
            marker = "chr1_30450000_30458900_insertion_8900"
            matrix_path = sv_root / "SV.386014.hmp.txt"
            matrix_path.write_text(
                "rs#\talleles\tchrom\tpos\tstrand\tassembly#\tcenter\tprotLSID\tassayLSID\tpanelLSID\tQCcode\tS1\tS2\tS3\n"
                f"{marker}\tA/T\t1\t30454500\t+\tNA\tNA\tNA\tNA\tNA\tNA\tAA\tNN\tTT\n",
                encoding="utf-8",
            )
            phenotype_path = tmp_path / "phenotypes.csv"
            phenotype_path.write_text(
                "<Trait>,100grainweight\n"
                "S1,25.1\n"
                "S2,29.3\n"
                "S3,-999\n",
                encoding="utf-8",
            )
            output_root = tmp_path / "db"
            candidate_output = tmp_path / "candidates.tsv"
            marker_output = tmp_path / "marker.tsv"

            stdout = StringIO()
            with redirect_stdout(stdout):
                rc = prepare_main([
                    "--sv-root", str(sv_root),
                    "--phenotype-table", str(phenotype_path),
                    "--output-root", str(output_root),
                    "--candidate-output", str(candidate_output),
                    "--marker-output", str(marker_output),
                ])

            self.assertEqual(rc, 0)
            gene_info = json.loads((output_root / "qHKW1" / "gene_info.json").read_text(encoding="utf-8"))
            self.assertEqual(gene_info["source"], "maizego_paper_marker")
            self.assertEqual(gene_info["source_marker"], marker)
            self.assertTrue(marker_output.exists())
            self.assertTrue(candidate_output.exists())

    def test_wheat_q7b_ph_prepare_builds_figure3g_database(self):
        from openpyxl import Workbook
        from prepare_wheat2024_q7b_ph_figure3g import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_path = tmp_path / "figure3g.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Figure_3_NIL_Q7B-PH_W141_field_"
            ws.append([
                "site", "year", "siteyear", "Accession", "allele",
                "Replicate", "PH_M_cm", "GY_Calc_tha", "GW_M_g1000grn",
            ])
            ws.append(["JIC", "H15", "JIC_H15", "WL0019", "W", 1, 103, 8.34, 48.55])
            ws.append(["JIC", "H15", "JIC_H15", "WL0019", "W", 2, 101, 8.10, 49.00])
            ws.append(["JIC", "H15", "JIC_H15", "WL0025", "P", 1, 97, 8.52, 50.19])
            ws.append(["JIC", "H15", "JIC_H15", "Paragon", "P", 1, 74, 7.90, 46.00])
            ws.append(["JIC", "H15", "JIC_H15", "W10074", "Par", 1, 98, 8.10, None])
            wb.save(source_path)

            output_root = tmp_path / "db"
            marker_output = tmp_path / "q7b_marker.tsv"
            phenotype_output = tmp_path / "q7b_pheno.tsv"

            stdout = StringIO()
            with redirect_stdout(stdout):
                rc = prepare_main([
                    "--source-xlsx", str(source_path),
                    "--output-root", str(output_root),
                    "--marker-output", str(marker_output),
                    "--phenotype-output", str(phenotype_output),
                ])

            self.assertEqual(rc, 0)
            db_dir = output_root / "Q7B-HT"
            gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
            self.assertEqual(gene_info["source"], "wwwg2b_q7b_ph_figure3g")
            self.assertEqual(gene_info["source_file"], str(source_path))
            self.assertEqual(gene_info["source_marker"], "Q7B-PH_allele")
            self.assertLess(gene_info["start"], gene_info["end"])
            self.assertTrue(marker_output.exists())
            self.assertTrue(phenotype_output.exists())
            hap_samples = (db_dir / "haplotype_samples.csv").read_text(encoding="utf-8")
            phenotype_data = (db_dir / "phenotype_data.csv").read_text(encoding="utf-8")
            self.assertIn("Q7B-PH_allele", (db_dir / "variant_info.csv").read_text(encoding="utf-8"))
            self.assertIn("WL0019__plot0001", hap_samples)
            self.assertIn("WL0019__plot0002", hap_samples)
            self.assertIn("WL0019__plot0001", phenotype_data)

    def test_wheat_tagw2_ad_prepare_builds_vcf_region_databases(self):
        from openpyxl import Workbook
        from prepare_wheat2024_tagw2_ad import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            vcf_a = tmp_path / "chr6A.vcf.gz"
            vcf_d = tmp_path / "chr6D.vcf.gz"

            vcf_header = (
                "##fileformat=VCFv4.2\n"
                "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\t"
                "WATDE0001\tWATDE0002\tWATDE0003\tModernA\n"
            )
            with gzip.open(vcf_a, "wt", encoding="utf-8", newline="") as f:
                f.write(vcf_header)
                f.write("6A\t237732100\toutside\tA\tG\t.\tPASS\t.\tGT\t0/0\t0/0\t0/0\t0/0\n")
                f.write("6A\t237734700\tvarA1\tA\tG\t.\tPASS\t.\tGT\t0/0\t1/1\t0/1\t0/0\n")
                f.write("6A\t237735000\tvarA2\tC\tT\t.\tPASS\t.\tGT\t1/1\t1/1\t0/0\t0/0\n")
                f.write("6A\t237761000\tafter\tC\tT\t.\tPASS\t.\tGT\t0/0\t0/0\t0/0\t0/0\n")
            with gzip.open(vcf_d, "wt", encoding="utf-8", newline="") as f:
                f.write(vcf_header.replace("6A", "6D"))
                f.write("6D\t175712300\tvarD1\tG\tA\t.\tPASS\t.\tGT\t0/0\t1/1\t1/1\t0/0\n")
                f.write("6D\t175721000\tvarD2\tT\tC\t.\tPASS\t.\tGT\t1/1\t1/1\t0/0\t0/0\n")

            phenotype_xlsx = tmp_path / "watkins.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "WISP_Watkins_JIC_CFLN10"
            ws.append(["StoreCode", "GW_M_g1000grn-CFLN10"])
            ws.append(["WATDE0001", 40.0])
            ws.append(["WATDE0002", 44.0])
            ws.append(["WATDE0003", 42.0])
            ws2 = wb.create_sheet("DFW_Watkins_JI_CFLN14")
            ws2.append(["StoreCode", "GW_M_g1000grn-CFLN14"])
            ws2.append(["WATDE0001", 41.0])
            ws2.append(["WATDE0002", 45.0])
            ws2.append(["WATDE0003", None])
            wb.save(phenotype_xlsx)

            output_root = tmp_path / "db"
            intermediate_root = tmp_path / "external"

            stdout = StringIO()
            with redirect_stdout(stdout):
                rc = prepare_main([
                    "--vcf-a", str(vcf_a),
                    "--vcf-d", str(vcf_d),
                    "--phenotype-xlsx", str(phenotype_xlsx),
                    "--output-root", str(output_root),
                    "--intermediate-root", str(intermediate_root),
                    "--min-haplotype-count", "1",
                ])

            self.assertEqual(rc, 0)
            for target_id in ["TaGW2-A1", "TaGW2-D1"]:
                db_dir = output_root / target_id
                self.assertTrue((db_dir / "gene_info.json").exists())
                gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
                self.assertEqual(gene_info["source"], "watseq_tagw2_ad_vcf")
                self.assertIn("TGW_mean", (db_dir / "phenotype_data.csv").read_text(encoding="utf-8"))
                self.assertIn("marker_id", (db_dir / "variant_info.csv").read_text(encoding="utf-8"))
                self.assertIn("WATDE0001", (db_dir / "haplotype_samples.csv").read_text(encoding="utf-8"))

    def test_wheat_rht1_prepare_builds_functional_snp_databases(self):
        import pandas as pd
        from openpyxl import Workbook
        from prepare_wheat2024_rht1_functional_snps import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            vcf_path = tmp_path / "rht1.vcf"
            vcf_path.write_text(
                "\n".join([
                    "##fileformat=VCFv4.2",
                    "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tWATDE0001\tWATDE0002\tWATDE0003\tWATDE0004",
                    "chr4B\t30861571\tchr4B_30861571\tC\tT\t.\t.\tANN=T|stop_gained|HIGH|TraesCS4B01G043100|TraesCS4B01G043100|transcript|TraesCS4B01G043100.1|protein_coding|1/1|c.190C>T|p.Gln64*|190/1866|190/1866|64/621||\tGT\t0/0\t1/1\t1/1\t0/0",
                    "chr4D\t18781242\tchr4D_18781242\tG\tT\t.\t.\tANN=T|stop_gained|HIGH|TraesCS4D01G040400|TraesCS4D01G040400|transcript|TraesCS4D01G040400.1|protein_coding|1/1|c.181G>T|p.Glu61*|181/1872|181/1872|61/623||\tGT\t0/0\t1/1\t0/0\t1/1",
                    "",
                ]),
                encoding="utf-8",
            )

            phenotype_xlsx = tmp_path / "phenotypes.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "WGIN_Watkins_JIC_CFLN06"
            ws.append(["StoreCode", "PH_M_cm-CFLN06", "GrwHabit_E_sw-CFLN06"])
            ws.append(["WATDE0001", 120, "w"])
            ws.append(["WATDE0002", 80, "s"])
            ws.append(["WATDE0003", 85, "s"])
            ws.append(["WATDE0004", 78, "w"])
            wb.save(phenotype_xlsx)

            output_root = tmp_path / "db"
            intermediate_root = tmp_path / "intermediate"
            rc = prepare_main([
                "--vcf", str(vcf_path),
                "--phenotype-xlsx", str(phenotype_xlsx),
                "--output-root", str(output_root),
                "--intermediate-root", str(intermediate_root),
                "--target", "Rht-B1b",
                "--target", "Rht-D1b",
                "--min-haplotype-count", "1",
            ])

            self.assertEqual(rc, 0)
            for target_id, marker_id in [
                ("Rht-B1b", "chr4B_30861571"),
                ("Rht-D1b", "chr4D_18781242"),
            ]:
                db_dir = output_root / target_id
                self.assertTrue((db_dir / "gene_info.json").exists())
                gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
                self.assertEqual(gene_info["source"], "wheatomics_remote_rht1_functional_snp_vcf")
                self.assertIn("PlantHeight_CFLN06", (db_dir / "phenotype_data.csv").read_text(encoding="utf-8"))
                variant_info = pd.read_csv(db_dir / "variant_info.csv")
                self.assertEqual(variant_info.loc[0, "marker_id"], marker_id)
                self.assertEqual(variant_info.loc[0, "annotation"], "stop_gained")
                self.assertIn("T", (db_dir / "haplotype_data.csv").read_text(encoding="utf-8"))

    def test_wheat_rht1_prepare_continues_when_one_functional_snp_is_not_segregating(self):
        from openpyxl import Workbook
        from prepare_wheat2024_rht1_functional_snps import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            vcf_path = tmp_path / "rht1.vcf"
            vcf_path.write_text(
                "\n".join([
                    "##fileformat=VCFv4.2",
                    "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tWATDE0001\tWATDE0002\tWATDE0003\tWATDE0004",
                    "chr4B\t30861571\tchr4B_30861571\tC\tT\t.\t.\tANN=T|stop_gained|HIGH|TraesCS4B01G043100|TraesCS4B01G043100|transcript|TraesCS4B01G043100.1|protein_coding|1/1|c.190C>T|p.Gln64*|190/1866|190/1866|64/621||\tGT\t0/0\t0/0\t0/0\t0/0",
                    "chr4D\t18781242\tchr4D_18781242\tG\tT\t.\t.\tANN=T|stop_gained|HIGH|TraesCS4D01G040400|TraesCS4D01G040400|transcript|TraesCS4D01G040400.1|protein_coding|1/1|c.181G>T|p.Glu61*|181/1872|181/1872|61/623||\tGT\t0/0\t1/1\t0/0\t1/1",
                    "",
                ]),
                encoding="utf-8",
            )

            phenotype_xlsx = tmp_path / "phenotypes.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "WGIN_Watkins_JIC_CFLN06"
            ws.append(["StoreCode", "PH_M_cm-CFLN06", "GrwHabit_E_sw-CFLN06"])
            for sample, height in [
                ("WATDE0001", 120),
                ("WATDE0002", 80),
                ("WATDE0003", 118),
                ("WATDE0004", 78),
            ]:
                ws.append([sample, height, "s"])
            wb.save(phenotype_xlsx)

            output_root = tmp_path / "db"
            rc = prepare_main([
                "--vcf", str(vcf_path),
                "--phenotype-xlsx", str(phenotype_xlsx),
                "--output-root", str(output_root),
                "--intermediate-root", str(tmp_path / "intermediate"),
                "--target", "Rht-B1b",
                "--target", "Rht-D1b",
                "--min-haplotype-count", "1",
            ])

            self.assertEqual(rc, 0)
            self.assertFalse((output_root / "Rht-B1b" / "gene_info.json").exists())
            self.assertTrue((output_root / "Rht-D1b" / "gene_info.json").exists())

    def test_wheat_rht_zanke2014_prepare_builds_external_marker_database(self):
        import pandas as pd
        from openpyxl import Workbook
        from prepare_wheat2024_rht_zanke2014 import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_path = tmp_path / "zanke_table_s2.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Table_S2"
            ws.append(["Table S2: List of varieties, genotyping data of candidate genes and phenotypic data."])
            ws.append([None, None, None, "Ellis et al., 2002, TAG", None, "Ellis et al., 2002, TAG"])
            ws.append([
                "Variety name", "GW no.", "habit", "Rht-B1", None, "Rht-D1", None,
                "Rht8", None, None, None, None, None, None, "Ppd-D1 wildtype",
                "Ppd-D1a mutant", "09.AND.PH", "09.SEL.PH", "09.WOH.PH",
                "10.AND.PH", "10.JAN.PH", "10.SAU.PH", "10.SEL.PH",
                "10.WOH.PH", "BLUES", "GAT_2012",
            ])
            ws.append([
                None, None, None, "mutant (dwarfed)", "wild type (tall)",
                "4D-mutant (DF-MR2)", "4D-wild type (DF2-WR2)",
            ])
            ws.append(["Tall1", "GW0001", "winter", None, 1, None, 1, None, None, None, None, None, None, None, 1, None, 110, 112, 108, 111, 109, 113, 107, 110, 110, 111])
            ws.append(["B1short", "GW0002", "winter", 1, None, None, 1, None, None, None, None, None, None, None, 1, None, 88, 90, 86, 89, 87, 91, 85, 88, 88, 89])
            ws.append(["D1short", "GW0003", "spring", None, 1, 1, None, None, None, None, None, None, None, None, 1, None, 82, 84, 80, 83, 81, 85, 79, 82, 82, 83])
            ws.append(["DoubleShort", "GW0004", "spring", 1, None, 1, None, None, None, None, None, None, None, None, 1, None, 70, 72, 68, 71, 69, 73, 67, 70, 70, 71])
            ws.append(["Sum", None, None, 2, 2, 2, 2])
            wb.save(source_path)

            output_root = tmp_path / "db"
            intermediate_root = tmp_path / "intermediate"
            rc = prepare_main([
                "--source-workbook", str(source_path),
                "--output-root", str(output_root),
                "--intermediate-root", str(intermediate_root),
                "--min-haplotype-count", "1",
            ])

            self.assertEqual(rc, 0)
            db_dir = output_root / "Rht-Zanke2014"
            gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
            self.assertEqual(gene_info["source"], "zanke2014_rht_candidate_gene_table_s2")
            self.assertEqual(gene_info["marker_panel"], ["Rht-B1", "Rht-D1"])
            variant_info = pd.read_csv(db_dir / "variant_info.csv")
            self.assertEqual(variant_info["marker_id"].tolist(), ["Rht-B1", "Rht-D1"])
            self.assertTrue((variant_info["annotation"] == "functional_marker").all())
            self.assertIn("PlantHeight_BLUE", (db_dir / "phenotype_data.csv").read_text(encoding="utf-8"))
            self.assertIn("B1b|Rht-D1a", (db_dir / "haplotype_data.csv").read_text(encoding="utf-8"))

    def test_wheat_rht_zanke2014_prepare_builds_single_marker_targets(self):
        import pandas as pd
        from openpyxl import Workbook
        from prepare_wheat2024_rht_zanke2014 import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_path = tmp_path / "zanke_table_s2.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Table_S2"
            ws.append(["Table S2: List of varieties, genotyping data of candidate genes and phenotypic data."])
            ws.append([None, None, None, "Ellis et al., 2002, TAG", None, "Ellis et al., 2002, TAG"])
            ws.append([
                "Variety name", "GW no.", "habit", "Rht-B1", None, "Rht-D1", None,
                "Rht8", None, None, None, None, None, None, "Ppd-D1 wildtype",
                "Ppd-D1a mutant", "09.AND.PH", "09.SEL.PH", "09.WOH.PH",
                "10.AND.PH", "10.JAN.PH", "10.SAU.PH", "10.SEL.PH",
                "10.WOH.PH", "BLUES", "GAT_2012",
            ])
            ws.append([
                None, None, None, "mutant (dwarfed)", "wild type (tall)",
                "4D-mutant (DF-MR2)", "4D-wild type (DF2-WR2)",
            ])
            ws.append(["Tall1", "GW0001", "winter", None, 1, None, 1, None, None, None, None, None, None, None, 1, None, 110, 112, 108, 111, 109, 113, 107, 110, 110, 111])
            ws.append(["B1short", "GW0002", "winter", 1, None, None, 1, None, None, None, None, None, None, None, 1, None, 88, 90, 86, 89, 87, 91, 85, 88, 88, 89])
            ws.append(["D1short", "GW0003", "spring", None, 1, 1, None, None, None, None, None, None, None, None, 1, None, 82, 84, 80, 83, 81, 85, 79, 82, 82, 83])
            ws.append(["DoubleShort", "GW0004", "spring", 1, None, 1, None, None, None, None, None, None, None, None, 1, None, 70, 72, 68, 71, 69, 73, 67, 70, 70, 71])
            ws.append(["Sum", None, None, 2, 2, 2, 2])
            wb.save(source_path)

            output_root = tmp_path / "db"
            intermediate_root = tmp_path / "intermediate"
            rc = prepare_main([
                "--source-workbook", str(source_path),
                "--output-root", str(output_root),
                "--intermediate-root", str(intermediate_root),
                "--min-haplotype-count", "1",
                "--single-marker-targets",
            ])

            self.assertEqual(rc, 0)
            db_dir = output_root / "Rht-D1-Zanke2014"
            gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
            self.assertEqual(gene_info["gene_symbol"], "Rht-D1")
            self.assertEqual(gene_info["marker_panel"], ["Rht-D1"])
            variant_info = pd.read_csv(db_dir / "variant_info.csv")
            self.assertEqual(variant_info["marker_id"].tolist(), ["Rht-D1"])
            self.assertEqual(variant_info["ref"].tolist(), ["Rht-D1a"])
            self.assertEqual(variant_info["alt"].tolist(), ["D1b"])
            hap_text = (db_dir / "haplotype_data.csv").read_text(encoding="utf-8")
            self.assertIn("D1b", hap_text)
            self.assertIn("Rht-D1a", hap_text)
            self.assertNotIn("|", hap_text)

    def test_wheat_vrn_kiss2014_prepare_builds_diagnostic_marker_database(self):
        from openpyxl import Workbook
        from prepare_wheat2024_vrn_kiss2014 import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_path = tmp_path / "kiss2014.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "stepwise_reg2011_2012a"
            ws.append([
                "Pedigre", "VRN-A1", "VRN-B1", "VRN-D1", "PPD-D1", "PPDB1",
                "DEV49_2011", "DEV49_2012",
            ])
            ws.append(["WINTER-1", 0, 0, 0, 0, 0, 214, 216])
            ws.append(["VRND-SPRING", 0, 0, 1, 0, 0, 205, 207])
            ws.append(["VRNA-SPRING", 1, 0, 0, 0, 0, 212, 211])
            ws.append(["VRNB-SPRING", 0, 1, 0, 0, 0, 213, 212])
            ws2 = wb.create_sheet("stepwise_reg2011_2012b")
            ws2.append([
                "Pedigre", "VRN-A1", "VRN-B1", "VRN-D1", "PPD-D1", "PPDB1",
                "DEV59_2011", "DEV59_2012",
            ])
            ws2.append(["WINTER-1", 0, 0, 0, 0, 0, 222, 224])
            ws2.append(["VRND-SPRING", 0, 0, 1, 0, 0, 214, 216])
            ws2.append(["VRNA-SPRING", 1, 0, 0, 0, 0, 220, 221])
            ws2.append(["VRNB-SPRING", 0, 1, 0, 0, 0, 221, 220])
            wb.save(source_path)

            output_root = tmp_path / "db"
            intermediate_root = tmp_path / "intermediate"
            rc = prepare_main([
                "--source-workbook", str(source_path),
                "--output-root", str(output_root),
                "--intermediate-root", str(intermediate_root),
                "--min-haplotype-count", "1",
            ])

            self.assertEqual(rc, 0)
            db_dir = output_root / "VRN-Kiss2014"
            gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
            self.assertEqual(gene_info["source"], "kiss2014_vrn_diagnostic_markers")
            self.assertEqual(gene_info["source_workbook"], str(source_path))
            self.assertEqual(gene_info["marker_panel"], ["VRN-A1", "VRN-B1", "VRN-D1"])
            self.assertTrue((intermediate_root / "VRN-Kiss2014" / "marker_matrix.tsv").exists())
            self.assertTrue((intermediate_root / "VRN-Kiss2014" / "phenotype.tsv").exists())

            variant_info = (db_dir / "variant_info.csv").read_text(encoding="utf-8")
            self.assertIn("VRN-D1", variant_info)
            self.assertIn("diagnostic_marker", variant_info)
            phenotype_data = (db_dir / "phenotype_data.csv").read_text(encoding="utf-8")
            self.assertIn("DEV49_mean", phenotype_data)
            self.assertIn("DEV59_mean", phenotype_data)
            hap_text = (db_dir / "haplotype_data.csv").read_text(encoding="utf-8")
            self.assertIn("0|0|1", hap_text)

    def test_wheat_vrn_kiss2014_prepare_builds_single_marker_targets(self):
        from openpyxl import Workbook
        from prepare_wheat2024_vrn_kiss2014 import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_path = tmp_path / "kiss2014.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "stepwise_reg2011_2012a"
            ws.append([
                "Pedigre", "VRN-A1", "VRN-B1", "VRN-D1", "PPD-D1", "PPDB1",
                "DEV49_2011", "DEV49_2012",
            ])
            ws.append(["WINTER-1", 0, 0, 0, 0, 0, 214, 216])
            ws.append(["VRND-SPRING", 0, 0, 1, 0, 0, 205, 207])
            ws.append(["VRNA-SPRING", 1, 0, 0, 0, 0, 212, 211])
            ws.append(["VRNB-SPRING", 0, 1, 0, 0, 0, 213, 212])
            ws2 = wb.create_sheet("stepwise_reg2011_2012b")
            ws2.append([
                "Pedigre", "VRN-A1", "VRN-B1", "VRN-D1", "PPD-D1", "PPDB1",
                "DEV59_2011", "DEV59_2012",
            ])
            ws2.append(["WINTER-1", 0, 0, 0, 0, 0, 222, 224])
            ws2.append(["VRND-SPRING", 0, 0, 1, 0, 0, 214, 216])
            ws2.append(["VRNA-SPRING", 1, 0, 0, 0, 0, 220, 221])
            ws2.append(["VRNB-SPRING", 0, 1, 0, 0, 0, 221, 220])
            wb.save(source_path)

            output_root = tmp_path / "db"
            intermediate_root = tmp_path / "intermediate"
            rc = prepare_main([
                "--source-workbook", str(source_path),
                "--output-root", str(output_root),
                "--intermediate-root", str(intermediate_root),
                "--min-haplotype-count", "1",
                "--single-marker-targets",
            ])

            self.assertEqual(rc, 0)
            for marker in ["VRN-A1", "VRN-B1", "VRN-D1"]:
                db_dir = output_root / f"{marker}-Kiss2014"
                self.assertTrue((db_dir / "gene_info.json").exists())
                gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
                self.assertEqual(gene_info["gene_symbol"], marker)
                self.assertEqual(gene_info["marker_panel"], [marker])
                variant_info = (db_dir / "variant_info.csv").read_text(encoding="utf-8")
                self.assertIn(marker, variant_info)
                self.assertNotIn("|", (db_dir / "haplotype_data.csv").read_text(encoding="utf-8"))

    def test_precomputed_single_marker_haplotypes_load_as_strings(self):
        import pandas as pd
        from haplotype_phenotype_analysis import HaplotypePhenotypeAnalyzer

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db" / "SingleMarker"
            db_dir.mkdir(parents=True)
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,maf,missing_rate,len_diff,is_sv,annotation,marker_id\n"
                "1,0,1,0.25,0,0,False,diagnostic_marker,Marker1\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\n"
                "0,2,Hap1,0\n"
                "1,2,Hap2,1\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\n"
                "S1,0,Hap1\n"
                "S2,1,Hap2\n",
                encoding="utf-8",
            )
            (db_dir / "phenotype_data.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name,Trait\n"
                "S1,0,Hap1,10\n"
                "S2,1,Hap2,20\n",
                encoding="utf-8",
            )
            (db_dir / "gene_info.json").write_text(
                json.dumps({"gene_id": "SingleMarker", "chrom": "diagnostic_marker_panel", "start": 1, "end": 1}),
                encoding="utf-8",
            )

            analyzer = HaplotypePhenotypeAnalyzer(
                vcf_file=str(tmp_path / "missing.vcf"),
                phenotype_file=str(db_dir / "phenotype_data.csv"),
                output_dir=str(tmp_path / "out"),
            )
            analyzer.analyze_gene(
                chrom="diagnostic_marker_panel",
                start=1,
                end=1,
                gene_id="SingleMarker",
                phenotype_cols=["Trait"],
                database_dir=str(tmp_path / "db"),
            )

            loaded = pd.read_csv(db_dir / "haplotype_samples.csv")
            self.assertIn(loaded["Haplotype_Seq"].dtype.kind, "iu")
            self.assertTrue((tmp_path / "out" / "SingleMarker.html").exists())
            scores = json.loads((tmp_path / "out" / "haplotype_scores.json").read_text(encoding="utf-8"))
            self.assertIn("Trait", scores)

    def test_precomputed_database_preserves_variant_info_row_order_for_positions(self):
        from haplotype_phenotype_analysis import HaplotypePhenotypeAnalyzer

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_dir = tmp_path / "db" / "OrderGene"
            db_dir.mkdir(parents=True)
            (db_dir / "variant_info.csv").write_text(
                "position,ref,alt,maf,missing_rate,len_diff,is_sv,annotation,marker_id\n"
                "100,A,G,0.5,0,0,False,base_snp,M100\n"
                "300,C,T,0.5,0,0,False,base_snp,M300\n"
                "200,DEL_837,INS_837,0.5,0,837,True,diagnostic_marker,M200\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_data.csv").write_text(
                "Haplotype_Seq,Count,Hap_Name,Alleles\n"
                "A|C|DEL_837,2,Hap1,A|C|DEL_837\n"
                "A|C|INS_837,2,Hap2,A|C|INS_837\n",
                encoding="utf-8",
            )
            (db_dir / "haplotype_samples.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name\n"
                "S1,A|C|DEL_837,Hap1\n"
                "S2,A|C|DEL_837,Hap1\n"
                "S3,A|C|INS_837,Hap2\n"
                "S4,A|C|INS_837,Hap2\n",
                encoding="utf-8",
            )
            (db_dir / "phenotype_data.csv").write_text(
                "SampleID,Haplotype_Seq,Hap_Name,Trait\n"
                "S1,A|C|DEL_837,Hap1,0\n"
                "S2,A|C|DEL_837,Hap1,0\n"
                "S3,A|C|INS_837,Hap2,1\n"
                "S4,A|C|INS_837,Hap2,1\n",
                encoding="utf-8",
            )
            (db_dir / "gene_info.json").write_text(
                json.dumps({
                    "gene_id": "OrderGene",
                    "chrom": "chrT",
                    "start": 100,
                    "end": 300,
                    "gene_start": 100,
                    "gene_end": 300,
                    "promoter_start": 1,
                    "promoter_end": 99,
                    "promoter_actual_length": 99,
                }),
                encoding="utf-8",
            )

            analyzer = HaplotypePhenotypeAnalyzer(
                vcf_file=str(tmp_path / "missing.vcf"),
                phenotype_file=str(db_dir / "phenotype_data.csv"),
                output_dir=str(tmp_path / "out"),
            )
            analyzer.analyze_gene(
                chrom="chrT",
                start=100,
                end=300,
                gene_id="OrderGene",
                phenotype_cols=["Trait"],
                database_dir=str(tmp_path / "db"),
            )

            self.assertEqual([100, 300, 200], analyzer.positions)

    def test_wheat_tagw2_b1_prepare_builds_remote_snp_database(self):
        import pandas as pd
        from openpyxl import Workbook
        from prepare_wheat2024_tagw2_b1_remote_snp import main as prepare_main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            vcf_path = tmp_path / "tagw2_b1.vcf"
            vcf_path.write_text(
                "\n".join([
                    "##fileformat=VCFv4.2",
                    "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tWATDE0001\tWATDE0002\tWATDE0003\tWATDE0004",
                    "chr6B\t291759689\tchr6B_291759689\tC\tA\t.\tPASS\t.\tGT\t1/1\t0/0\t1/1\t0/0",
                    "chr6B\t291760677\tchr6B_291760677\tG\tA\t.\tPASS\t.\tGT\t0/0\t1/1\t0/0\t1/1",
                    "chr6B\t291761315\tchr6B_291761315\tT\tC\t.\tPASS\t.\tGT\t1/1\t0/0\t1/1\t0/0",
                    "chr6B\t291762000\tchr6B_291762000\tG\tT\t.\tPASS\t.\tGT\t1/1\t0/0\t1/1\t0/0",
                    "",
                ]),
                encoding="utf-8",
            )

            phenotype_xlsx = tmp_path / "watkins.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "WISP_Watkins_JIC_CFLN10"
            ws.append(["StoreCode", "GW_M_g1000grn-CFLN10"])
            ws.append(["WATDE0001", 48.0])
            ws.append(["WATDE0002", 40.0])
            ws.append(["WATDE0003", 47.0])
            ws.append(["WATDE0004", 41.0])
            ws2 = wb.create_sheet("DFW_Watkins_JI_CFLN14")
            ws2.append(["StoreCode", "GW_M_g1000grn-CFLN14"])
            ws2.append(["WATDE0001", 49.0])
            ws2.append(["WATDE0002", 39.0])
            ws2.append(["WATDE0003", 46.0])
            ws2.append(["WATDE0004", 42.0])
            wb.save(phenotype_xlsx)

            output_root = tmp_path / "db"
            intermediate_root = tmp_path / "intermediate"
            rc = prepare_main([
                "--vcf", str(vcf_path),
                "--phenotype-xlsx", str(phenotype_xlsx),
                "--output-root", str(output_root),
                "--intermediate-root", str(intermediate_root),
                "--min-haplotype-count", "1",
            ])

            self.assertEqual(rc, 0)
            db_dir = output_root / "TaGW2-B1-remoteSNP"
            gene_info = json.loads((db_dir / "gene_info.json").read_text(encoding="utf-8"))
            self.assertEqual(gene_info["source"], "wwwg2b_tagw2_b1_full_region_snp_vcf")
            self.assertEqual(gene_info["atg_position"], 291761397)
            variant_info = pd.read_csv(db_dir / "variant_info.csv")
            self.assertEqual(len(variant_info), 4)
            for marker in ["chr6B_291759689", "chr6B_291760677", "chr6B_291761315"]:
                self.assertIn(marker, variant_info["marker_id"].tolist())
            self.assertIn("chr6B_291762000", variant_info["marker_id"].tolist())
            self.assertEqual(
                variant_info.set_index("marker_id").loc["chr6B_291762000", "annotation"],
                "intron",
            )
            hap_text = (db_dir / "haplotype_data.csv").read_text(encoding="utf-8")
            self.assertIn("A|G|C|T", hap_text)

    def test_tagw2_hap8_maps_published_promoter_offsets_to_refseq_positions(self):
        from analyze_tagw2_hap8_functional_groups import map_promoter_offset_to_position

        gene_start = 237_734_651
        cds_start_offset_1based = 185

        self.assertEqual(
            map_promoter_offset_to_position(gene_start, cds_start_offset_1based, -739),
            237_734_096,
        )
        self.assertEqual(
            map_promoter_offset_to_position(gene_start, cds_start_offset_1based, -593),
            237_734_242,
        )

    def test_tagw2_hap8_classifies_literature_and_background_patterns(self):
        import pandas as pd
        from analyze_tagw2_hap8_functional_groups import classify_literature_background_groups

        df = pd.DataFrame({
            "SampleID": ["S1", "S2", "S3", "S4"],
            "Hap_Name": ["Hap8", "Hap7", "Hap1", "Hap15"],
            "lit1": ["G", "G", "A", "G"],
            "lit2": ["A", "A", "G", "A"],
            "bg1": ["C", "T", "T", "C"],
            "bg2": ["T", "T", "T", "G"],
        })

        classified = classify_literature_background_groups(
            df,
            literature_alleles={"lit1": "G", "lit2": "A"},
            background_alleles={"bg1": "C", "bg2": "T"},
            full_haplotype_name="Hap8",
        )

        self.assertEqual(classified["PublishedPair"].tolist(), [True, True, False, True])
        self.assertEqual(classified["Hap8Background"].tolist(), [True, False, False, False])
        self.assertEqual(classified["FullHap8"].tolist(), [True, False, False, False])
        self.assertEqual(
            classified["Hap8SplitGroup"].tolist(),
            ["Published+Background", "PublishedOnly", "Neither", "PublishedOnly"],
        )

    def test_tagw2_hap8_binary_trait_summary_reports_effect_and_pvalue(self):
        import pandas as pd
        from analyze_tagw2_hap8_functional_groups import summarize_binary_trait

        df = pd.DataFrame({
            "SampleID": ["S1", "S2", "S3", "S4", "S5"],
            "group": [True, True, False, False, False],
            "TGW_mean": [30.0, 32.0, 40.0, 42.0, 44.0],
        })

        summary = summarize_binary_trait(df, "group", "TGW_mean", "case", "control")

        self.assertEqual(summary["case_label"], "case")
        self.assertEqual(summary["control_label"], "control")
        self.assertEqual(summary["case_n"], 2)
        self.assertEqual(summary["control_n"], 3)
        self.assertLess(summary["effect_case_minus_control"], 0)
        self.assertTrue(0 <= summary["welch_pvalue"] <= 1)


if __name__ == "__main__":
    unittest.main()
